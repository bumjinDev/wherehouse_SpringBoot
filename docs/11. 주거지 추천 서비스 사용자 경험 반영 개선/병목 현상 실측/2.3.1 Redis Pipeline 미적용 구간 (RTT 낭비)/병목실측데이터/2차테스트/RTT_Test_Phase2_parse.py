#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Redis Pipeline 병목 측정 로그 파서 - Phase 2 (Pipeline 적용 버전)
================================================================
Phase 1 (Sequential)과 Phase 2 (Pipeline) 로그를 모두 파싱하여
Before/After 비교 분석을 지원한다.

지원 로그 패턴:
- [Metrics-Sequential]: Phase 1 순차 호출 측정
- [Metrics-Pipeline]: Phase 2 Pipeline 호출 측정
- [Metrics-LoopSummary]: 루프 전체 요약

사용법:
    python RTT_Test_Phase2_parse.py <log_file_path> [options]
    
예시:
    python RTT_Test_Phase2_parse.py ./phase2_log.txt --output xlsx
    python RTT_Test_Phase2_parse.py ./phase2_log.txt --compare phase1_result.xlsx
"""

import re
import sys
import json
import argparse
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, asdict
from typing import List, Dict, Optional, Tuple
from statistics import mean, stdev, median
from collections import defaultdict


# =============================================================================
# 데이터 클래스 정의
# =============================================================================

@dataclass
class SequentialMetric:
    """[Metrics-Sequential] 로그 파싱 결과 (Phase 1)"""
    timestamp: Optional[str]
    thread_name: Optional[str]
    district: str
    commands: int
    method_time_ms: float
    total_cmd_latency_ms: float
    io_ratio_pct: float
    non_io_time_ms: float
    cmd1_latency_ms: float
    cmd1_count: int
    cmd2_latency_ms: float
    cmd2_count: int
    cmd3_latency_ms: float
    cmd3_count: int
    status: str
    intersection_count: Optional[int] = None


@dataclass
class PipelineMetric:
    """[Metrics-Pipeline] 로그 파싱 결과 (Phase 2)"""
    timestamp: Optional[str]
    thread_name: Optional[str]
    district: str
    commands: str  # "3(batched)"
    method_time_ms: float
    pipeline_latency_ms: float
    io_ratio_pct: float
    non_io_time_ms: float
    result1_count: int
    result2_count: int
    result3_count: int
    status: str
    intersection_count: Optional[int] = None


@dataclass
class LoopSummaryMetric:
    """[Metrics-LoopSummary] 로그 파싱 결과"""
    timestamp: Optional[str]
    thread_name: Optional[str]
    mode: str  # SEQUENTIAL 또는 PIPELINE
    total_districts: int
    success_districts: int
    empty_districts: int
    total_properties: int
    loop_time_ms: float
    avg_per_district_ms: float


@dataclass
class AnalysisSummary:
    """분석 요약 결과"""
    total_records: int
    avg_method_time_ms: float
    avg_cmd_latency_ms: float
    avg_io_ratio_pct: float
    min_method_time_ms: float
    max_method_time_ms: float
    std_method_time_ms: float
    success_rate_pct: float
    early_termination_rate_pct: float
    mode: str  # "SEQUENTIAL" 또는 "PIPELINE"


# =============================================================================
# 정규식 패턴 정의
# =============================================================================

TIMESTAMP_PATTERNS = [
    r'(\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}[.,]?\d*)',
    r'(\d{2}-\w{3}-\d{4} \d{2}:\d{2}:\d{2})',
    r'(\d{2}:\d{2}:\d{2}[.,]\d+)',
]

THREAD_PATTERNS = [
    r'\[([^\]]*exec[^\]]*)\]',
    r'\[([^\]]*thread[^\]]*)\]',
    r'\[(scheduling-\d+)\]',
    r'\[(main)\]',
    r'--- \[([^\]]+)\]',
    r'\] \[([^\]]+)\] [a-z]',
    r'\[([a-zA-Z]+-[a-zA-Z]+-\d+-[a-zA-Z]+-\d+)\]',
]

# [Metrics-Sequential] 로그 패턴 (Phase 1)
SEQUENTIAL_PATTERN = re.compile(
    r'\[Metrics-Sequential\]\s*'
    r'district=([^,]+),\s*'
    r'commands=(-?\d+),\s*'
    r'methodTime=([\d.]+)\s*ms,\s*'
    r'totalCmdLatency=([\d.]+)\s*ms,\s*'
    r'ioRatio=([\d.]+)\s*%,\s*'
    r'nonIoTime=([\d.]+)\s*ms,\s*'
    r'cmd1=([\d.]+)\s*ms\s*\((\d+)건\),\s*'
    r'cmd2=([\d.]+)\s*ms\s*\((\d+)건\),\s*'
    r'cmd3=([\d.]+)\s*ms\s*\((\d+)건\),\s*'
    r'status=(.+?)(?:\s*$|\s*\n)',
    re.UNICODE
)

# [Metrics-Pipeline] 로그 패턴 (Phase 2)
PIPELINE_PATTERN = re.compile(
    r'\[Metrics-Pipeline\]\s*'
    r'district=([^,]+),\s*'
    r'commands=([^,]+),\s*'
    r'methodTime=([\d.]+)\s*ms,\s*'
    r'pipelineLatency=([\d.]+)\s*ms,\s*'
    r'ioRatio=([\d.]+)\s*%,\s*'
    r'nonIoTime=([\d.]+)\s*ms,\s*'
    r'result1=(\d+)\s*건,\s*'
    r'result2=(\d+)\s*건,\s*'
    r'result3=(\d+)\s*건,\s*'
    r'status=(.+?)(?:\s*$|\s*\n)',
    re.UNICODE
)

# [Metrics-LoopSummary] 로그 패턴
LOOP_SUMMARY_PATTERN = re.compile(
    r'\[Metrics-LoopSummary\]\s*'
    r'mode=(\w+),\s*'
    r'totalDistricts=(\d+),\s*'
    r'successDistricts=(\d+),\s*'
    r'emptyDistricts=(\d+),\s*'
    r'totalProperties=(\d+),\s*'
    r'loopTime=([\d.]+)\s*ms,\s*'
    r'avgPerDistrict=([\d.]+)\s*ms',
    re.UNICODE
)


# =============================================================================
# 파싱 함수
# =============================================================================

def extract_timestamp(line: str) -> Optional[str]:
    for pattern in TIMESTAMP_PATTERNS:
        match = re.search(pattern, line)
        if match:
            return match.group(1)
    return None


def extract_thread_name(line: str) -> Optional[str]:
    for pattern in THREAD_PATTERNS:
        match = re.search(pattern, line)
        if match:
            return match.group(1)
    return None


def parse_sequential_log(line: str) -> Optional[SequentialMetric]:
    """[Metrics-Sequential] 로그 파싱 (Phase 1)"""
    match = SEQUENTIAL_PATTERN.search(line)
    if not match:
        return None
    
    timestamp = extract_timestamp(line)
    thread_name = extract_thread_name(line)
    status = match.group(13).strip()
    
    intersection_count = None
    if 'intersection=' in status:
        try:
            intersection_count = int(re.search(r'intersection=(\d+)', status).group(1))
        except (AttributeError, ValueError):
            pass
    
    return SequentialMetric(
        timestamp=timestamp,
        thread_name=thread_name,
        district=match.group(1).strip(),
        commands=int(match.group(2)),
        method_time_ms=float(match.group(3)),
        total_cmd_latency_ms=float(match.group(4)),
        io_ratio_pct=float(match.group(5)),
        non_io_time_ms=float(match.group(6)),
        cmd1_latency_ms=float(match.group(7)),
        cmd1_count=int(match.group(8)),
        cmd2_latency_ms=float(match.group(9)),
        cmd2_count=int(match.group(10)),
        cmd3_latency_ms=float(match.group(11)),
        cmd3_count=int(match.group(12)),
        status=status,
        intersection_count=intersection_count
    )


def parse_pipeline_log(line: str) -> Optional[PipelineMetric]:
    """[Metrics-Pipeline] 로그 파싱 (Phase 2)"""
    match = PIPELINE_PATTERN.search(line)
    if not match:
        return None
    
    timestamp = extract_timestamp(line)
    thread_name = extract_thread_name(line)
    status = match.group(10).strip()
    
    intersection_count = None
    if 'intersection=' in status:
        try:
            intersection_count = int(re.search(r'intersection=(\d+)', status).group(1))
        except (AttributeError, ValueError):
            pass
    
    return PipelineMetric(
        timestamp=timestamp,
        thread_name=thread_name,
        district=match.group(1).strip(),
        commands=match.group(2).strip(),
        method_time_ms=float(match.group(3)),
        pipeline_latency_ms=float(match.group(4)),
        io_ratio_pct=float(match.group(5)),
        non_io_time_ms=float(match.group(6)),
        result1_count=int(match.group(7)),
        result2_count=int(match.group(8)),
        result3_count=int(match.group(9)),
        status=status,
        intersection_count=intersection_count
    )


def parse_loop_summary_log(line: str) -> Optional[LoopSummaryMetric]:
    """[Metrics-LoopSummary] 로그 파싱"""
    match = LOOP_SUMMARY_PATTERN.search(line)
    if not match:
        return None
    
    timestamp = extract_timestamp(line)
    thread_name = extract_thread_name(line)
    
    return LoopSummaryMetric(
        timestamp=timestamp,
        thread_name=thread_name,
        mode=match.group(1),
        total_districts=int(match.group(2)),
        success_districts=int(match.group(3)),
        empty_districts=int(match.group(4)),
        total_properties=int(match.group(5)),
        loop_time_ms=float(match.group(6)),
        avg_per_district_ms=float(match.group(7))
    )


def parse_log_file(filepath: str) -> Tuple[List[SequentialMetric], List[PipelineMetric], List[LoopSummaryMetric]]:
    """로그 파일 전체 파싱 (Phase 1 + Phase 2)"""
    sequential_metrics = []
    pipeline_metrics = []
    loop_summary_metrics = []
    
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            # Sequential 패턴 체크 (Phase 1)
            if '[Metrics-Sequential]' in line:
                metric = parse_sequential_log(line)
                if metric:
                    sequential_metrics.append(metric)
            
            # Pipeline 패턴 체크 (Phase 2)
            elif '[Metrics-Pipeline]' in line:
                metric = parse_pipeline_log(line)
                if metric:
                    pipeline_metrics.append(metric)
            
            # LoopSummary 패턴 체크
            elif '[Metrics-LoopSummary]' in line:
                metric = parse_loop_summary_log(line)
                if metric:
                    loop_summary_metrics.append(metric)
    
    return sequential_metrics, pipeline_metrics, loop_summary_metrics


# =============================================================================
# 분석 함수
# =============================================================================

def analyze_sequential_metrics(metrics: List[SequentialMetric]) -> Optional[AnalysisSummary]:
    """Sequential 메트릭 분석 (Phase 1)"""
    if not metrics:
        return None
    
    method_times = [m.method_time_ms for m in metrics]
    cmd_latencies = [m.total_cmd_latency_ms for m in metrics]
    io_ratios = [m.io_ratio_pct for m in metrics]
    
    success_count = sum(1 for m in metrics if 'SUCCESS' in m.status)
    early_term_count = sum(1 for m in metrics if 'EARLY_RETURN' in m.status)
    
    return AnalysisSummary(
        total_records=len(metrics),
        avg_method_time_ms=mean(method_times),
        avg_cmd_latency_ms=mean(cmd_latencies),
        avg_io_ratio_pct=mean(io_ratios),
        min_method_time_ms=min(method_times),
        max_method_time_ms=max(method_times),
        std_method_time_ms=stdev(method_times) if len(method_times) > 1 else 0,
        success_rate_pct=(success_count / len(metrics)) * 100,
        early_termination_rate_pct=(early_term_count / len(metrics)) * 100,
        mode="SEQUENTIAL"
    )


def analyze_pipeline_metrics(metrics: List[PipelineMetric]) -> Optional[AnalysisSummary]:
    """Pipeline 메트릭 분석 (Phase 2)"""
    if not metrics:
        return None
    
    method_times = [m.method_time_ms for m in metrics]
    pipeline_latencies = [m.pipeline_latency_ms for m in metrics]
    io_ratios = [m.io_ratio_pct for m in metrics]
    
    success_count = sum(1 for m in metrics if 'SUCCESS' in m.status)
    empty_count = sum(1 for m in metrics if 'EMPTY' in m.status)
    
    return AnalysisSummary(
        total_records=len(metrics),
        avg_method_time_ms=mean(method_times),
        avg_cmd_latency_ms=mean(pipeline_latencies),
        avg_io_ratio_pct=mean(io_ratios),
        min_method_time_ms=min(method_times),
        max_method_time_ms=max(method_times),
        std_method_time_ms=stdev(method_times) if len(method_times) > 1 else 0,
        success_rate_pct=(success_count / len(metrics)) * 100,
        early_termination_rate_pct=(empty_count / len(metrics)) * 100,
        mode="PIPELINE"
    )


def analyze_by_district(metrics: List) -> Dict[str, AnalysisSummary]:
    """지역구별 분석 (Sequential/Pipeline 공통)"""
    district_metrics = defaultdict(list)
    
    for m in metrics:
        district_metrics[m.district].append(m)
    
    result = {}
    for district, district_data in district_metrics.items():
        if isinstance(district_data[0], SequentialMetric):
            result[district] = analyze_sequential_metrics(district_data)
        elif isinstance(district_data[0], PipelineMetric):
            result[district] = analyze_pipeline_metrics(district_data)
    
    return result


def compare_phases(seq_summary: AnalysisSummary, pipe_summary: AnalysisSummary) -> Dict:
    """Phase 1 vs Phase 2 비교 분석"""
    if not seq_summary or not pipe_summary:
        return {}
    
    method_time_improvement = ((seq_summary.avg_method_time_ms - pipe_summary.avg_method_time_ms) 
                               / seq_summary.avg_method_time_ms) * 100
    
    latency_improvement = ((seq_summary.avg_cmd_latency_ms - pipe_summary.avg_cmd_latency_ms) 
                           / seq_summary.avg_cmd_latency_ms) * 100
    
    io_ratio_change = pipe_summary.avg_io_ratio_pct - seq_summary.avg_io_ratio_pct
    
    return {
        "before": {
            "mode": seq_summary.mode,
            "avg_method_time_ms": round(seq_summary.avg_method_time_ms, 4),
            "avg_cmd_latency_ms": round(seq_summary.avg_cmd_latency_ms, 4),
            "avg_io_ratio_pct": round(seq_summary.avg_io_ratio_pct, 2),
            "total_records": seq_summary.total_records
        },
        "after": {
            "mode": pipe_summary.mode,
            "avg_method_time_ms": round(pipe_summary.avg_method_time_ms, 4),
            "avg_cmd_latency_ms": round(pipe_summary.avg_cmd_latency_ms, 4),
            "avg_io_ratio_pct": round(pipe_summary.avg_io_ratio_pct, 2),
            "total_records": pipe_summary.total_records
        },
        "improvement": {
            "method_time_pct": round(method_time_improvement, 2),
            "latency_pct": round(latency_improvement, 2),
            "io_ratio_change_pct": round(io_ratio_change, 2)
        }
    }


# =============================================================================
# 출력 함수
# =============================================================================

def print_pipeline_table(metrics: List[PipelineMetric], limit: int = 50):
    """Pipeline 메트릭 테이블 출력"""
    print("\n" + "=" * 120)
    print("📊 [Metrics-Pipeline] 상세 데이터 (Phase 2)")
    print("=" * 120)
    
    headers = ['Timestamp', 'Thread', 'District', 'MethodTime', 'PipelineLatency', 
               'IO%', 'R1', 'R2', 'R3', 'Status']
    
    col_widths = [23, 25, 10, 12, 15, 8, 5, 5, 5, 25]
    
    header_line = ""
    for i, (h, w) in enumerate(zip(headers, col_widths)):
        header_line += f"{h:<{w}}"
    print(header_line)
    print("-" * 120)
    
    for m in metrics[:limit]:
        ts = m.timestamp[:23] if m.timestamp else "-"
        thread = (m.thread_name[:23] if m.thread_name else "-")
        district = m.district[:8] if len(m.district) > 8 else m.district
        status = m.status[:23] if len(m.status) > 23 else m.status
        
        row = (f"{ts:<23}{thread:<25}{district:<10}"
               f"{m.method_time_ms:>10.4f}ms"
               f"{m.pipeline_latency_ms:>13.4f}ms"
               f"{m.io_ratio_pct:>7.2f}%"
               f"{m.result1_count:>5}"
               f"{m.result2_count:>5}"
               f"{m.result3_count:>5}"
               f"  {status}")
        print(row)
    
    if len(metrics) > limit:
        print(f"\n... 외 {len(metrics) - limit}건 생략")


def print_comparison_summary(comparison: Dict):
    """Phase 1 vs Phase 2 비교 요약 출력"""
    if not comparison:
        print("\n⚠️ 비교할 데이터가 부족합니다.")
        return
    
    print("\n" + "=" * 80)
    print("📈 Phase 1 (Sequential) vs Phase 2 (Pipeline) 비교 분석")
    print("=" * 80)
    
    before = comparison["before"]
    after = comparison["after"]
    improvement = comparison["improvement"]
    
    print(f"\n{'지표':<25} {'Before (Sequential)':>20} {'After (Pipeline)':>20} {'개선율':>15}")
    print("-" * 80)
    print(f"{'평균 Method Time':<25} {before['avg_method_time_ms']:>18.4f}ms {after['avg_method_time_ms']:>18.4f}ms {improvement['method_time_pct']:>13.2f}%")
    print(f"{'평균 Cmd/Pipeline Latency':<25} {before['avg_cmd_latency_ms']:>18.4f}ms {after['avg_cmd_latency_ms']:>18.4f}ms {improvement['latency_pct']:>13.2f}%")
    print(f"{'평균 I/O 비율':<25} {before['avg_io_ratio_pct']:>18.2f}% {after['avg_io_ratio_pct']:>18.2f}% {improvement['io_ratio_change_pct']:>+13.2f}%p")
    print(f"{'총 레코드 수':<25} {before['total_records']:>18}건 {after['total_records']:>18}건")
    
    print("\n" + "-" * 80)
    print("📌 해석:")
    if improvement['latency_pct'] > 0:
        print(f"   • Command Latency가 {improvement['latency_pct']:.2f}% 감소했습니다.")
        print(f"     → 3회 RTT → 1회 RTT 효과로 네트워크 대기 시간이 줄었습니다.")
    if improvement['method_time_pct'] > 0:
        print(f"   • 메서드 총 실행 시간이 {improvement['method_time_pct']:.2f}% 감소했습니다.")
    if improvement['io_ratio_change_pct'] < 0:
        print(f"   • I/O 비율이 {abs(improvement['io_ratio_change_pct']):.2f}%p 감소했습니다.")
        print(f"     → 상대적으로 비-I/O 작업(byte[]→String 변환 등)의 비중이 증가했습니다.")


def print_analysis_summary(summary: AnalysisSummary):
    """분석 요약 출력"""
    if not summary:
        return
    
    print("\n" + "=" * 60)
    print(f"📊 분석 요약 ({summary.mode})")
    print("=" * 60)
    
    print(f"  총 레코드 수        : {summary.total_records}건")
    print(f"  평균 Method Time    : {summary.avg_method_time_ms:.4f} ms")
    print(f"  평균 Cmd Latency    : {summary.avg_cmd_latency_ms:.4f} ms")
    print(f"  평균 I/O 비율       : {summary.avg_io_ratio_pct:.2f} %")
    print(f"  최소 Method Time    : {summary.min_method_time_ms:.4f} ms")
    print(f"  최대 Method Time    : {summary.max_method_time_ms:.4f} ms")
    print(f"  표준편차            : {summary.std_method_time_ms:.4f} ms")
    print(f"  성공률              : {summary.success_rate_pct:.2f} %")
    print(f"  빈 결과율           : {summary.early_termination_rate_pct:.2f} %")


def print_loop_summary_table(metrics: List[LoopSummaryMetric]):
    """LoopSummary 테이블 출력"""
    if not metrics:
        return
    
    print("\n" + "=" * 100)
    print("📊 [Metrics-LoopSummary] 루프 전체 요약")
    print("=" * 100)
    
    headers = ['Timestamp', 'Thread', 'Mode', 'Districts', 'Success', 'Empty', 'Properties', 'LoopTime', 'Avg/District']
    print(f"{'Timestamp':<24}{'Thread':<25}{'Mode':<12}{'Districts':>10}{'Success':>8}{'Empty':>7}{'Props':>8}{'LoopTime':>12}{'Avg':>10}")
    print("-" * 100)
    
    for m in metrics:
        ts = m.timestamp[:23] if m.timestamp else "-"
        thread = (m.thread_name[:23] if m.thread_name else "-")
        print(f"{ts:<24}{thread:<25}{m.mode:<12}{m.total_districts:>10}"
              f"{m.success_districts:>8}{m.empty_districts:>7}{m.total_properties:>8}"
              f"{m.loop_time_ms:>10.4f}ms{m.avg_per_district_ms:>8.4f}ms")


# =============================================================================
# Excel 출력 함수
# =============================================================================

def export_to_xlsx(sequential: List[SequentialMetric],
                   pipeline: List[PipelineMetric],
                   loop_summary: List[LoopSummaryMetric],
                   seq_summary: Optional[AnalysisSummary],
                   pipe_summary: Optional[AnalysisSummary],
                   comparison: Dict,
                   filepath: str):
    """Excel 파일로 내보내기"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ openpyxl 라이브러리가 필요합니다: pip install openpyxl")
        return
    
    wb = Workbook()
    
    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    before_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    after_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def apply_header_style(ws, row, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    
    def apply_data_style(ws, start_row, end_row, num_cols):
        for row in range(start_row, end_row + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')
    
    def auto_column_width(ws):
        for col in ws.columns:
            max_length = 0
            column = None
            for cell in col:
                try:
                    # MergedCell은 column_letter가 없으므로 건너뜀
                    if column is None and hasattr(cell, 'column_letter'):
                        column = cell.column_letter
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column:
                ws.column_dimensions[column].width = min(max_length + 2, 40)
    
    # ========================================
    # Sheet 1: Comparison Summary (Before/After 비교)
    # ========================================
    ws1 = wb.active
    ws1.title = "Comparison_Summary"
    
    ws1.append(["📈 Phase 1 vs Phase 2 비교 분석"])
    ws1.merge_cells('A1:D1')
    ws1.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    ws1.append([])
    ws1.append(["지표", "Before (Sequential)", "After (Pipeline)", "개선율"])
    apply_header_style(ws1, 3, 4)
    
    if comparison:
        before = comparison["before"]
        after = comparison["after"]
        improvement = comparison["improvement"]
        
        ws1.append(["평균 Method Time (ms)", before['avg_method_time_ms'], after['avg_method_time_ms'], f"{improvement['method_time_pct']:.2f}%"])
        ws1.append(["평균 Latency (ms)", before['avg_cmd_latency_ms'], after['avg_cmd_latency_ms'], f"{improvement['latency_pct']:.2f}%"])
        ws1.append(["평균 I/O 비율 (%)", before['avg_io_ratio_pct'], after['avg_io_ratio_pct'], f"{improvement['io_ratio_change_pct']:+.2f}%p"])
        ws1.append(["총 레코드 수", before['total_records'], after['total_records'], "-"])
        
        apply_data_style(ws1, 4, 7, 4)
        
        # Before/After 색상
        for row in range(4, 8):
            ws1.cell(row=row, column=2).fill = before_fill
            ws1.cell(row=row, column=3).fill = after_fill
    
    auto_column_width(ws1)
    
    # ========================================
    # Sheet 2: Pipeline Metrics (Phase 2 상세)
    # ========================================
    ws2 = wb.create_sheet("Pipeline_Metrics")
    
    headers = ['Timestamp', 'Thread', 'District', 'Commands', 'MethodTime(ms)', 
               'PipelineLatency(ms)', 'IO_Ratio(%)', 'NonIO(ms)',
               'Result1_Count', 'Result2_Count', 'Result3_Count', 'Status', 'Intersection']
    ws2.append(headers)
    apply_header_style(ws2, 1, len(headers))
    
    for m in pipeline:
        ws2.append([
            m.timestamp, m.thread_name, m.district, m.commands,
            m.method_time_ms, m.pipeline_latency_ms, m.io_ratio_pct, m.non_io_time_ms,
            m.result1_count, m.result2_count, m.result3_count, m.status, m.intersection_count
        ])
    
    apply_data_style(ws2, 2, len(pipeline) + 1, len(headers))
    auto_column_width(ws2)
    
    # ========================================
    # Sheet 3: Sequential Metrics (Phase 1 상세, 있는 경우)
    # ========================================
    if sequential:
        ws3 = wb.create_sheet("Sequential_Metrics")
        
        headers = ['Timestamp', 'Thread', 'District', 'Commands', 'MethodTime(ms)', 
                   'CmdLatency(ms)', 'IO_Ratio(%)', 'NonIO(ms)',
                   'Cmd1(ms)', 'Cmd1_Count', 'Cmd2(ms)', 'Cmd2_Count', 
                   'Cmd3(ms)', 'Cmd3_Count', 'Status', 'Intersection']
        ws3.append(headers)
        apply_header_style(ws3, 1, len(headers))
        
        for m in sequential:
            ws3.append([
                m.timestamp, m.thread_name, m.district, m.commands,
                m.method_time_ms, m.total_cmd_latency_ms, m.io_ratio_pct, m.non_io_time_ms,
                m.cmd1_latency_ms, m.cmd1_count, m.cmd2_latency_ms, m.cmd2_count,
                m.cmd3_latency_ms, m.cmd3_count, m.status, m.intersection_count
            ])
        
        apply_data_style(ws3, 2, len(sequential) + 1, len(headers))
        auto_column_width(ws3)
    
    # ========================================
    # Sheet 4: LoopSummary Metrics
    # ========================================
    ws4 = wb.create_sheet("LoopSummary_Metrics")
    
    headers = ['Timestamp', 'Thread', 'Mode', 'TotalDistricts', 'SuccessDistricts',
               'EmptyDistricts', 'TotalProperties', 'LoopTime(ms)', 'AvgPerDistrict(ms)']
    ws4.append(headers)
    apply_header_style(ws4, 1, len(headers))
    
    for m in loop_summary:
        ws4.append([
            m.timestamp, m.thread_name, m.mode, m.total_districts,
            m.success_districts, m.empty_districts, m.total_properties,
            m.loop_time_ms, m.avg_per_district_ms
        ])
    
    apply_data_style(ws4, 2, len(loop_summary) + 1, len(headers))
    auto_column_width(ws4)
    
    # ========================================
    # Sheet 5: Analysis Summary
    # ========================================
    ws5 = wb.create_sheet("Analysis_Summary")
    
    ws5.append(["📊 Phase 2 (Pipeline) 분석 요약"])
    ws5.merge_cells('A1:C1')
    ws5.cell(row=1, column=1).font = Font(bold=True, size=12)
    
    ws5.append([])
    ws5.append(["지표", "값", "단위"])
    apply_header_style(ws5, 3, 3)
    
    if pipe_summary:
        data = [
            ("총 레코드 수", pipe_summary.total_records, "건"),
            ("평균 Method Time", round(pipe_summary.avg_method_time_ms, 4), "ms"),
            ("평균 Pipeline Latency", round(pipe_summary.avg_cmd_latency_ms, 4), "ms"),
            ("평균 I/O 비율", round(pipe_summary.avg_io_ratio_pct, 2), "%"),
            ("최소 Method Time", round(pipe_summary.min_method_time_ms, 4), "ms"),
            ("최대 Method Time", round(pipe_summary.max_method_time_ms, 4), "ms"),
            ("표준편차", round(pipe_summary.std_method_time_ms, 4), "ms"),
            ("성공률", round(pipe_summary.success_rate_pct, 2), "%"),
            ("빈 결과율", round(pipe_summary.early_termination_rate_pct, 2), "%"),
        ]
        
        for label, value, unit in data:
            ws5.append([label, value, unit])
        
        apply_data_style(ws5, 4, len(data) + 3, 3)
    
    auto_column_width(ws5)
    
    # ========================================
    # Sheet 6: District Analysis
    # ========================================
    ws6 = wb.create_sheet("District_Analysis")
    
    if pipeline:
        district_summaries = analyze_by_district(pipeline)
        
        headers = ['지역구', '건수', 'Avg_MethodTime(ms)', 'Avg_PipelineLatency(ms)', 
                   'IO_Ratio(%)', '성공률(%)']
        ws6.append(headers)
        apply_header_style(ws6, 1, len(headers))
        
        sorted_districts = sorted(
            district_summaries.items(),
            key=lambda x: x[1].avg_method_time_ms if x[1] else 0,
            reverse=True
        )
        
        for district, s in sorted_districts:
            if s:
                ws6.append([
                    district,
                    s.total_records,
                    round(s.avg_method_time_ms, 4),
                    round(s.avg_cmd_latency_ms, 4),
                    round(s.avg_io_ratio_pct, 2),
                    round(s.success_rate_pct, 2)
                ])
        
        apply_data_style(ws6, 2, len(sorted_districts) + 1, len(headers))
        auto_column_width(ws6)
    
    # 파일 저장
    wb.save(filepath)
    print(f"✅ Excel 파일 저장: {filepath}")
    print(f"   - Comparison_Summary: Before/After 비교")
    print(f"   - Pipeline_Metrics: {len(pipeline)}건 (Phase 2)")
    print(f"   - Sequential_Metrics: {len(sequential)}건 (Phase 1)")
    print(f"   - LoopSummary_Metrics: {len(loop_summary)}건")
    print(f"   - Analysis_Summary: Phase 2 분석 요약")
    print(f"   - District_Analysis: 지역구별 분석")


def export_to_json(sequential: List[SequentialMetric],
                   pipeline: List[PipelineMetric],
                   loop_summary: List[LoopSummaryMetric],
                   seq_summary: Optional[AnalysisSummary],
                   pipe_summary: Optional[AnalysisSummary],
                   comparison: Dict,
                   filepath: str):
    """JSON 파일로 내보내기"""
    data = {
        'metadata': {
            'generated_at': datetime.now().isoformat(),
            'sequential_count': len(sequential),
            'pipeline_count': len(pipeline),
            'loop_summary_count': len(loop_summary)
        },
        'comparison': comparison,
        'sequential_summary': asdict(seq_summary) if seq_summary else None,
        'pipeline_summary': asdict(pipe_summary) if pipe_summary else None,
        'sequential_metrics': [asdict(m) for m in sequential],
        'pipeline_metrics': [asdict(m) for m in pipeline],
        'loop_summary_metrics': [asdict(m) for m in loop_summary]
    }
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON 데이터 저장: {filepath}")


# =============================================================================
# 메인 함수
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Redis Pipeline 병목 측정 로그 파서 (Phase 1 + Phase 2)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  python RTT_Test_Phase2_parse.py phase2_log.txt
  python RTT_Test_Phase2_parse.py phase2_log.txt --output xlsx --export results.xlsx
  python RTT_Test_Phase2_parse.py combined_log.txt --output json --export results.json
        """
    )
    
    parser.add_argument('logfile', help='분석할 로그 파일 경로')
    parser.add_argument('--output', choices=['table', 'json', 'xlsx'], 
                        default='table', help='출력 형식 (기본: table)')
    parser.add_argument('--export', metavar='FILE', help='결과 내보내기 파일 경로')
    parser.add_argument('--limit', type=int, default=50, 
                        help='테이블 출력 시 최대 행 수 (기본: 50)')
    parser.add_argument('--quiet', '-q', action='store_true',
                        help='상세 데이터 출력 생략, 요약만 출력')
    
    args = parser.parse_args()
    
    try:
        print(f"\n🔍 로그 파일 분석 중: {args.logfile}")
        sequential_metrics, pipeline_metrics, loop_summary_metrics = parse_log_file(args.logfile)
        
        print(f"   - [Metrics-Sequential] 발견: {len(sequential_metrics)}건 (Phase 1)")
        print(f"   - [Metrics-Pipeline] 발견: {len(pipeline_metrics)}건 (Phase 2)")
        print(f"   - [Metrics-LoopSummary] 발견: {len(loop_summary_metrics)}건")
        
        if not sequential_metrics and not pipeline_metrics:
            print("\n⚠️  측정 로그를 찾을 수 없습니다.")
            sys.exit(1)
        
        # 분석 수행
        seq_summary = analyze_sequential_metrics(sequential_metrics)
        pipe_summary = analyze_pipeline_metrics(pipeline_metrics)
        comparison = compare_phases(seq_summary, pipe_summary)
        
        # 출력 형식에 따른 처리
        if args.output == 'xlsx':
            if not args.export:
                args.export = args.logfile.rsplit('.', 1)[0] + '_phase2_metrics.xlsx'
            export_to_xlsx(sequential_metrics, pipeline_metrics, loop_summary_metrics,
                          seq_summary, pipe_summary, comparison, args.export)
            print("\n✅ 분석 완료")
            return
        
        if args.output == 'json':
            if not args.export:
                args.export = args.logfile.rsplit('.', 1)[0] + '_phase2_metrics.json'
            export_to_json(sequential_metrics, pipeline_metrics, loop_summary_metrics,
                          seq_summary, pipe_summary, comparison, args.export)
            print("\n✅ 분석 완료")
            return
        
        # table 출력
        if not args.quiet:
            if pipeline_metrics:
                print_pipeline_table(pipeline_metrics, args.limit)
            
            if loop_summary_metrics:
                print_loop_summary_table(loop_summary_metrics)
        
        if pipe_summary:
            print_analysis_summary(pipe_summary)
        
        if comparison:
            print_comparison_summary(comparison)
        
        print("\n✅ 분석 완료")
        
    except FileNotFoundError as e:
        print(f"\n❌ 오류: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
