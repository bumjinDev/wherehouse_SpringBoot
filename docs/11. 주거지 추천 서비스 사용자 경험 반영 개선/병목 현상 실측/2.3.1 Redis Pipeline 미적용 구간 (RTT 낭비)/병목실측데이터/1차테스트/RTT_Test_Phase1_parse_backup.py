#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Redis Pipeline 병목 측정 로그 파서
==================================
Tomcat 로그 파일에서 [Metrics-Sequential] 및 [Metrics-LoopSummary] 로그를 
파싱하여 분석용 데이터로 변환한다.

사용법:
    python parse_redis_metrics.py <log_file_path> [options]
    
예시:
    python parse_redis_metrics.py /var/log/tomcat/catalina.out
    python parse_redis_metrics.py ./spring.log --output csv
    python parse_redis_metrics.py ./spring.log --output json --export results.json
"""

import re
import sys
import json
import argparse
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, asdict
from typing import List, Dict, Optional, Tuple
from statistics import mean, stdev, median
from collections import defaultdict


# =============================================================================
# 데이터 클래스 정의
# =============================================================================

@dataclass
class SequentialMetric:
    """[Metrics-Sequential] 로그 파싱 결과"""
    timestamp: Optional[str]
    thread_name: Optional[str]  # Tomcat 스레드 이름 (예: http-nio-8080-exec-1)
    district: str
    commands: int
    method_time_ms: float
    total_cmd_latency_ms: float
    io_ratio_pct: float
    non_io_time_ms: float
    cmd1_latency_ms: float
    cmd1_count: int
    cmd2_latency_ms: float
    cmd2_count: int
    cmd3_latency_ms: float
    cmd3_count: int
    status: str
    intersection_count: Optional[int] = None


@dataclass
class LoopSummaryMetric:
    """[Metrics-LoopSummary] 로그 파싱 결과"""
    timestamp: Optional[str]
    thread_name: Optional[str]  # Tomcat 스레드 이름
    mode: str
    total_districts: int
    success_districts: int
    empty_districts: int
    total_properties: int
    loop_time_ms: float
    avg_per_district_ms: float


@dataclass
class AnalysisSummary:
    """분석 요약 결과"""
    total_records: int
    avg_method_time_ms: float
    avg_cmd_latency_ms: float
    avg_io_ratio_pct: float
    min_method_time_ms: float
    max_method_time_ms: float
    std_method_time_ms: float
    avg_cmd1_ms: float
    avg_cmd2_ms: float
    avg_cmd3_ms: float
    success_rate_pct: float
    early_termination_rate_pct: float


# =============================================================================
# 정규식 패턴 정의
# =============================================================================

# 타임스탬프 패턴 (다양한 로그 포맷 대응)
TIMESTAMP_PATTERNS = [
    r'(\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}[.,]?\d*)',  # ISO 형식
    r'(\d{2}-\w{3}-\d{4} \d{2}:\d{2}:\d{2})',              # Tomcat 기본
    r'(\d{2}:\d{2}:\d{2}[.,]\d+)',                          # 시간만
]

# 스레드 이름 패턴 (다양한 로그 포맷 대응)
# 예: [http-nio-8080-exec-1], [main], [scheduling-1], [pool-1-thread-1]
THREAD_PATTERNS = [
    r'\[([^\]]*exec[^\]]*)\]',           # http-nio-8080-exec-1, http-apr-8080-exec-3
    r'\[([^\]]*thread[^\]]*)\]',         # pool-1-thread-1
    r'\[(scheduling-\d+)\]',             # scheduling-1
    r'\[(main)\]',                       # main
    r'--- \[([^\]]+)\]',                 # Spring Boot 형식: --- [thread-name]
    r'\] \[([^\]]+)\] [a-z]',            # 일반 형식: ] [thread-name] c.w.r
    r'\[([a-zA-Z]+-[a-zA-Z]+-\d+-[a-zA-Z]+-\d+)\]',  # 일반적인 스레드 패턴
]

# [Metrics-Sequential] 로그 패턴
SEQUENTIAL_PATTERN = re.compile(
    r'\[Metrics-Sequential\]\s*'
    r'district=([^,]+),\s*'
    r'commands=(-?\d+),\s*'
    r'methodTime=([\d.]+)\s*ms,\s*'
    r'totalCmdLatency=([\d.]+)\s*ms,\s*'
    r'ioRatio=([\d.]+)\s*%,\s*'
    r'nonIoTime=([\d.]+)\s*ms,\s*'
    r'cmd1=([\d.]+)\s*ms\s*\((\d+)건\),\s*'
    r'cmd2=([\d.]+)\s*ms\s*\((\d+)건\),\s*'
    r'cmd3=([\d.]+)\s*ms\s*\((\d+)건\),\s*'
    r'status=(.+?)(?:\s*$|\s*\n)',
    re.UNICODE
)

# [Metrics-LoopSummary] 로그 패턴
LOOP_SUMMARY_PATTERN = re.compile(
    r'\[Metrics-LoopSummary\]\s*'
    r'mode=(\w+),\s*'
    r'totalDistricts=(\d+),\s*'
    r'successDistricts=(\d+),\s*'
    r'emptyDistricts=(\d+),\s*'
    r'totalProperties=(\d+),\s*'
    r'loopTime=([\d.]+)\s*ms,\s*'
    r'avgPerDistrict=([\d.]+)\s*ms',
    re.UNICODE
)


# =============================================================================
# 파싱 함수
# =============================================================================

def extract_timestamp(line: str) -> Optional[str]:
    """로그 라인에서 타임스탬프 추출"""
    for pattern in TIMESTAMP_PATTERNS:
        match = re.search(pattern, line)
        if match:
            return match.group(1)
    return None


def extract_thread_name(line: str) -> Optional[str]:
    """로그 라인에서 Tomcat 스레드 이름 추출
    
    지원 형식:
    - [http-nio-8080-exec-1] 
    - [pool-1-thread-1]
    - --- [thread-name] (Spring Boot)
    """
    for pattern in THREAD_PATTERNS:
        match = re.search(pattern, line)
        if match:
            return match.group(1)
    return None


def parse_sequential_log(line: str) -> Optional[SequentialMetric]:
    """[Metrics-Sequential] 로그 파싱"""
    match = SEQUENTIAL_PATTERN.search(line)
    if not match:
        return None
    
    timestamp = extract_timestamp(line)
    thread_name = extract_thread_name(line)
    status = match.group(13).strip()
    
    # intersection 값 추출
    intersection_count = None
    if 'intersection=' in status:
        try:
            intersection_count = int(re.search(r'intersection=(\d+)', status).group(1))
        except (AttributeError, ValueError):
            pass
    
    return SequentialMetric(
        timestamp=timestamp,
        thread_name=thread_name,
        district=match.group(1).strip(),
        commands=int(match.group(2)),
        method_time_ms=float(match.group(3)),
        total_cmd_latency_ms=float(match.group(4)),
        io_ratio_pct=float(match.group(5)),
        non_io_time_ms=float(match.group(6)),
        cmd1_latency_ms=float(match.group(7)),
        cmd1_count=int(match.group(8)),
        cmd2_latency_ms=float(match.group(9)),
        cmd2_count=int(match.group(10)),
        cmd3_latency_ms=float(match.group(11)),
        cmd3_count=int(match.group(12)),
        status=status,
        intersection_count=intersection_count
    )


def parse_loop_summary_log(line: str) -> Optional[LoopSummaryMetric]:
    """[Metrics-LoopSummary] 로그 파싱"""
    match = LOOP_SUMMARY_PATTERN.search(line)
    if not match:
        return None
    
    timestamp = extract_timestamp(line)
    thread_name = extract_thread_name(line)
    
    return LoopSummaryMetric(
        timestamp=timestamp,
        thread_name=thread_name,
        mode=match.group(1),
        total_districts=int(match.group(2)),
        success_districts=int(match.group(3)),
        empty_districts=int(match.group(4)),
        total_properties=int(match.group(5)),
        loop_time_ms=float(match.group(6)),
        avg_per_district_ms=float(match.group(7))
    )


def parse_log_file(filepath: str) -> Tuple[List[SequentialMetric], List[LoopSummaryMetric]]:
    """로그 파일 전체 파싱"""
    sequential_metrics = []
    loop_summary_metrics = []
    
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"로그 파일을 찾을 수 없습니다: {filepath}")
    
    # 인코딩 자동 감지 시도
    encodings = ['utf-8', 'cp949', 'euc-kr', 'latin-1']
    content = None
    
    for encoding in encodings:
        try:
            with open(filepath, 'r', encoding=encoding) as f:
                content = f.read()
            break
        except UnicodeDecodeError:
            continue
    
    if content is None:
        raise ValueError(f"로그 파일 인코딩을 감지할 수 없습니다: {filepath}")
    
    for line in content.splitlines():
        # Sequential 메트릭 파싱
        if '[Metrics-Sequential]' in line:
            metric = parse_sequential_log(line)
            if metric:
                sequential_metrics.append(metric)
        
        # LoopSummary 메트릭 파싱
        if '[Metrics-LoopSummary]' in line:
            metric = parse_loop_summary_log(line)
            if metric:
                loop_summary_metrics.append(metric)
    
    return sequential_metrics, loop_summary_metrics


# =============================================================================
# 분석 함수
# =============================================================================

def analyze_sequential_metrics(metrics: List[SequentialMetric]) -> Optional[AnalysisSummary]:
    """Sequential 메트릭 통계 분석"""
    if not metrics:
        return None
    
    method_times = [m.method_time_ms for m in metrics]
    cmd_latencies = [m.total_cmd_latency_ms for m in metrics]
    io_ratios = [m.io_ratio_pct for m in metrics]
    cmd1_times = [m.cmd1_latency_ms for m in metrics]
    cmd2_times = [m.cmd2_latency_ms for m in metrics if m.cmd2_latency_ms > 0]
    cmd3_times = [m.cmd3_latency_ms for m in metrics if m.cmd3_latency_ms > 0]
    
    success_count = sum(1 for m in metrics if 'SUCCESS' in m.status)
    early_term_count = sum(1 for m in metrics if 'EARLY_RETURN' in m.status)
    
    return AnalysisSummary(
        total_records=len(metrics),
        avg_method_time_ms=mean(method_times),
        avg_cmd_latency_ms=mean(cmd_latencies),
        avg_io_ratio_pct=mean(io_ratios),
        min_method_time_ms=min(method_times),
        max_method_time_ms=max(method_times),
        std_method_time_ms=stdev(method_times) if len(method_times) > 1 else 0,
        avg_cmd1_ms=mean(cmd1_times),
        avg_cmd2_ms=mean(cmd2_times) if cmd2_times else 0,
        avg_cmd3_ms=mean(cmd3_times) if cmd3_times else 0,
        success_rate_pct=(success_count / len(metrics)) * 100,
        early_termination_rate_pct=(early_term_count / len(metrics)) * 100
    )


def analyze_by_district(metrics: List[SequentialMetric]) -> Dict[str, AnalysisSummary]:
    """지역구별 분석"""
    by_district = defaultdict(list)
    for m in metrics:
        by_district[m.district].append(m)
    
    results = {}
    for district, district_metrics in by_district.items():
        summary = analyze_sequential_metrics(district_metrics)
        if summary:
            results[district] = summary
    
    return results


def analyze_by_thread(metrics: List[SequentialMetric]) -> Dict[str, AnalysisSummary]:
    """스레드별 분석"""
    by_thread = defaultdict(list)
    for m in metrics:
        thread_key = m.thread_name if m.thread_name else "(unknown)"
        by_thread[thread_key].append(m)
    
    results = {}
    for thread_name, thread_metrics in by_thread.items():
        summary = analyze_sequential_metrics(thread_metrics)
        if summary:
            results[thread_name] = summary
    
    return results


def group_by_request(sequential: List[SequentialMetric], 
                     loop_summary: List[LoopSummaryMetric]) -> Dict[str, Dict]:
    """스레드+타임스탬프 기반으로 요청(Request) 단위 그룹화
    
    하나의 API 요청에서 발생한 25개 지역구 처리 + 1개 LoopSummary를 묶음
    """
    requests = defaultdict(lambda: {'sequential': [], 'loop_summary': None})
    
    # LoopSummary를 기준으로 요청 식별
    for ls in loop_summary:
        key = f"{ls.thread_name}|{ls.timestamp}" if ls.thread_name else f"(unknown)|{ls.timestamp}"
        requests[key]['loop_summary'] = ls
    
    # Sequential 메트릭을 해당 요청에 매핑
    for seq in sequential:
        # 가장 가까운 LoopSummary 찾기 (같은 스레드, 시간 근접)
        thread = seq.thread_name if seq.thread_name else "(unknown)"
        
        # 단순 매핑: 같은 스레드의 모든 sequential을 해당 스레드로 그룹화
        matched = False
        for key in requests:
            if key.startswith(thread + "|"):
                requests[key]['sequential'].append(seq)
                matched = True
                break
        
        if not matched:
            # LoopSummary 없이 단독 sequential인 경우
            key = f"{thread}|{seq.timestamp}"
            requests[key]['sequential'].append(seq)
    
    return dict(requests)


# =============================================================================
# 출력 함수
# =============================================================================

def print_sequential_table(metrics: List[SequentialMetric], limit: int = 50):
    """Sequential 메트릭 테이블 출력 (스레드 정보 포함)"""
    print("\n" + "=" * 140)
    print("[Metrics-Sequential] 상세 데이터")
    print("=" * 140)
    
    header = f"{'스레드':<25} {'지역구':<10} {'Cmd':>3} {'MethodTime':>12} {'CmdLatency':>12} {'I/O%':>8} {'Cmd1':>10} {'Cmd2':>10} {'Cmd3':>10} {'Status':<25}"
    print(header)
    print("-" * 140)
    
    for i, m in enumerate(metrics[:limit]):
        thread_display = (m.thread_name[:23] + "..") if m.thread_name and len(m.thread_name) > 25 else (m.thread_name or "(unknown)")
        status_display = m.status[:23] + ".." if len(m.status) > 25 else m.status
        print(f"{thread_display:<25} {m.district:<10} {m.commands:>3} "
              f"{m.method_time_ms:>10.4f}ms {m.total_cmd_latency_ms:>10.4f}ms "
              f"{m.io_ratio_pct:>6.2f}% "
              f"{m.cmd1_latency_ms:>8.4f}ms {m.cmd2_latency_ms:>8.4f}ms {m.cmd3_latency_ms:>8.4f}ms "
              f"{status_display:<25}")
    
    if len(metrics) > limit:
        print(f"\n... 외 {len(metrics) - limit}건 (--limit 옵션으로 조정 가능)")


def print_loop_summary_table(metrics: List[LoopSummaryMetric]):
    """LoopSummary 메트릭 테이블 출력 (스레드 정보 포함)"""
    print("\n" + "=" * 130)
    print("[Metrics-LoopSummary] 전체 순회 요약")
    print("=" * 130)
    
    header = f"{'스레드':<25} {'Mode':<12} {'총지역구':>8} {'성공':>6} {'빈결과':>6} {'총매물':>8} {'LoopTime':>12} {'Avg/District':>14}"
    print(header)
    print("-" * 130)
    
    for m in metrics:
        thread_display = (m.thread_name[:23] + "..") if m.thread_name and len(m.thread_name) > 25 else (m.thread_name or "(unknown)")
        print(f"{thread_display:<25} {m.mode:<12} {m.total_districts:>8} {m.success_districts:>6} "
              f"{m.empty_districts:>6} {m.total_properties:>8} "
              f"{m.loop_time_ms:>10.4f}ms {m.avg_per_district_ms:>12.4f}ms")


def print_analysis_summary(summary: AnalysisSummary, title: str = "전체"):
    """분석 요약 출력"""
    print(f"\n{'─' * 60}")
    print(f"📊 분석 요약 ({title})")
    print(f"{'─' * 60}")
    print(f"  총 레코드 수        : {summary.total_records}건")
    print(f"  평균 Method Time    : {summary.avg_method_time_ms:.4f} ms")
    print(f"  평균 Cmd Latency    : {summary.avg_cmd_latency_ms:.4f} ms")
    print(f"  평균 I/O 비율       : {summary.avg_io_ratio_pct:.2f} %")
    print(f"  최소 Method Time    : {summary.min_method_time_ms:.4f} ms")
    print(f"  최대 Method Time    : {summary.max_method_time_ms:.4f} ms")
    print(f"  표준편차            : {summary.std_method_time_ms:.4f} ms")
    print(f"  평균 Cmd1 (보증금)  : {summary.avg_cmd1_ms:.4f} ms")
    print(f"  평균 Cmd2 (월세)    : {summary.avg_cmd2_ms:.4f} ms")
    print(f"  평균 Cmd3 (평수)    : {summary.avg_cmd3_ms:.4f} ms")
    print(f"  성공률              : {summary.success_rate_pct:.2f} %")
    print(f"  조기종료율          : {summary.early_termination_rate_pct:.2f} %")


def print_district_analysis(district_summaries: Dict[str, AnalysisSummary]):
    """지역구별 분석 출력"""
    print("\n" + "=" * 80)
    print("📍 지역구별 분석")
    print("=" * 80)
    
    header = f"{'지역구':<10} {'건수':>6} {'Avg Method':>12} {'Avg CmdLat':>12} {'I/O%':>8} {'성공률':>8}"
    print(header)
    print("-" * 80)
    
    # 평균 Method Time 기준 정렬
    sorted_districts = sorted(
        district_summaries.items(), 
        key=lambda x: x[1].avg_method_time_ms, 
        reverse=True
    )
    
    for district, summary in sorted_districts:
        print(f"{district:<10} {summary.total_records:>6} "
              f"{summary.avg_method_time_ms:>10.4f}ms "
              f"{summary.avg_cmd_latency_ms:>10.4f}ms "
              f"{summary.avg_io_ratio_pct:>6.2f}% "
              f"{summary.success_rate_pct:>6.2f}%")


def print_thread_analysis(thread_summaries: Dict[str, AnalysisSummary]):
    """스레드별 분석 출력"""
    print("\n" + "=" * 100)
    print("🧵 스레드별 분석")
    print("=" * 100)
    
    header = f"{'스레드':<30} {'건수':>6} {'Avg Method':>12} {'Avg CmdLat':>12} {'I/O%':>8} {'성공률':>8}"
    print(header)
    print("-" * 100)
    
    # 건수 기준 정렬
    sorted_threads = sorted(
        thread_summaries.items(), 
        key=lambda x: x[1].total_records, 
        reverse=True
    )
    
    for thread_name, summary in sorted_threads:
        thread_display = (thread_name[:28] + "..") if len(thread_name) > 30 else thread_name
        print(f"{thread_display:<30} {summary.total_records:>6} "
              f"{summary.avg_method_time_ms:>10.4f}ms "
              f"{summary.avg_cmd_latency_ms:>10.4f}ms "
              f"{summary.avg_io_ratio_pct:>6.2f}% "
              f"{summary.success_rate_pct:>6.2f}%")


def print_request_analysis(requests: Dict[str, Dict]):
    """요청 단위 분석 출력 (스레드별 하나의 완전한 요청)"""
    print("\n" + "=" * 120)
    print("📦 요청(Request) 단위 분석 - 각 API 호출별 메트릭")
    print("=" * 120)
    
    header = f"{'스레드':<25} {'Timestamp':<26} {'지역구':>6} {'LoopTime':>12} {'Avg/District':>14} {'총매물':>8}"
    print(header)
    print("-" * 120)
    
    for key, data in requests.items():
        parts = key.split("|", 1)
        thread_name = parts[0] if len(parts) > 0 else "(unknown)"
        timestamp = parts[1] if len(parts) > 1 else "-"
        
        thread_display = (thread_name[:23] + "..") if len(thread_name) > 25 else thread_name
        
        seq_count = len(data['sequential'])
        ls = data['loop_summary']
        
        if ls:
            print(f"{thread_display:<25} {timestamp:<26} {seq_count:>6} "
                  f"{ls.loop_time_ms:>10.4f}ms {ls.avg_per_district_ms:>12.4f}ms "
                  f"{ls.total_properties:>8}")
        else:
            # LoopSummary 없이 Sequential만 있는 경우
            total_method_time = sum(m.method_time_ms for m in data['sequential'])
            avg_method_time = total_method_time / seq_count if seq_count > 0 else 0
            print(f"{thread_display:<25} {timestamp:<26} {seq_count:>6} "
                  f"{total_method_time:>10.4f}ms {avg_method_time:>12.4f}ms "
                  f"{'(N/A)':>8}")


def print_before_after_template(summary: AnalysisSummary):
    """Before/After 비교용 템플릿 출력"""
    print("\n" + "=" * 70)
    print("📋 Before/After 비교용 데이터 (복사하여 사용)")
    print("=" * 70)
    print(f"""
| 지표                  | Before (순차)         | After (Pipeline)      | 개선율      |
|-----------------------|-----------------------|-----------------------|-------------|
| 평균 Method Time      | {summary.avg_method_time_ms:.4f} ms         | ___.___ ms            | __.___%     |
| 평균 Cmd Latency      | {summary.avg_cmd_latency_ms:.4f} ms         | ___.___ ms            | __.___%     |
| 평균 I/O 비율         | {summary.avg_io_ratio_pct:.2f} %           | __.___%               | -           |
| Cmd1 평균 (보증금)    | {summary.avg_cmd1_ms:.4f} ms         | -                     | -           |
| Cmd2 평균 (월세)      | {summary.avg_cmd2_ms:.4f} ms         | -                     | -           |
| Cmd3 평균 (평수)      | {summary.avg_cmd3_ms:.4f} ms         | -                     | -           |
| 측정 건수             | {summary.total_records}건                | ___건                 | -           |
| 성공률                | {summary.success_rate_pct:.2f} %           | __.___%               | -           |
| 조기종료율            | {summary.early_termination_rate_pct:.2f} %           | 0.00% (불가)          | -           |
""")


def export_to_csv(sequential: List[SequentialMetric], 
                  loop_summary: List[LoopSummaryMetric], 
                  filepath: str):
    """CSV 파일로 내보내기"""
    import csv
    
    # Sequential 데이터
    seq_path = filepath.replace('.csv', '_sequential.csv')
    with open(seq_path, 'w', newline='', encoding='utf-8-sig') as f:
        if sequential:
            writer = csv.DictWriter(f, fieldnames=asdict(sequential[0]).keys())
            writer.writeheader()
            for m in sequential:
                writer.writerow(asdict(m))
    print(f"✅ Sequential 데이터 저장: {seq_path}")
    
    # LoopSummary 데이터
    loop_path = filepath.replace('.csv', '_loop_summary.csv')
    with open(loop_path, 'w', newline='', encoding='utf-8-sig') as f:
        if loop_summary:
            writer = csv.DictWriter(f, fieldnames=asdict(loop_summary[0]).keys())
            writer.writeheader()
            for m in loop_summary:
                writer.writerow(asdict(m))
    print(f"✅ LoopSummary 데이터 저장: {loop_path}")


def export_to_xlsx(sequential: List[SequentialMetric],
                   loop_summary: List[LoopSummaryMetric],
                   summary: Optional[AnalysisSummary],
                   filepath: str):
    """Excel 파일로 내보내기 - 데이터 유형별 시트 분리"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    
    # 스타일 정의
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    number_font = Font(name='Consolas')
    
    def apply_header_style(sheet, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = sheet.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def apply_data_style(sheet, start_row, end_row, col_count):
        for row in range(start_row, end_row + 1):
            for col in range(1, col_count + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
    
    def auto_column_width(sheet):
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[column].width = min(max_length + 2, 50)
    
    # ========================================
    # Sheet 1: Sequential Metrics (상세 데이터)
    # ========================================
    ws1 = wb.active
    ws1.title = "Sequential_Metrics"
    
    if sequential:
        headers = ['Timestamp', 'Thread', 'District', 'Commands', 
                   'MethodTime(ms)', 'CmdLatency(ms)', 'IO_Ratio(%)', 'NonIO(ms)',
                   'Cmd1(ms)', 'Cmd1_Count', 'Cmd2(ms)', 'Cmd2_Count', 
                   'Cmd3(ms)', 'Cmd3_Count', 'Status', 'Intersection']
        ws1.append(headers)
        apply_header_style(ws1, 1, len(headers))
        
        for m in sequential:
            ws1.append([
                m.timestamp or '',
                m.thread_name or '',
                m.district,
                m.commands,
                round(m.method_time_ms, 4),
                round(m.total_cmd_latency_ms, 4),
                round(m.io_ratio_pct, 2),
                round(m.non_io_time_ms, 4),
                round(m.cmd1_latency_ms, 4),
                m.cmd1_count,
                round(m.cmd2_latency_ms, 4),
                m.cmd2_count,
                round(m.cmd3_latency_ms, 4),
                m.cmd3_count,
                m.status,
                m.intersection_count if m.intersection_count is not None else ''
            ])
        
        apply_data_style(ws1, 2, len(sequential) + 1, len(headers))
        auto_column_width(ws1)
    
    # ========================================
    # Sheet 2: LoopSummary Metrics (순회 요약)
    # ========================================
    ws2 = wb.create_sheet("LoopSummary_Metrics")
    
    if loop_summary:
        headers = ['Timestamp', 'Thread', 'Mode', 'TotalDistricts', 
                   'SuccessDistricts', 'EmptyDistricts', 'TotalProperties',
                   'LoopTime(ms)', 'AvgPerDistrict(ms)']
        ws2.append(headers)
        apply_header_style(ws2, 1, len(headers))
        
        for m in loop_summary:
            ws2.append([
                m.timestamp or '',
                m.thread_name or '',
                m.mode,
                m.total_districts,
                m.success_districts,
                m.empty_districts,
                m.total_properties,
                round(m.loop_time_ms, 4),
                round(m.avg_per_district_ms, 4)
            ])
        
        apply_data_style(ws2, 2, len(loop_summary) + 1, len(headers))
        auto_column_width(ws2)
    
    # ========================================
    # Sheet 3: Analysis Summary (분석 요약)
    # ========================================
    ws3 = wb.create_sheet("Analysis_Summary")
    
    if summary:
        ws3.append(['지표', '값', '단위'])
        apply_header_style(ws3, 1, 3)
        
        data = [
            ('총 레코드 수', summary.total_records, '건'),
            ('평균 Method Time', round(summary.avg_method_time_ms, 4), 'ms'),
            ('평균 Cmd Latency', round(summary.avg_cmd_latency_ms, 4), 'ms'),
            ('평균 I/O 비율', round(summary.avg_io_ratio_pct, 2), '%'),
            ('최소 Method Time', round(summary.min_method_time_ms, 4), 'ms'),
            ('최대 Method Time', round(summary.max_method_time_ms, 4), 'ms'),
            ('표준편차', round(summary.std_method_time_ms, 4), 'ms'),
            ('평균 Cmd1 (보증금)', round(summary.avg_cmd1_ms, 4), 'ms'),
            ('평균 Cmd2 (월세)', round(summary.avg_cmd2_ms, 4), 'ms'),
            ('평균 Cmd3 (평수)', round(summary.avg_cmd3_ms, 4), 'ms'),
            ('성공률', round(summary.success_rate_pct, 2), '%'),
            ('조기종료율', round(summary.early_termination_rate_pct, 2), '%'),
        ]
        
        for row in data:
            ws3.append(row)
        
        apply_data_style(ws3, 2, len(data) + 1, 3)
        auto_column_width(ws3)
    
    # ========================================
    # Sheet 4: District Analysis (지역구별 분석)
    # ========================================
    ws4 = wb.create_sheet("District_Analysis")
    
    if sequential:
        district_summaries = analyze_by_district(sequential)
        
        headers = ['지역구', '건수', 'Avg_MethodTime(ms)', 'Avg_CmdLatency(ms)', 
                   'IO_Ratio(%)', '성공률(%)']
        ws4.append(headers)
        apply_header_style(ws4, 1, len(headers))
        
        sorted_districts = sorted(
            district_summaries.items(),
            key=lambda x: x[1].avg_method_time_ms,
            reverse=True
        )
        
        for district, s in sorted_districts:
            ws4.append([
                district,
                s.total_records,
                round(s.avg_method_time_ms, 4),
                round(s.avg_cmd_latency_ms, 4),
                round(s.avg_io_ratio_pct, 2),
                round(s.success_rate_pct, 2)
            ])
        
        apply_data_style(ws4, 2, len(sorted_districts) + 1, len(headers))
        auto_column_width(ws4)
    
    # ========================================
    # Sheet 5: Thread Analysis (스레드별 분석)
    # ========================================
    ws5 = wb.create_sheet("Thread_Analysis")
    
    if sequential:
        thread_summaries = analyze_by_thread(sequential)
        
        headers = ['스레드', '건수', 'Avg_MethodTime(ms)', 'Avg_CmdLatency(ms)',
                   'IO_Ratio(%)', '성공률(%)']
        ws5.append(headers)
        apply_header_style(ws5, 1, len(headers))
        
        sorted_threads = sorted(
            thread_summaries.items(),
            key=lambda x: x[1].total_records,
            reverse=True
        )
        
        for thread_name, s in sorted_threads:
            ws5.append([
                thread_name,
                s.total_records,
                round(s.avg_method_time_ms, 4),
                round(s.avg_cmd_latency_ms, 4),
                round(s.avg_io_ratio_pct, 2),
                round(s.success_rate_pct, 2)
            ])
        
        apply_data_style(ws5, 2, len(sorted_threads) + 1, len(headers))
        auto_column_width(ws5)
    
    # ========================================
    # Sheet 6: Before/After Template (비교 템플릿)
    # ========================================
    ws6 = wb.create_sheet("Before_After_Compare")
    
    if summary:
        headers = ['지표', 'Before (순차)', 'After (Pipeline)', '개선율']
        ws6.append(headers)
        apply_header_style(ws6, 1, len(headers))
        
        template_data = [
            ('평균 Method Time (ms)', round(summary.avg_method_time_ms, 4), '', ''),
            ('평균 Cmd Latency (ms)', round(summary.avg_cmd_latency_ms, 4), '', ''),
            ('평균 I/O 비율 (%)', round(summary.avg_io_ratio_pct, 2), '', '-'),
            ('Cmd1 평균 - 보증금 (ms)', round(summary.avg_cmd1_ms, 4), '-', '-'),
            ('Cmd2 평균 - 월세 (ms)', round(summary.avg_cmd2_ms, 4), '-', '-'),
            ('Cmd3 평균 - 평수 (ms)', round(summary.avg_cmd3_ms, 4), '-', '-'),
            ('측정 건수', summary.total_records, '', '-'),
            ('성공률 (%)', round(summary.success_rate_pct, 2), '', '-'),
            ('조기종료율 (%)', round(summary.early_termination_rate_pct, 2), '0.00 (불가)', '-'),
        ]
        
        for row in template_data:
            ws6.append(row)
        
        apply_data_style(ws6, 2, len(template_data) + 1, len(headers))
        
        # After, 개선율 컬럼 노란색 배경 (입력 대기)
        yellow_fill = PatternFill('solid', fgColor='FFFF00')
        for row in range(2, len(template_data) + 2):
            for col in [3, 4]:  # C, D 컬럼
                cell = ws6.cell(row=row, column=col)
                if cell.value == '':
                    cell.fill = yellow_fill
        
        auto_column_width(ws6)
    
    # 파일 저장
    wb.save(filepath)
    print(f"✅ Excel 파일 저장: {filepath}")
    print(f"   - Sequential_Metrics: {len(sequential)}건")
    print(f"   - LoopSummary_Metrics: {len(loop_summary)}건")
    print(f"   - Analysis_Summary: 분석 요약")
    print(f"   - District_Analysis: 지역구별 분석")
    print(f"   - Thread_Analysis: 스레드별 분석")
    print(f"   - Before_After_Compare: 비교 템플릿")


def export_to_json(sequential: List[SequentialMetric], 
                   loop_summary: List[LoopSummaryMetric],
                   summary: Optional[AnalysisSummary],
                   filepath: str):
    """JSON 파일로 내보내기"""
    data = {
        'metadata': {
            'generated_at': datetime.now().isoformat(),
            'sequential_count': len(sequential),
            'loop_summary_count': len(loop_summary)
        },
        'analysis_summary': asdict(summary) if summary else None,
        'sequential_metrics': [asdict(m) for m in sequential],
        'loop_summary_metrics': [asdict(m) for m in loop_summary]
    }
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON 데이터 저장: {filepath}")


# =============================================================================
# 메인 함수
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Redis Pipeline 병목 측정 로그 파서',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  python parse_redis_metrics.py catalina.out
  python parse_redis_metrics.py spring.log --output csv --export results.csv
  python parse_redis_metrics.py spring.log --output xlsx --export results.xlsx
  python parse_redis_metrics.py app.log --district-analysis
  python parse_redis_metrics.py app.log --thread-analysis
  python parse_redis_metrics.py app.log --request-analysis
  python parse_redis_metrics.py app.log --limit 100 --template
        """
    )
    
    parser.add_argument('logfile', help='분석할 로그 파일 경로')
    parser.add_argument('--output', choices=['table', 'csv', 'json', 'xlsx'], 
                        default='table', help='출력 형식 (기본: table)')
    parser.add_argument('--export', metavar='FILE', help='결과 내보내기 파일 경로')
    parser.add_argument('--limit', type=int, default=50, 
                        help='테이블 출력 시 최대 행 수 (기본: 50)')
    parser.add_argument('--district-analysis', action='store_true',
                        help='지역구별 상세 분석 출력')
    parser.add_argument('--thread-analysis', action='store_true',
                        help='스레드별 상세 분석 출력')
    parser.add_argument('--request-analysis', action='store_true',
                        help='요청(Request) 단위 분석 출력')
    parser.add_argument('--template', action='store_true',
                        help='Before/After 비교 템플릿 출력')
    parser.add_argument('--quiet', '-q', action='store_true',
                        help='상세 데이터 출력 생략, 요약만 출력')
    parser.add_argument('--filter-thread', metavar='PATTERN',
                        help='특정 스레드만 필터링 (예: exec-1, exec-*)')
    
    args = parser.parse_args()
    
    try:
        print(f"\n🔍 로그 파일 분석 중: {args.logfile}")
        sequential_metrics, loop_summary_metrics = parse_log_file(args.logfile)
        
        # 스레드 필터 적용
        if args.filter_thread:
            pattern = args.filter_thread.replace('*', '.*')
            sequential_metrics = [m for m in sequential_metrics 
                                  if m.thread_name and re.search(pattern, m.thread_name)]
            loop_summary_metrics = [m for m in loop_summary_metrics 
                                    if m.thread_name and re.search(pattern, m.thread_name)]
            print(f"   - 스레드 필터 적용: '{args.filter_thread}'")
        
        print(f"   - [Metrics-Sequential] 발견: {len(sequential_metrics)}건")
        print(f"   - [Metrics-LoopSummary] 발견: {len(loop_summary_metrics)}건")
        
        # 스레드 목록 출력
        threads = set(m.thread_name for m in sequential_metrics if m.thread_name)
        if threads:
            print(f"   - 감지된 스레드: {len(threads)}개")
            for t in sorted(threads)[:5]:
                print(f"       • {t}")
            if len(threads) > 5:
                print(f"       ... 외 {len(threads) - 5}개")
        
        if not sequential_metrics and not loop_summary_metrics:
            print("\n⚠️  측정 로그를 찾을 수 없습니다.")
            print("    로그 파일에 [Metrics-Sequential] 또는 [Metrics-LoopSummary] 패턴이 있는지 확인하세요.")
            sys.exit(1)
        
        # 분석 수행
        summary = analyze_sequential_metrics(sequential_metrics)
        
        # xlsx 출력인 경우 바로 파일 생성
        if args.output == 'xlsx':
            if not args.export:
                # export 미지정 시 기본 파일명 생성
                args.export = args.logfile.rsplit('.', 1)[0] + '_metrics.xlsx'
            export_to_xlsx(sequential_metrics, loop_summary_metrics, summary, args.export)
            print("\n✅ 분석 완료")
            return
        
        # 출력 형식에 따른 처리
        if args.output == 'table':
            if not args.quiet and sequential_metrics:
                print_sequential_table(sequential_metrics, args.limit)
            
            if not args.quiet and loop_summary_metrics:
                print_loop_summary_table(loop_summary_metrics)
            
            if summary:
                print_analysis_summary(summary)
            
            if args.district_analysis and sequential_metrics:
                district_summaries = analyze_by_district(sequential_metrics)
                print_district_analysis(district_summaries)
            
            if args.thread_analysis and sequential_metrics:
                thread_summaries = analyze_by_thread(sequential_metrics)
                print_thread_analysis(thread_summaries)
            
            if args.request_analysis:
                requests = group_by_request(sequential_metrics, loop_summary_metrics)
                print_request_analysis(requests)
            
            if args.template and summary:
                print_before_after_template(summary)
        
        # 파일 내보내기
        if args.export:
            if args.output == 'csv':
                export_to_csv(sequential_metrics, loop_summary_metrics, args.export)
            elif args.output == 'json':
                export_to_json(sequential_metrics, loop_summary_metrics, summary, args.export)
            else:
                # table 모드에서도 export 지정 시 JSON으로 저장
                export_to_json(sequential_metrics, loop_summary_metrics, summary, args.export)
        
        print("\n✅ 분석 완료")
        
    except FileNotFoundError as e:
        print(f"\n❌ 오류: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
