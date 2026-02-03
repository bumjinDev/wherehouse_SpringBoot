#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Wherehouse OOM 병목 측정 로그 파서 (2차 테스트 - Slice 청크 처리)
- 기반: OOM_병목_측정_명세서_v1_1.docx
- 2차 테스트: Slice 기반 청크 처리 로그 포맷 대응
- 출력: 콘솔 (기본) / Excel 파일 (--xlsx 옵션)

사용법:
    python perf_log_parser_phase2.py <로그파일경로>
    python perf_log_parser_phase2.py wherehouse.log --xlsx
    python perf_log_parser_phase2.py wherehouse.log --xlsx output.xlsx
"""

import re
import sys
from dataclasses import dataclass, field
from typing import Dict, List, Optional
from pathlib import Path
from statistics import mean, stdev

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# =============================================================================
# 데이터 클래스 정의
# =============================================================================

@dataclass
class ChunkMetric:
    """청크별 측정 데이터"""
    chunk_index: int
    chunk_size: int
    cumulative: int
    load_ms: int
    transform_ms: int
    total_ms: int
    has_next: bool


@dataclass
class DbLoadMetric:
    """DB 로드 총계 측정 데이터"""
    data_type: str  # CHARTER or MONTHLY
    total_count: int = 0
    total_chunks: int = 0
    elapsed_ms: int = 0
    start_ts: Optional[int] = None
    end_ts: Optional[int] = None
    chunks: List[ChunkMetric] = field(default_factory=list)


@dataclass
class BatchMetric:
    """배치 전체 측정 데이터"""
    elapsed_ms: int = 0
    start_ts: Optional[int] = None
    end_ts: Optional[int] = None


@dataclass
class ParseResult:
    """파싱 결과 집계"""
    chunk_size: int = 0
    batch: BatchMetric = field(default_factory=BatchMetric)
    charter_load: DbLoadMetric = field(default_factory=lambda: DbLoadMetric(data_type='CHARTER'))
    monthly_load: DbLoadMetric = field(default_factory=lambda: DbLoadMetric(data_type='MONTHLY'))
    transform_charter_ms: int = 0
    transform_monthly_ms: int = 0
    raw_lines: List[str] = field(default_factory=list)


# =============================================================================
# 정규식 패턴 정의
# =============================================================================

# CHUNK_SIZE = 10000 건/청크
PATTERN_CHUNK_SIZE = re.compile(r'CHUNK_SIZE\s*=\s*(\d+)\s*건/청크')

# [PERF:BATCH:TOTAL] ... phase=START|END ... ts=... elapsed_ms=...
PATTERN_BATCH = re.compile(
    r'\[PERF:BATCH:TOTAL\].*thread=(?P<thread>[\w-]+).*phase=(?P<phase>START|END).*ts=(?P<ts>\d+)'
    r'(?:.*elapsed_ms=(?P<elapsed_ms>\d+))?'
)

# [PERF:DBLOAD:CHARTER|MONTHLY] ... phase=START ... chunkSize=10000
PATTERN_DBLOAD_START = re.compile(
    r'\[PERF:DBLOAD:(?P<type>CHARTER|MONTHLY)\].*phase=START.*ts=(?P<ts>\d+).*chunkSize=(?P<chunk_size>\d+)'
)

# [PERF:DBLOAD:CHARTER|MONTHLY] ... phase=END ... totalCount=58660 | totalChunks=6 | elapsed_ms=1743
PATTERN_DBLOAD_END = re.compile(
    r'\[PERF:DBLOAD:(?P<type>CHARTER|MONTHLY)\].*phase=END.*ts=(?P<ts>\d+).*'
    r'totalCount=(?P<total_count>\d+).*totalChunks=(?P<total_chunks>\d+).*elapsed_ms=(?P<elapsed_ms>\d+)'
)

# [PERF:CHUNK:CHARTER|MONTHLY] ... phase=COMPLETE | chunkIndex=0 | chunkSize=10000 | cumulative=10000 | load_ms=248 | transform_ms=5 | total_ms=253 | hasNext=true
PATTERN_CHUNK = re.compile(
    r'\[PERF:CHUNK:(?P<type>CHARTER|MONTHLY)\].*phase=COMPLETE.*'
    r'chunkIndex=(?P<chunk_index>\d+).*chunkSize=(?P<chunk_size>\d+).*cumulative=(?P<cumulative>\d+).*'
    r'load_ms=(?P<load_ms>\d+).*transform_ms=(?P<transform_ms>\d+).*total_ms=(?P<total_ms>\d+).*hasNext=(?P<has_next>true|false)'
)

# TRANSFORM:CHARTER = 26 ms (누적)
PATTERN_TRANSFORM = re.compile(r'TRANSFORM:(?P<type>CHARTER|MONTHLY)\s*=\s*(?P<ms>\d+)\s*ms')


# =============================================================================
# 파서 클래스
# =============================================================================

class PerfLogParser:
    """2차 테스트 성능 로그 파서"""
    
    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self.result = ParseResult()
    
    def parse(self) -> ParseResult:
        if not self.filepath.exists():
            raise FileNotFoundError(f"로그 파일 없음: {self.filepath}")
        
        with open(self.filepath, 'r', encoding='utf-8') as f:
            for line in f:
                self._parse_line(line.strip())
        
        return self.result
    
    def _parse_line(self, line: str) -> None:
        # CHUNK_SIZE
        match = PATTERN_CHUNK_SIZE.search(line)
        if match:
            self.result.chunk_size = int(match.group(1))
            self.result.raw_lines.append(line)
            return
        
        # BATCH:TOTAL
        match = PATTERN_BATCH.search(line)
        if match:
            self.result.raw_lines.append(line)
            phase = match.group('phase')
            ts = int(match.group('ts'))
            if phase == 'START':
                self.result.batch.start_ts = ts
            elif phase == 'END':
                self.result.batch.end_ts = ts
                if match.group('elapsed_ms'):
                    self.result.batch.elapsed_ms = int(match.group('elapsed_ms'))
            return
        
        # DBLOAD START
        match = PATTERN_DBLOAD_START.search(line)
        if match:
            self.result.raw_lines.append(line)
            data_type = match.group('type')
            ts = int(match.group('ts'))
            if data_type == 'CHARTER':
                self.result.charter_load.start_ts = ts
            else:
                self.result.monthly_load.start_ts = ts
            return
        
        # DBLOAD END
        match = PATTERN_DBLOAD_END.search(line)
        if match:
            self.result.raw_lines.append(line)
            data_type = match.group('type')
            load = self.result.charter_load if data_type == 'CHARTER' else self.result.monthly_load
            load.end_ts = int(match.group('ts'))
            load.total_count = int(match.group('total_count'))
            load.total_chunks = int(match.group('total_chunks'))
            load.elapsed_ms = int(match.group('elapsed_ms'))
            return
        
        # CHUNK
        match = PATTERN_CHUNK.search(line)
        if match:
            self.result.raw_lines.append(line)
            chunk = ChunkMetric(
                chunk_index=int(match.group('chunk_index')),
                chunk_size=int(match.group('chunk_size')),
                cumulative=int(match.group('cumulative')),
                load_ms=int(match.group('load_ms')),
                transform_ms=int(match.group('transform_ms')),
                total_ms=int(match.group('total_ms')),
                has_next=match.group('has_next') == 'true'
            )
            if match.group('type') == 'CHARTER':
                self.result.charter_load.chunks.append(chunk)
            else:
                self.result.monthly_load.chunks.append(chunk)
            return
        
        # TRANSFORM
        match = PATTERN_TRANSFORM.search(line)
        if match:
            self.result.raw_lines.append(line)
            if match.group('type') == 'CHARTER':
                self.result.transform_charter_ms = int(match.group('ms'))
            else:
                self.result.transform_monthly_ms = int(match.group('ms'))
            return


# =============================================================================
# 통계 계산 함수
# =============================================================================

def calculate_chunk_stats(chunks: List[ChunkMetric]) -> Dict:
    """청크 리스트에서 통계 계산"""
    if not chunks:
        return {}
    
    load_times = [c.load_ms for c in chunks]
    transform_times = [c.transform_ms for c in chunks]
    total_times = [c.total_ms for c in chunks]
    
    def calc_stats(values):
        return {
            'min': min(values),
            'max': max(values),
            'avg': round(mean(values), 2),
            'stdev': round(stdev(values), 2) if len(values) > 1 else 0
        }
    
    return {
        'count': len(chunks),
        'load_ms': calc_stats(load_times),
        'transform_ms': calc_stats(transform_times),
        'total_ms': calc_stats(total_times)
    }


# =============================================================================
# 콘솔 리포트 출력
# =============================================================================

def print_console_report(result: ParseResult):
    """콘솔에 분석 결과 출력"""
    
    print("=" * 80)
    print("Wherehouse 프로젝트 - OOM 병목 2차 테스트 분석 결과")
    print("Slice 기반 청크 처리 성능 측정")
    print("=" * 80)
    print()
    
    # 1. 기본 설정
    print("[1] 테스트 설정")
    print("-" * 40)
    print(f"  CHUNK_SIZE: {result.chunk_size:,} 건/청크")
    print()
    
    # 2. 전체 요약
    print("[2] 측정 결과 요약")
    print("-" * 40)
    print(f"  DBLOAD:CHARTER    = {result.charter_load.elapsed_ms:,} ms ({result.charter_load.total_count:,}건, {result.charter_load.total_chunks}청크)")
    print(f"  DBLOAD:MONTHLY    = {result.monthly_load.elapsed_ms:,} ms ({result.monthly_load.total_count:,}건, {result.monthly_load.total_chunks}청크)")
    print(f"  TRANSFORM:CHARTER = {result.transform_charter_ms} ms (누적)")
    print(f"  TRANSFORM:MONTHLY = {result.transform_monthly_ms} ms (누적)")
    print(f"  BATCH:TOTAL       = {result.batch.elapsed_ms:,} ms")
    print()
    
    # 3. 청크별 상세 - 전세
    if result.charter_load.chunks:
        print("[3] 전세(CHARTER) 청크별 상세")
        print("-" * 40)
        print(f"  {'Index':<6} {'Size':<8} {'Cumul':<10} {'Load(ms)':<10} {'Trans(ms)':<10} {'Total(ms)':<10}")
        print(f"  {'-'*6} {'-'*8} {'-'*10} {'-'*10} {'-'*10} {'-'*10}")
        for chunk in result.charter_load.chunks:
            print(f"  {chunk.chunk_index:<6} {chunk.chunk_size:<8,} {chunk.cumulative:<10,} {chunk.load_ms:<10} {chunk.transform_ms:<10} {chunk.total_ms:<10}")
        
        stats = calculate_chunk_stats(result.charter_load.chunks)
        print()
        print(f"  [통계] Load(ms):      min={stats['load_ms']['min']}, max={stats['load_ms']['max']}, avg={stats['load_ms']['avg']}, stdev={stats['load_ms']['stdev']}")
        print(f"  [통계] Transform(ms): min={stats['transform_ms']['min']}, max={stats['transform_ms']['max']}, avg={stats['transform_ms']['avg']}, stdev={stats['transform_ms']['stdev']}")
        print(f"  [통계] Total(ms):     min={stats['total_ms']['min']}, max={stats['total_ms']['max']}, avg={stats['total_ms']['avg']}, stdev={stats['total_ms']['stdev']}")
        print()
    
    # 4. 청크별 상세 - 월세
    if result.monthly_load.chunks:
        print("[4] 월세(MONTHLY) 청크별 상세")
        print("-" * 40)
        print(f"  {'Index':<6} {'Size':<8} {'Cumul':<10} {'Load(ms)':<10} {'Trans(ms)':<10} {'Total(ms)':<10}")
        print(f"  {'-'*6} {'-'*8} {'-'*10} {'-'*10} {'-'*10} {'-'*10}")
        for chunk in result.monthly_load.chunks:
            print(f"  {chunk.chunk_index:<6} {chunk.chunk_size:<8,} {chunk.cumulative:<10,} {chunk.load_ms:<10} {chunk.transform_ms:<10} {chunk.total_ms:<10}")
        
        stats = calculate_chunk_stats(result.monthly_load.chunks)
        print()
        print(f"  [통계] Load(ms):      min={stats['load_ms']['min']}, max={stats['load_ms']['max']}, avg={stats['load_ms']['avg']}, stdev={stats['load_ms']['stdev']}")
        print(f"  [통계] Transform(ms): min={stats['transform_ms']['min']}, max={stats['transform_ms']['max']}, avg={stats['transform_ms']['avg']}, stdev={stats['transform_ms']['stdev']}")
        print(f"  [통계] Total(ms):     min={stats['total_ms']['min']}, max={stats['total_ms']['max']}, avg={stats['total_ms']['avg']}, stdev={stats['total_ms']['stdev']}")
        print()
    
    # 5. 1차 테스트 대비 비교
    print("[5] 1차 테스트 대비 비교")
    print("-" * 40)
    print("  ※ 1차 테스트 결과 (findAll 방식):")
    print("     - DBLOAD:CHARTER    = 851 ms (58,660건)")
    print("     - DBLOAD:MONTHLY    = 876 ms (56,272건)")
    print("     - TRANSFORM:CHARTER = 17 ms")
    print("     - TRANSFORM:MONTHLY = 12 ms")
    print("     - BATCH:TOTAL       = 7,248 ms")
    print("     - 힙 피크 점유율    = 64.4% (94.6MB / 256MB)")
    print()
    
    phase1_dbload_charter = 851
    phase1_dbload_monthly = 876
    phase1_batch_total = 7248
    
    charter_diff = result.charter_load.elapsed_ms - phase1_dbload_charter
    monthly_diff = result.monthly_load.elapsed_ms - phase1_dbload_monthly
    batch_diff = result.batch.elapsed_ms - phase1_batch_total
    
    charter_pct = (charter_diff / phase1_dbload_charter) * 100
    monthly_pct = (monthly_diff / phase1_dbload_monthly) * 100
    batch_pct = (batch_diff / phase1_batch_total) * 100
    
    print("  ※ 2차 테스트 대비 변화:")
    print(f"     - DBLOAD:CHARTER: {phase1_dbload_charter} → {result.charter_load.elapsed_ms} ms ({charter_diff:+} ms, {charter_pct:+.1f}%)")
    print(f"     - DBLOAD:MONTHLY: {phase1_dbload_monthly} → {result.monthly_load.elapsed_ms} ms ({monthly_diff:+} ms, {monthly_pct:+.1f}%)")
    print(f"     - BATCH:TOTAL:    {phase1_batch_total} → {result.batch.elapsed_ms} ms ({batch_diff:+} ms, {batch_pct:+.1f}%)")
    print()
    print("  ※ 트레이드오프 분석:")
    print(f"     - DB 로드 시간 증가: 페이징 쿼리 오버헤드 (OFFSET 증가에 따른 성능 저하)")
    if result.chunk_size > 0 and result.charter_load.total_count > 0:
        estimated_peak = result.chunk_size / result.charter_load.total_count * 64.4
        print(f"     - 예상 힙 피크 감소: 64.4% → ~{estimated_peak:.1f}% (청크 크기 기준)")
    print(f"     - OOM 안전성: 데이터 증가 시에도 힙 피크 일정 유지")
    print()
    print("=" * 80)


# =============================================================================
# Excel 리포트 생성기
# =============================================================================

class ExcelReportGenerator:
    """Excel 리포트 생성기"""
    
    # 구간 코드 설명 정의
    METRIC_DESCRIPTIONS = {
        "DBLOAD:CHARTER": "전세 데이터 DB 로드 (Slice 페이징)",
        "DBLOAD:MONTHLY": "월세 데이터 DB 로드 (Slice 페이징)",
        "TRANSFORM:CHARTER": "전세 Entity → Property DTO 변환 (누적)",
        "TRANSFORM:MONTHLY": "월세 Entity → Property DTO 변환 (누적)",
        "CHUNK:CHARTER": "전세 청크별 처리 (로드 + 변환)",
        "CHUNK:MONTHLY": "월세 청크별 처리 (로드 + 변환)",
        "BATCH:TOTAL": "배치 프로세스 전체 실행 시간",
    }
    
    # 1차 테스트 기준값
    PHASE1_BASELINE = {
        'dbload_charter_ms': 851,
        'dbload_monthly_ms': 876,
        'transform_charter_ms': 17,
        'transform_monthly_ms': 12,
        'batch_total_ms': 7248,
        'heap_peak_pct': 64.4,
        'heap_peak_mb': 94.6,
        'heap_max_mb': 256,
        'charter_count': 58660,
        'monthly_count': 56272,
    }
    
    def __init__(self, result: ParseResult):
        self.result = result
        self.wb = Workbook()
        
        # 스타일 정의
        self.header_font = Font(bold=True, size=11)
        self.title_font = Font(bold=True, size=14)
        self.header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.desc_header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
    
    def generate(self, output_path: str) -> None:
        # 기본 시트 제거
        self.wb.remove(self.wb.active)
        
        # 시트별 데이터 생성
        self._create_summary_sheet()
        self._create_duration_sheet()
        self._create_ratio_sheet()
        self._create_count_sheet()
        self._create_chunk_detail_sheet('CHARTER')
        self._create_chunk_detail_sheet('MONTHLY')
        self._create_chunk_stats_sheet()
        self._create_comparison_sheet()
        
        self.wb.save(output_path)
        print(f"[완료] Excel 파일 생성: {output_path}")
    
    def _apply_header_style(self, ws, row: int, col_count: int) -> None:
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_align
    
    def _apply_cell_border(self, ws, row: int, col_count: int) -> None:
        for col in range(1, col_count + 1):
            ws.cell(row=row, column=col).border = self.border
    
    def _auto_column_width(self, ws) -> None:
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column_cells:
                try:
                    if cell.value and not isinstance(cell, type(None)):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max(max_length + 2, 12)
    
    def _add_description_table(self, ws, start_row: int) -> None:
        """구간 코드 설명 테이블 추가"""
        start_row += 2
        
        ws.cell(row=start_row, column=1, value="[구간 코드 설명]")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        start_row += 1
        
        ws.cell(row=start_row, column=1, value="구간 코드")
        ws.cell(row=start_row, column=2, value="설명")
        for col in range(1, 3):
            cell = ws.cell(row=start_row, column=col)
            cell.font = self.header_font
            cell.fill = self.desc_header_fill
            cell.border = self.border
            cell.alignment = self.center_align
        start_row += 1
        
        for code, desc in self.METRIC_DESCRIPTIONS.items():
            ws.cell(row=start_row, column=1, value=code)
            ws.cell(row=start_row, column=2, value=desc)
            for col in range(1, 3):
                ws.cell(row=start_row, column=col).border = self.border
            start_row += 1
    
    def _create_summary_sheet(self) -> None:
        """요약 시트"""
        ws = self.wb.create_sheet("요약")
        
        ws['A1'] = "Wherehouse OOM 병목 측정 결과 (2차 테스트 - Slice 청크 처리)"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        
        ws['A3'] = f"CHUNK_SIZE: {self.result.chunk_size:,} 건/청크"
        ws['A3'].font = Font(bold=True)
        
        headers = ["구간", "소요시간(ms)", "소요시간(sec)", "건수", "청크 수"]
        ws.append([])
        ws.append(headers)
        self._apply_header_style(ws, 5, 5)
        
        data = [
            ("DBLOAD:CHARTER", self.result.charter_load.elapsed_ms, 
             round(self.result.charter_load.elapsed_ms / 1000, 2),
             self.result.charter_load.total_count, self.result.charter_load.total_chunks),
            ("DBLOAD:MONTHLY", self.result.monthly_load.elapsed_ms,
             round(self.result.monthly_load.elapsed_ms / 1000, 2),
             self.result.monthly_load.total_count, self.result.monthly_load.total_chunks),
            ("TRANSFORM:CHARTER", self.result.transform_charter_ms,
             round(self.result.transform_charter_ms / 1000, 3), "-", "-"),
            ("TRANSFORM:MONTHLY", self.result.transform_monthly_ms,
             round(self.result.transform_monthly_ms / 1000, 3), "-", "-"),
            ("BATCH:TOTAL", self.result.batch.elapsed_ms,
             round(self.result.batch.elapsed_ms / 1000, 2), "-", "-"),
        ]
        
        row = 6
        for item in data:
            ws.append(item)
            self._apply_cell_border(ws, row, 5)
            row += 1
        
        self._add_description_table(ws, row)
        self._auto_column_width(ws)
    
    def _create_duration_sheet(self) -> None:
        """구간별 소요시간 시트"""
        ws = self.wb.create_sheet("구간별_소요시간")
        
        headers = ["구간", "카테고리", "항목", "소요시간(ms)", "소요시간(sec)", "소요시간(min)", "건수", "청크 수"]
        ws.append(headers)
        self._apply_header_style(ws, 1, 8)
        
        data = [
            ("DBLOAD:CHARTER", "DBLOAD", "CHARTER", self.result.charter_load.elapsed_ms,
             round(self.result.charter_load.elapsed_ms / 1000, 2),
             round(self.result.charter_load.elapsed_ms / 60000, 3),
             self.result.charter_load.total_count, self.result.charter_load.total_chunks),
            ("DBLOAD:MONTHLY", "DBLOAD", "MONTHLY", self.result.monthly_load.elapsed_ms,
             round(self.result.monthly_load.elapsed_ms / 1000, 2),
             round(self.result.monthly_load.elapsed_ms / 60000, 3),
             self.result.monthly_load.total_count, self.result.monthly_load.total_chunks),
            ("TRANSFORM:CHARTER", "TRANSFORM", "CHARTER", self.result.transform_charter_ms,
             round(self.result.transform_charter_ms / 1000, 3),
             round(self.result.transform_charter_ms / 60000, 4), None, None),
            ("TRANSFORM:MONTHLY", "TRANSFORM", "MONTHLY", self.result.transform_monthly_ms,
             round(self.result.transform_monthly_ms / 1000, 3),
             round(self.result.transform_monthly_ms / 60000, 4), None, None),
            ("BATCH:TOTAL", "BATCH", "TOTAL", self.result.batch.elapsed_ms,
             round(self.result.batch.elapsed_ms / 1000, 2),
             round(self.result.batch.elapsed_ms / 60000, 2), None, None),
        ]
        
        row = 2
        for item in data:
            ws.append(item)
            self._apply_cell_border(ws, row, 8)
            row += 1
        
        self._add_description_table(ws, row)
        self._auto_column_width(ws)
    
    def _create_ratio_sheet(self) -> None:
        """비율 분석 시트"""
        ws = self.wb.create_sheet("비율_분석")
        
        headers = ["단계", "소요시간(ms)", "비율(%)"]
        ws.append(headers)
        self._apply_header_style(ws, 1, 3)
        
        total_dbload = self.result.charter_load.elapsed_ms + self.result.monthly_load.elapsed_ms
        total_transform = self.result.transform_charter_ms + self.result.transform_monthly_ms
        total = self.result.batch.elapsed_ms if self.result.batch.elapsed_ms else 1
        other_time = total - total_dbload - total_transform
        
        data = [
            ("DB 로드 (DBLOAD)", total_dbload, round(total_dbload / total * 100, 2)),
            ("  - 전세 (CHARTER)", self.result.charter_load.elapsed_ms, 
             round(self.result.charter_load.elapsed_ms / total * 100, 2)),
            ("  - 월세 (MONTHLY)", self.result.monthly_load.elapsed_ms,
             round(self.result.monthly_load.elapsed_ms / total * 100, 2)),
            ("Entity→DTO 변환 (TRANSFORM)", total_transform, round(total_transform / total * 100, 2)),
            ("  - 전세 (CHARTER)", self.result.transform_charter_ms,
             round(self.result.transform_charter_ms / total * 100, 2)),
            ("  - 월세 (MONTHLY)", self.result.transform_monthly_ms,
             round(self.result.transform_monthly_ms / total * 100, 2)),
            ("기타 처리 (Redis 등)", other_time, round(other_time / total * 100, 2)),
            ("배치 전체 (BATCH:TOTAL)", total, 100.0),
        ]
        
        row = 2
        for item in data:
            ws.append(item)
            self._apply_cell_border(ws, row, 3)
            row += 1
        
        self._auto_column_width(ws)
    
    def _create_count_sheet(self) -> None:
        """데이터 건수 시트"""
        ws = self.wb.create_sheet("데이터_건수")
        
        headers = ["데이터 유형", "건수", "청크 수", "청크 크기"]
        ws.append(headers)
        self._apply_header_style(ws, 1, 4)
        
        charter_count = self.result.charter_load.total_count
        monthly_count = self.result.monthly_load.total_count
        total_count = charter_count + monthly_count
        total_chunks = self.result.charter_load.total_chunks + self.result.monthly_load.total_chunks
        
        data = [
            ("전세 (CHARTER)", charter_count, self.result.charter_load.total_chunks, self.result.chunk_size),
            ("월세 (MONTHLY)", monthly_count, self.result.monthly_load.total_chunks, self.result.chunk_size),
            ("총 매물 수", total_count, total_chunks, "-"),
        ]
        
        row = 2
        for item in data:
            ws.append(item)
            self._apply_cell_border(ws, row, 4)
            row += 1
        
        self._auto_column_width(ws)
    
    def _create_chunk_detail_sheet(self, data_type: str) -> None:
        """청크별 상세 시트"""
        ws = self.wb.create_sheet(f"청크별_상세_{data_type}")
        
        load = self.result.charter_load if data_type == 'CHARTER' else self.result.monthly_load
        type_name = "전세" if data_type == 'CHARTER' else "월세"
        
        ws['A1'] = f"{type_name}({data_type}) 청크별 처리 상세"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:G1')
        
        headers = ["청크 인덱스", "청크 크기", "누적 건수", "로드(ms)", "변환(ms)", "총계(ms)", "hasNext"]
        ws.append([])
        ws.append(headers)
        self._apply_header_style(ws, 3, 7)
        
        row = 4
        for chunk in load.chunks:
            ws.append([
                chunk.chunk_index,
                chunk.chunk_size,
                chunk.cumulative,
                chunk.load_ms,
                chunk.transform_ms,
                chunk.total_ms,
                str(chunk.has_next)
            ])
            self._apply_cell_border(ws, row, 7)
            row += 1
        
        self._auto_column_width(ws)
    
    def _create_chunk_stats_sheet(self) -> None:
        """청크 통계 시트"""
        ws = self.wb.create_sheet("청크_통계")
        
        ws['A1'] = "청크별 처리 시간 통계"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        headers = ["데이터 유형", "측정 항목", "최소(ms)", "최대(ms)", "평균(ms)", "표준편차(ms)"]
        ws.append([])
        ws.append(headers)
        self._apply_header_style(ws, 3, 6)
        
        row = 4
        for data_type, load in [('CHARTER', self.result.charter_load), ('MONTHLY', self.result.monthly_load)]:
            if not load.chunks:
                continue
            
            stats = calculate_chunk_stats(load.chunks)
            type_name = "전세" if data_type == 'CHARTER' else "월세"
            
            for metric_name, metric_key in [('로드', 'load_ms'), ('변환', 'transform_ms'), ('총계', 'total_ms')]:
                s = stats[metric_key]
                ws.append([
                    f"{type_name}({data_type})",
                    metric_name,
                    s['min'],
                    s['max'],
                    s['avg'],
                    s['stdev']
                ])
                self._apply_cell_border(ws, row, 6)
                row += 1
        
        self._auto_column_width(ws)
    
    def _create_comparison_sheet(self) -> None:
        """1차 테스트 대비 비교 시트"""
        ws = self.wb.create_sheet("1차_대비_비교")
        
        ws['A1'] = "1차 테스트 (findAll) vs 2차 테스트 (Slice 청크 처리) 비교"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        headers = ["측정 항목", "1차 테스트", "2차 테스트", "차이(ms)", "변화율(%)", "비고"]
        ws.append([])
        ws.append(headers)
        self._apply_header_style(ws, 3, 6)
        
        p1 = self.PHASE1_BASELINE
        
        comparisons = [
            ("DBLOAD:CHARTER (ms)", p1['dbload_charter_ms'], self.result.charter_load.elapsed_ms),
            ("DBLOAD:MONTHLY (ms)", p1['dbload_monthly_ms'], self.result.monthly_load.elapsed_ms),
            ("TRANSFORM:CHARTER (ms)", p1['transform_charter_ms'], self.result.transform_charter_ms),
            ("TRANSFORM:MONTHLY (ms)", p1['transform_monthly_ms'], self.result.transform_monthly_ms),
            ("BATCH:TOTAL (ms)", p1['batch_total_ms'], self.result.batch.elapsed_ms),
        ]
        
        row = 4
        for item, phase1_val, phase2_val in comparisons:
            diff = phase2_val - phase1_val
            pct = round((diff / phase1_val) * 100, 1) if phase1_val else 0
            
            note = ""
            if "DBLOAD" in item:
                note = "페이징 오버헤드" if diff > 0 else "개선"
            elif "BATCH" in item:
                note = "총 처리 시간 증가" if diff > 0 else "개선"
            
            ws.append([item, phase1_val, phase2_val, diff, pct, note])
            self._apply_cell_border(ws, row, 6)
            
            # 증가한 항목 강조
            if diff > 0:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = self.warning_fill
            
            row += 1
        
        # 힙 메모리 비교 (예상치)
        row += 1
        ws.cell(row=row, column=1, value="[메모리 사용량 비교 (예상)]")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        headers2 = ["측정 항목", "1차 테스트", "2차 테스트 (예상)", "감소율(%)"]
        for col, header in enumerate(headers2, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.highlight_fill
            cell.border = self.border
        row += 1
        
        estimated_peak = round(self.result.chunk_size / p1['charter_count'] * p1['heap_peak_pct'], 1) if self.result.chunk_size else 0
        estimated_mb = round(self.result.chunk_size / p1['charter_count'] * p1['heap_peak_mb'], 1) if self.result.chunk_size else 0
        reduction = round((p1['heap_peak_pct'] - estimated_peak) / p1['heap_peak_pct'] * 100, 1)
        
        memory_data = [
            ("힙 피크 점유율 (%)", f"{p1['heap_peak_pct']}%", f"~{estimated_peak}%", f"-{reduction}%"),
            ("힙 피크 사용량 (MB)", f"{p1['heap_peak_mb']}MB", f"~{estimated_mb}MB", f"-{reduction}%"),
        ]
        
        for item in memory_data:
            ws.append(item)
            self._apply_cell_border(ws, row, 4)
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = self.highlight_fill
            row += 1
        
        # 트레이드오프 분석
        row += 1
        ws.cell(row=row, column=1, value="[트레이드오프 분석]")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        tradeoffs = [
            "• DB 로드 시간 증가: Oracle OFFSET 페이징의 구조적 오버헤드 (OFFSET 증가에 따른 스캔 범위 증가)",
            "• 힙 피크 메모리 감소: 청크 단위 처리로 동시 적재 Entity 수 제한",
            "• OOM 안전성 확보: 데이터 증가 시에도 힙 피크 일정 유지 (청크 크기에 비례)",
            f"• 현재 청크 크기: {self.result.chunk_size:,}건 → Entity 최대 ~{self.result.chunk_size:,}건만 힙에 동시 적재",
        ]
        
        for text in tradeoffs:
            ws.cell(row=row, column=1, value=text)
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
        
        self._auto_column_width(ws)


# =============================================================================
# 메인 실행
# =============================================================================

def main():
    if len(sys.argv) < 2:
        print("사용법: python perf_log_parser_phase2.py <로그파일> [--xlsx [출력파일.xlsx]]")
        print("예시:")
        print("  python perf_log_parser_phase2.py wherehouse.log")
        print("  python perf_log_parser_phase2.py wherehouse.log --xlsx")
        print("  python perf_log_parser_phase2.py wherehouse.log --xlsx output.xlsx")
        sys.exit(1)
    
    filepath = sys.argv[1]
    
    # --xlsx 옵션 파싱
    xlsx_mode = '--xlsx' in sys.argv
    xlsx_output = None
    if xlsx_mode:
        xlsx_idx = sys.argv.index('--xlsx')
        if xlsx_idx + 1 < len(sys.argv) and not sys.argv[xlsx_idx + 1].startswith('-'):
            xlsx_output = sys.argv[xlsx_idx + 1]
        else:
            xlsx_output = "perf_result_phase2.xlsx"
        
        if not xlsx_output.endswith('.xlsx'):
            xlsx_output += '.xlsx'
    
    try:
        parser = PerfLogParser(filepath)
        result = parser.parse()
        
        if not result.raw_lines:
            print(f"[경고] PERF 측정 데이터 없음: {filepath}")
            sys.exit(1)
        
        # 콘솔 출력 (항상 실행)
        print_console_report(result)
        
        # Excel 출력 (옵션)
        if xlsx_mode:
            if not OPENPYXL_AVAILABLE:
                print("[오류] Excel 출력을 위해 openpyxl 패키지 필요: pip install openpyxl")
                sys.exit(1)
            
            generator = ExcelReportGenerator(result)
            generator.generate(xlsx_output)
        
    except FileNotFoundError as e:
        print(f"[오류] {e}")
        sys.exit(1)
    except Exception as e:
        print(f"[오류] 파싱 중 예외 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
