#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Wherehouse OOM 병목 측정 로그 파서
- 명세서: OOM_병목_측정_명세서_v1_1.docx 섹션 4.1 로그 포맷 기반
- 출력: Excel 파일 (데이터 유형별 시트 분리)

사용법:
    python perf_log_parser.py <로그파일경로>
    python perf_log_parser.py wherehouse.log
    python perf_log_parser.py wherehouse.log output.xlsx
"""

import re
import sys
from dataclasses import dataclass, field
from typing import Dict, List, Optional
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[오류] openpyxl 패키지 필요: pip install openpyxl")
    sys.exit(1)


@dataclass
class PerfMetric:
    """개별 측정 구간 데이터"""
    category: str
    item: str
    thread: str
    start_ts: Optional[int] = None
    end_ts: Optional[int] = None
    elapsed_ms: Optional[int] = None
    count: Optional[int] = None


@dataclass
class ParseResult:
    """파싱 결과 집계"""
    metrics: Dict[str, PerfMetric] = field(default_factory=dict)
    raw_lines: List[str] = field(default_factory=list)
    
    def get_key(self, category: str, item: str) -> str:
        return f"{category}:{item}"


class PerfLogParser:
    """성능 로그 파서"""
    
    PERF_PATTERN = re.compile(
        r'\[PERF:(?P<category>\w+):(?P<item>\w+)\]\s+'
        r'thread=(?P<thread>[\w-]+)\s*\|\s*'
        r'phase=(?P<phase>START|END)\s*\|\s*'
        r'ts=(?P<ts>\d+)'
        r'(?:\s*\|\s*count=(?P<count>\d+))?'
        r'(?:\s*\|\s*elapsed_ms=(?P<elapsed_ms>\d+))?'
    )
    
    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self.result = ParseResult()
    
    def parse(self) -> ParseResult:
        if not self.filepath.exists():
            raise FileNotFoundError(f"로그 파일 없음: {self.filepath}")
        
        with open(self.filepath, 'r', encoding='utf-8') as f:
            for line in f:
                self._parse_line(line.strip())
        
        return self.result
    
    def _parse_line(self, line: str) -> None:
        match = self.PERF_PATTERN.search(line)
        if not match:
            return
        
        self.result.raw_lines.append(line)
        
        category = match.group('category')
        item = match.group('item')
        thread = match.group('thread')
        phase = match.group('phase')
        ts = int(match.group('ts'))
        count = int(match.group('count')) if match.group('count') else None
        elapsed_ms = int(match.group('elapsed_ms')) if match.group('elapsed_ms') else None
        
        key = self.result.get_key(category, item)
        
        if key not in self.result.metrics:
            self.result.metrics[key] = PerfMetric(
                category=category,
                item=item,
                thread=thread
            )
        
        metric = self.result.metrics[key]
        
        if phase == 'START':
            metric.start_ts = ts
        elif phase == 'END':
            metric.end_ts = ts
            metric.elapsed_ms = elapsed_ms
            metric.count = count


class ExcelReportGenerator:
    """Excel 리포트 생성기"""
    
    # 구간 코드 설명 정의
    METRIC_DESCRIPTIONS = {
        "DBLOAD:CHARTER": "전세 데이터 DB 로드 (JPA findAll)",
        "DBLOAD:MONTHLY": "월세 데이터 DB 로드 (JPA findAll)",
        "TRANSFORM:CHARTER": "전세 Entity → Property DTO 변환",
        "TRANSFORM:MONTHLY": "월세 Entity → Property DTO 변환",
        "BATCH:TOTAL": "배치 프로세스 전체 실행 시간",
    }
    
    def __init__(self, result: ParseResult):
        self.result = result
        self.wb = Workbook()
        
        # 스타일 정의
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.desc_header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
    
    def generate(self, output_path: str) -> None:
        # 기본 시트 제거
        self.wb.remove(self.wb.active)
        
        # 시트별 데이터 생성
        self._create_summary_sheet()
        self._create_duration_sheet()
        self._create_ratio_sheet()
        self._create_count_sheet()
        self._create_performance_sheet()
        
        self.wb.save(output_path)
        print(f"[완료] Excel 파일 생성: {output_path}")
    
    def _apply_header_style(self, ws, row: int, col_count: int) -> None:
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_align
    
    def _apply_cell_border(self, ws, row: int, col_count: int) -> None:
        for col in range(1, col_count + 1):
            ws.cell(row=row, column=col).border = self.border
    
    def _auto_column_width(self, ws) -> None:
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column_cells:
                try:
                    if cell.value and not isinstance(cell, type(None)):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max(max_length + 2, 12)
    
    def _add_description_table(self, ws, start_row: int) -> None:
        """구간 코드 설명 테이블 추가"""
        # 빈 행 추가
        start_row += 2
        
        # 설명 테이블 제목
        ws.cell(row=start_row, column=1, value="[구간 코드 설명]")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        start_row += 1
        
        # 헤더
        ws.cell(row=start_row, column=1, value="구간 코드")
        ws.cell(row=start_row, column=2, value="설명")
        for col in range(1, 3):
            cell = ws.cell(row=start_row, column=col)
            cell.font = self.header_font
            cell.fill = self.desc_header_fill
            cell.border = self.border
            cell.alignment = self.center_align
        start_row += 1
        
        # 설명 데이터
        for code, desc in self.METRIC_DESCRIPTIONS.items():
            ws.cell(row=start_row, column=1, value=code)
            ws.cell(row=start_row, column=2, value=desc)
            for col in range(1, 3):
                ws.cell(row=start_row, column=col).border = self.border
            start_row += 1
    
    def _create_summary_sheet(self) -> None:
        """요약 시트"""
        ws = self.wb.create_sheet("요약")
        
        # 제목
        ws['A1'] = "Wherehouse OOM 병목 측정 결과 (1차 테스트)"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # 구간별 소요시간 테이블
        headers = ["구간", "소요시간(ms)", "소요시간(sec)", "건수"]
        ws.append([])
        ws.append(headers)
        self._apply_header_style(ws, 3, 4)
        
        order = [
            ('DBLOAD', 'CHARTER'),
            ('DBLOAD', 'MONTHLY'),
            ('TRANSFORM', 'CHARTER'),
            ('TRANSFORM', 'MONTHLY'),
            ('BATCH', 'TOTAL'),
        ]
        
        row = 4
        for category, item in order:
            key = f"{category}:{item}"
            metric = self.result.metrics.get(key)
            if metric:
                elapsed = metric.elapsed_ms or 0
                ws.append([
                    key,
                    elapsed,
                    round(elapsed / 1000, 2),
                    metric.count or "-"
                ])
                self._apply_cell_border(ws, row, 4)
                row += 1
        
        # 구간 코드 설명 테이블 추가
        self._add_description_table(ws, row)
        
        self._auto_column_width(ws)
    
    def _create_duration_sheet(self) -> None:
        """구간별 소요시간 시트"""
        ws = self.wb.create_sheet("구간별_소요시간")
        
        headers = ["구간", "카테고리", "항목", "소요시간(ms)", "소요시간(sec)", "소요시간(min)", "건수"]
        ws.append(headers)
        self._apply_header_style(ws, 1, 7)
        
        order = [
            ('DBLOAD', 'CHARTER'),
            ('DBLOAD', 'MONTHLY'),
            ('TRANSFORM', 'CHARTER'),
            ('TRANSFORM', 'MONTHLY'),
            ('BATCH', 'TOTAL'),
        ]
        
        row = 2
        for category, item in order:
            key = f"{category}:{item}"
            metric = self.result.metrics.get(key)
            if metric:
                elapsed = metric.elapsed_ms or 0
                ws.append([
                    key,
                    category,
                    item,
                    elapsed,
                    round(elapsed / 1000, 2),
                    round(elapsed / 60000, 2),
                    metric.count or None
                ])
                self._apply_cell_border(ws, row, 7)
                row += 1
        
        # 구간 코드 설명 테이블 추가
        self._add_description_table(ws, row)
        
        self._auto_column_width(ws)
    
    def _create_ratio_sheet(self) -> None:
        """비율 분석 시트"""
        ws = self.wb.create_sheet("비율_분석")
        
        headers = ["단계", "소요시간(ms)", "비율(%)"]
        ws.append(headers)
        self._apply_header_style(ws, 1, 3)
        
        # 집계 계산
        total_dbload = 0
        total_transform = 0
        
        for key, metric in self.result.metrics.items():
            if metric.category == 'DBLOAD' and metric.elapsed_ms:
                total_dbload += metric.elapsed_ms
            elif metric.category == 'TRANSFORM' and metric.elapsed_ms:
                total_transform += metric.elapsed_ms
        
        batch_total = self.result.metrics.get('BATCH:TOTAL')
        total = batch_total.elapsed_ms if batch_total and batch_total.elapsed_ms else 1
        other_time = total - total_dbload - total_transform
        
        data = [
            ("DB 로드 (DBLOAD)", total_dbload, round(total_dbload / total * 100, 2)),
            ("Entity→DTO 변환 (TRANSFORM)", total_transform, round(total_transform / total * 100, 2)),
            ("기타 처리 (Redis 등)", other_time, round(other_time / total * 100, 2)),
            ("배치 전체 (BATCH:TOTAL)", total, 100.0),
        ]
        
        row = 2
        for item in data:
            ws.append(item)
            self._apply_cell_border(ws, row, 3)
            row += 1
        
        self._auto_column_width(ws)
    
    def _create_count_sheet(self) -> None:
        """데이터 건수 시트"""
        ws = self.wb.create_sheet("데이터_건수")
        
        headers = ["데이터 유형", "건수"]
        ws.append(headers)
        self._apply_header_style(ws, 1, 2)
        
        charter = self.result.metrics.get('DBLOAD:CHARTER')
        monthly = self.result.metrics.get('DBLOAD:MONTHLY')
        
        charter_count = charter.count if charter and charter.count else 0
        monthly_count = monthly.count if monthly and monthly.count else 0
        total_count = charter_count + monthly_count
        
        data = [
            ("전세 (CHARTER)", charter_count),
            ("월세 (MONTHLY)", monthly_count),
            ("총 매물 수", total_count),
        ]
        
        row = 2
        for item in data:
            ws.append(item)
            self._apply_cell_border(ws, row, 2)
            row += 1
        
        self._auto_column_width(ws)
    
    def _create_performance_sheet(self) -> None:
        """건당 처리 성능 시트"""
        ws = self.wb.create_sheet("건당_처리성능")
        
        headers = ["데이터 유형", "총 소요시간(ms)", "건수", "건당 처리시간(ms/건)"]
        ws.append(headers)
        self._apply_header_style(ws, 1, 4)
        
        charter = self.result.metrics.get('DBLOAD:CHARTER')
        monthly = self.result.metrics.get('DBLOAD:MONTHLY')
        
        row = 2
        if charter and charter.elapsed_ms and charter.count:
            per_item = round(charter.elapsed_ms / charter.count, 4)
            ws.append(["전세 (CHARTER)", charter.elapsed_ms, charter.count, per_item])
            self._apply_cell_border(ws, row, 4)
            row += 1
        
        if monthly and monthly.elapsed_ms and monthly.count:
            per_item = round(monthly.elapsed_ms / monthly.count, 4)
            ws.append(["월세 (MONTHLY)", monthly.elapsed_ms, monthly.count, per_item])
            self._apply_cell_border(ws, row, 4)
            row += 1
        
        self._auto_column_width(ws)


def main():
    if len(sys.argv) < 2:
        print("사용법: python perf_log_parser.py <로그파일> [출력파일.xlsx]")
        sys.exit(1)
    
    filepath = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else "perf_result.xlsx"
    
    if not output_path.endswith('.xlsx'):
        output_path += '.xlsx'
    
    try:
        parser = PerfLogParser(filepath)
        result = parser.parse()
        
        if not result.metrics:
            print(f"[경고] PERF 측정 데이터 없음: {filepath}")
            sys.exit(1)
        
        generator = ExcelReportGenerator(result)
        generator.generate(output_path)
        
    except FileNotFoundError as e:
        print(f"[오류] {e}")
        sys.exit(1)
    except Exception as e:
        print(f"[오류] 파싱 중 예외 발생: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
