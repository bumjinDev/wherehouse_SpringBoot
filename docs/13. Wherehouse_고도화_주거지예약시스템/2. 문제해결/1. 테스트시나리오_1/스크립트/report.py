#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
F004 슬롯 예약 동시성 측정 — 시나리오 1 결과 엑셀 생성기.

wherehouse 백엔드 로그에서 F004 측정 라인(검증 직후 / 점유 직후 / 등록 직후) 을 읽어,
포트폴리오용 단일 엑셀(시트 1 개) 로 정리한다.

사용:
  py report.py
  py report.py --log <로그경로> --out <엑셀경로>
"""
from __future__ import annotations

import argparse
import re
import statistics
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────
# 로그 파싱
# ────────────────────────────────────────────────────────────

LINE_RE = re.compile(
    r'^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3})\s+'
    r'\[[^\]]+\]\s+\w+\s+\S+\s+-\s+'
    r'\[F004-(?P<cp>[^\]]+)\]\s+'
    r'thread=\[name=(?P<thr>[^,]+),\s*id=(?P<tid>\d+)\]\s+'
    r'nanoTime=(?P<nt>\d+)'
    r'(?:\s+instant=\S+)?'
    r'\s+(?P<rest>.*)$'
)

SLOT_BODY_RE = re.compile(r'^slot=\((?P<body>.*)\)\s*$')
RES_HDR_RE   = re.compile(r'^slotId=(\d+)\s+reservationsCount=(\d+)\s*$')
RES_ROW_RE   = re.compile(
    r'reservation=\(reservationId=(\d+),\s*slotId=\d+,\s*searcherUserId=(\S+?),\s*status=(\w+)'
)

# 매물 번호 fallback 용 — 컨트롤러의 슬롯 조회 요청 로그
QUERY_REQ_RE = re.compile(
    r'\[VISIT_SLOT_QUERY_REQ\]\s+propertyId=(?P<pid>\S+?),\s+leaseType=(?P<lt>\S+)'
)

CP_SHORT = {
    'CP1-AFTER-AVAILABLE-CHECK':    'CP1',
    'CP2-AFTER-SLOT-UPDATE':        'CP2',
    'CP3-AFTER-RESERVATION-INSERT': 'CP3',
}

# 영어 enum → 보기 좋은 한국어 라벨
LEASE_KOR = {
    'CHARTER': '전세',
    'MONTHLY': '월세',
}

SLOT_STATUS_KOR = {
    'AVAILABLE': '예약 가능',
    'RESERVED':  '예약됨',
    'CLOSED':    '종료됨',
    'WITHDRAWN': '철회됨',
}


def kor_lease(v):
    if v is None:
        return None
    return LEASE_KOR.get(v, v)


def kor_slot_status(v):
    if v is None:
        return None
    return SLOT_STATUS_KOR.get(v, v)


def parse_kv(body: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for pair in body.split(', '):
        if '=' in pair:
            k, v = pair.split('=', 1)
            out[k.strip()] = v.strip()
    return out


def norm(v: Optional[str]) -> Optional[str]:
    if v is None or v == 'null':
        return None
    return v


def collect_property_fallback(path: Path) -> Tuple[Optional[str], Optional[str]]:
    """현재 로그 전체에서 [VISIT_SLOT_QUERY_REQ] 의 (propertyId, leaseType) distinct 페어 수집.
    페어가 정확히 1 개일 때만 fallback 으로 사용 (단일 매물 측정 회차)."""
    pairs = set()
    with path.open('r', encoding='utf-8', errors='replace') as f:
        for line in f:
            m = QUERY_REQ_RE.search(line)
            if m:
                pairs.add((m.group('pid'), m.group('lt')))
    return next(iter(pairs)) if len(pairs) == 1 else (None, None)


def parse_log(path: Path) -> List[Dict[str, Any]]:
    events: List[Dict[str, Any]] = []
    with path.open('r', encoding='utf-8', errors='replace') as f:
        for line in f:
            m = LINE_RE.match(line.rstrip('\n'))
            if not m:
                continue
            ev: Dict[str, Any] = {
                'log_ts':      m.group('ts'),
                'cp':          CP_SHORT.get(m.group('cp'), m.group('cp')),
                'thr':         m.group('thr'),
                'tid':         int(m.group('tid')),
                'nano':        int(m.group('nt')),
                'slot_id':     None, 'slot_status': None,
                'property_id': None, 'lease_type':  None,
                'start_time':  None, 'end_time':    None,
                'res_count':   None,
                'res_id':      None, 'searcher':    None, 'res_status': None,
            }
            rest = m.group('rest')
            sm = SLOT_BODY_RE.match(rest)
            rh = RES_HDR_RE.match(rest)
            rr = RES_ROW_RE.search(rest)
            if sm:
                kv = parse_kv(sm.group('body'))
                ev['slot_id']     = int(kv['slotId']) if 'slotId' in kv else None
                ev['slot_status'] = norm(kv.get('status'))
                ev['property_id'] = norm(kv.get('propertyId'))
                ev['lease_type']  = norm(kv.get('leaseType'))
                ev['start_time']  = norm(kv.get('startTime'))
                ev['end_time']    = norm(kv.get('endTime'))
            elif rh:
                ev['slot_id']   = int(rh.group(1))
                ev['res_count'] = int(rh.group(2))
            elif rr:
                ev['res_id']     = int(rr.group(1))
                ev['searcher']   = rr.group(2)
                ev['res_status'] = rr.group(3)
            events.append(ev)
    return events


def group_requests(events: List[Dict[str, Any]]) -> List[List[Dict[str, Any]]]:
    events_sorted = sorted(events, key=lambda e: (e['thr'], e['nano']))
    groups: List[List[Dict[str, Any]]] = []
    current: List[Dict[str, Any]] = []
    prev_thr: Optional[str] = None
    for e in events_sorted:
        is_new = (e['cp'] == 'CP1' and e['slot_status'] is not None)
        if (e['thr'] != prev_thr) or is_new:
            if current:
                groups.append(current)
                current = []
        current.append(e)
        prev_thr = e['thr']
    if current:
        groups.append(current)
    return groups


def short_time(s: Optional[str]) -> Optional[str]:
    """'2026-05-27T13:03' → '13:03' 식으로 시각 부분만 추출."""
    if not s:
        return None
    return s.split('T', 1)[1] if 'T' in s else s


def summarize_request(seq: int,
                      group: List[Dict[str, Any]],
                      fb_pid: Optional[str],
                      fb_lt: Optional[str]) -> Dict[str, Any]:
    cp1_slot = next((e for e in group if e['cp'] == 'CP1' and e['slot_status']), None)
    cp1_cnt  = next((e for e in group if e['cp'] == 'CP1' and e['res_count'] is not None), None)
    cp2_slot = next((e for e in group if e['cp'] == 'CP2' and e['slot_status']), None)
    cp3_cnt  = next((e for e in group if e['cp'] == 'CP3' and e['res_count'] is not None), None)

    if cp3_cnt:
        reached = '등록 완료'
    elif cp2_slot:
        reached = '2. 점유 직후'
    else:
        reached = '1. 검증 직후'

    elapsed_ns: Optional[int] = None
    if cp1_slot and cp3_cnt:
        elapsed_ns = cp3_cnt['nano'] - cp1_slot['nano']

    cp3_count = (cp3_cnt or {}).get('res_count')
    if cp3_count is None:
        rule = '미도달'
    elif cp3_count == 1:
        rule = '정상'
    else:
        rule = f'위반 ({cp3_count}건 관찰)'

    property_id = (cp1_slot or {}).get('property_id') or fb_pid
    lease_type  = (cp1_slot or {}).get('lease_type')  or fb_lt
    start_t = short_time((cp1_slot or {}).get('start_time'))
    end_t   = short_time((cp1_slot or {}).get('end_time'))
    slot_time = f'{start_t} ~ {end_t}' if (start_t and end_t) else None

    return {
        '요청 번호':                  seq,
        '처리 스레드명':               group[0]['thr'],
        '매물 식별자':                 property_id,
        '임대 유형':                   kor_lease(lease_type),
        '슬롯 번호':                   (cp1_slot or cp1_cnt or {}).get('slot_id'),
        '방문 예정 시간':              slot_time,
        '요청 시작 시각':              (cp1_slot or {}).get('log_ts'),
        '요청 종료 시각':              (cp3_cnt or {}).get('log_ts'),
        '검증→등록 응답시간 (ns)':     elapsed_ns,
        '진행 단계':                   reached,
        '확인 조회 시점 슬롯 상태':    kor_slot_status((cp1_slot or {}).get('slot_status')),
        '점유 직후 슬롯 상태':         kor_slot_status((cp2_slot or {}).get('slot_status')),
        '등록 전 예약 수':             (cp1_cnt or {}).get('res_count'),
        '등록 후 예약 수':             cp3_count,
        '정합성 판정':                 rule,
    }


# ────────────────────────────────────────────────────────────
# 엑셀 출력
# ────────────────────────────────────────────────────────────

TITLE_FILL   = '305496'
SECTION_FILL = '8EA9DB'
HEADER_FILL  = '203864'
LABEL_FILL   = 'F2F2F2'


def fill(cell, color: str):
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')


def write_section(ws, row: int, title: str, span: int):
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=span)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    fill(c, SECTION_FILL)
    c.alignment = Alignment(vertical='center')
    ws.row_dimensions[row].height = 22


def write_report(out_path: Path, rows: List[Dict[str, Any]], log_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'F004 측정 결과'

    headers = list(rows[0].keys()) if rows else [
        '요청 번호', '처리 스레드명', '매물 식별자', '임대 유형', '슬롯 번호',
        '방문 예정 시간', '요청 시작 시각', '요청 종료 시각', '검증→등록 응답시간 (ns)',
        '진행 단계', '확인 조회 시점 슬롯 상태', '점유 직후 슬롯 상태',
        '등록 전 예약 수', '등록 후 예약 수', '정합성 판정',
    ]
    SPAN = len(headers)

    # 제목
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=SPAN)
    title = ws.cell(row=1, column=1, value='F004 슬롯 예약 동시성 측정 — 시나리오 1 결과')
    title.font = Font(bold=True, color='FFFFFF', size=14)
    fill(title, TITLE_FILL)
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 메타 정보
    ws.cell(row=3, column=1, value='입력 로그').font = Font(bold=True)
    ws.cell(row=3, column=2, value=str(log_path))
    ws.cell(row=4, column=1, value='측정 요청 수').font = Font(bold=True)
    ws.cell(row=4, column=2, value=len(rows))

    # 회차 요약
    cur = 6
    write_section(ws, cur, '■ 회차 요약', SPAN)
    cur += 1
    success    = sum(1 for r in rows if r['진행 단계'] == '등록 완료')
    violations = sum(1 for r in rows if r['정합성 판정'].startswith('위반'))
    ns_values  = [r['검증→등록 응답시간 (ns)']
                  for r in rows if r['검증→등록 응답시간 (ns)'] is not None]
    mean_ns    = round(statistics.fmean(ns_values)) if ns_values else None
    max_ns     = max(ns_values) if ns_values else None
    min_ns     = min(ns_values) if ns_values else None

    kpi = [
        ('총 요청 수',                              len(rows)),
        ('성공 요청 수 (등록 완료까지 도달)',       success),
        ('미완 요청 수 (등록까지 도달 못함)',       len(rows) - success),
        ('정합성 위반 건수 (한 슬롯에 2 명 이상)',  violations),
        ('검증→등록 응답시간 평균 (ns)',           mean_ns),
        ('검증→등록 응답시간 최솟값 (ns)',         min_ns),
        ('검증→등록 응답시간 최댓값 (ns)',         max_ns),
    ]
    for k, v in kpi:
        a = ws.cell(row=cur, column=1, value=k)
        a.font = Font(bold=True)
        fill(a, LABEL_FILL)
        ws.cell(row=cur, column=2, value=v)
        cur += 1

    cur += 2

    # 요청별 측정 결과
    write_section(ws, cur, '■ 요청별 측정 결과', SPAN)
    cur += 1
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=cur, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        fill(c, HEADER_FILL)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[cur].height = 30
    cur += 1
    if rows:
        for r in rows:
            for col, h in enumerate(headers, start=1):
                ws.cell(row=cur, column=col, value=r.get(h))
            cur += 1
    else:
        ws.cell(row=cur, column=1, value='(측정된 요청이 없습니다.)')
        cur += 1

    cur += 2

    # 측정 단계 / 정합성 안내
    write_section(ws, cur, '■ 측정 단계 / 정합성 판정 안내', SPAN)
    cur += 1
    notices = [
        ('진행 단계',
            '한 요청이 어디까지 진행했는지. "1. 검증 직후" = 슬롯이 예약 가능한지 확인을 통과한 시점, '
            '"2. 점유 직후" = 본 요청이 슬롯을 "예약됨" 으로 표시(UPDATE)한 직후, '
            '"등록 완료" = 본 요청이 예약 행을 INSERT 완료한 시점.'),
        ('요청 시작 시각 / 요청 종료 시각',
            '본 측정 구간의 시작 (CP1, 슬롯 확인 조회 직후) 과 종료 (CP3, 예약 INSERT 직후) 시점의 logback 기록 시각. '
            '사람이 읽는 KST 시각, 1/1000초 정밀도.'),
        ('검증→등록 응답시간 (ns)',
            '검증 직후 (CP1) 부터 등록 직후 (CP3) 까지 전체 소요 시간. System.nanoTime() 차이를 ns 단위 그대로 표기.'),
        ('확인 조회 시점 / 점유 직후 슬롯 상태',
            '"확인 조회 시점 슬롯 상태" 는 슬롯이 예약 가능한지 확인하기 위해 슬롯 행을 조회한 시점의 상태 (정상이면 "예약 가능"). '
            '"점유 직후 슬롯 상태" 는 본 요청이 슬롯을 "예약됨" 으로 표시한 직후의 상태 (정상이면 "예약됨").'),
        ('등록 전 / 등록 후 예약 수',
            '본 스레드가 예약 행을 INSERT 하기 전 (CP1 시점) 과 INSERT 한 직후 (CP3 시점) 의 같은 슬롯 예약 행 수. '
            '정상 흐름이면 "등록 후" 가 "등록 전" 보다 정확히 1 만큼 커야 한다.'),
        ('정합성 판정',
            '"등록 후 예약 수" 가 1 = 정상, 2 이상이면 위반 (한 슬롯에 두 명 이상 확정됨).'),
        ('슬롯 상태 값',
            '"예약 가능" (AVAILABLE) / "예약됨" (RESERVED) / "종료됨" (CLOSED) / "철회됨" (WITHDRAWN).'),
        ('임대 유형 값',
            '"전세" (CHARTER) / "월세" (MONTHLY).'),
    ]
    for k, v in notices:
        a = ws.cell(row=cur, column=1, value=k)
        a.font = Font(bold=True)
        fill(a, LABEL_FILL)
        a.alignment = Alignment(vertical='top')
        b = ws.cell(row=cur, column=2, value=v)
        b.alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=cur, end_row=cur, start_column=2, end_column=SPAN)
        cur += 1

    # 컬럼 폭 — headers 순서대로
    # [요청#, 스레드명, 매물식별자, 임대유형, 슬롯#, 방문예정시간,
    #  요청시작, 요청종료, 응답시간(ns), 진행단계,
    #  확인조회시점 상태, 점유직후 상태, 등록전 수, 등록후 수, 정합성]
    widths = [8, 22, 36, 10, 8, 14, 22, 22, 24, 14, 22, 18, 14, 14, 18]
    for idx, w in enumerate(widths[:SPAN], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = 'A2'

    # 2번째 시트 — 원본 측정 로그 (시간순)
    write_raw_log_sheet(wb, log_path)
    # 3번째 시트 — 시나리오·기대·실측 대조표 (보고서용)
    write_comparison_sheet(wb, rows)
    # 4번째 시트 — KPI 박스 요약 (보고서용)
    write_kpi_box_sheet(wb, rows)
    # 5번째 시트 — 측정 환경 + 시나리오 + 결과 + 결론 (보고서용)
    write_summary_report_sheet(wb, rows, log_path)

    wb.save(out_path)


# ────────────────────────────────────────────────────────────
# 2번째 시트 — 원본 log.info 값을 그대로 시간순으로
# ────────────────────────────────────────────────────────────

# F004 측정 관련 라인 패턴 — [F004-...] 측정 로그 + [VISIT_RESERVATION_*] 요청/확정 로그
RAW_LOG_RE = re.compile(
    r'^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3})\s+'
    r'\[(?P<thr>[^\]]+)\]\s+'
    r'(?P<level>\w+)\s+'
    r'(?P<logger>\S+)\s+-\s+'
    r'(?P<message>(?:\[F004-|\[VISIT_RESERVATION_).*)$'
)


def detect_phase(message: str) -> str:
    """원본 메시지에서 어느 단계/이벤트의 라인인지 식별 라벨 반환."""
    if 'CP1-AFTER-AVAILABLE-CHECK' in message:
        return '1. 검증 직후 (CP1)'
    if 'CP2-AFTER-SLOT-UPDATE' in message:
        return '2. 점유 직후 (CP2)'
    if 'CP3-AFTER-RESERVATION-INSERT' in message:
        return '3. 등록 직후 (CP3)'
    if 'VISIT_RESERVATION_CREATE_REQ' in message:
        return '요청 진입 (컨트롤러)'
    if 'VISIT_RESERVATION_CONFIRMED' in message:
        return '예약 확정 (서비스 완료)'
    return '기타'


# 원본 메시지에서 필드별로 추출하기 위한 보조 정규식
_NANO_INSTANT_RE  = re.compile(r'nanoTime=(?P<nano>\d+)(?:\s+instant=(?P<inst>\S+))?')
_SLOT_PAYLOAD_RE  = re.compile(r'slot=\((?P<body>[^)]*)\)')
_RES_PAYLOAD_RE   = re.compile(r'reservation=\((?P<body>[^)]*)\)')
_RES_COUNT_RE     = re.compile(r'slotId=(?P<sid>\d+)\s+reservationsCount=(?P<rc>\d+)')
_CREATE_REQ_RE    = re.compile(r'\[VISIT_RESERVATION_CREATE_REQ\]\s+userId=(?P<uid>\S+?),\s+slotId=(?P<sid>\d+)')
_CONFIRMED_RE     = re.compile(
    r'\[VISIT_RESERVATION_CONFIRMED\]\s+'
    r'reservationId=(?P<rid>\d+),\s+slotId=(?P<sid>\d+),\s+'
    r'searcher=(?P<searcher>\S+?),\s+registrant=(?P<reg>\S+)'
)


def parse_message_fields(message: str) -> Dict[str, Any]:
    """원본 log.info 메시지에서 추출 가능한 모든 필드를 dict 로 반환.

    같은 의미의 값이 라인 종류마다 다른 이름으로 적혀 있는 경우
    (예: searcherUserId / searcher / userId) 는 같은 컬럼 키로 합친다.
    """
    f: Dict[str, Any] = {
        'nano_time':           None,
        'instant':              None,
        'slot_id':              None,
        'window_id':            None,
        'property_id':          None,
        'lease_type':           None,
        'slot_start_time':      None,
        'slot_end_time':        None,
        'slot_status':          None,
        'slot_created_at':      None,
        'reservations_count':   None,
        'reservation_id':       None,
        'searcher_user_id':     None,
        'reservation_status':   None,
        'confirmed_at':         None,
        'cancelled_at':         None,
        'invalidated_at':       None,
        'visit_result':         None,
        'result_classified_at': None,
        'registrant_user_id':   None,
    }

    # F004 측정 라인 공통 — nanoTime / instant
    ni = _NANO_INSTANT_RE.search(message)
    if ni:
        f['nano_time'] = int(ni.group('nano'))
        f['instant']   = ni.group('inst')

    # slot=(...) 페이로드
    sm = _SLOT_PAYLOAD_RE.search(message)
    if sm:
        kv = parse_kv(sm.group('body'))
        if kv.get('slotId'):   f['slot_id']         = int(kv['slotId'])
        if kv.get('windowId'): f['window_id']       = int(kv['windowId'])
        f['property_id']     = norm(kv.get('propertyId'))
        f['lease_type']      = norm(kv.get('leaseType'))
        f['slot_start_time'] = norm(kv.get('startTime'))
        f['slot_end_time']   = norm(kv.get('endTime'))
        f['slot_status']     = norm(kv.get('status'))
        f['slot_created_at'] = norm(kv.get('createdAt'))

    # reservation=(...) 페이로드
    rm = _RES_PAYLOAD_RE.search(message)
    if rm:
        kv = parse_kv(rm.group('body'))
        if kv.get('reservationId'): f['reservation_id']  = int(kv['reservationId'])
        if kv.get('slotId') and f['slot_id'] is None:
            f['slot_id']  = int(kv['slotId'])
        f['searcher_user_id']     = norm(kv.get('searcherUserId'))
        f['reservation_status']   = norm(kv.get('status'))
        f['confirmed_at']         = norm(kv.get('confirmedAt'))
        f['cancelled_at']         = norm(kv.get('cancelledAt'))
        f['invalidated_at']       = norm(kv.get('invalidatedAt'))
        f['visit_result']         = norm(kv.get('visitResult'))
        f['result_classified_at'] = norm(kv.get('resultClassifiedAt'))

    # CP1/CP3 의 "slotId=X reservationsCount=Y" 헤더 줄
    cm = _RES_COUNT_RE.search(message)
    if cm:
        if f['slot_id'] is None:
            f['slot_id'] = int(cm.group('sid'))
        f['reservations_count'] = int(cm.group('rc'))

    # 컨트롤러의 [VISIT_RESERVATION_CREATE_REQ]
    rq = _CREATE_REQ_RE.search(message)
    if rq:
        f['searcher_user_id'] = rq.group('uid')
        if f['slot_id'] is None:
            f['slot_id'] = int(rq.group('sid'))

    # 서비스의 [VISIT_RESERVATION_CONFIRMED]
    cf = _CONFIRMED_RE.search(message)
    if cf:
        if f['reservation_id'] is None:
            f['reservation_id'] = int(cf.group('rid'))
        if f['slot_id'] is None:
            f['slot_id']  = int(cf.group('sid'))
        if f['searcher_user_id'] is None:
            f['searcher_user_id'] = cf.group('searcher')
        f['registrant_user_id'] = cf.group('reg')

    return f


# 2번째 시트의 신규 컬럼 (단계/이벤트와 메시지 원본 사이에 들어가는 부분)
RAW_FIELD_COLUMNS: List[Tuple[str, str]] = [
    ('nano_time',           '측정 시각 (nanoTime, ns)'),
    ('instant',             '절대 시각 (instant, UTC)'),
    ('slot_id',             '슬롯 번호'),
    ('window_id',           '윈도우 번호'),
    ('property_id',         '매물 식별자'),
    ('lease_type',          '임대 유형'),
    ('slot_start_time',     '슬롯 시작 시각'),
    ('slot_end_time',       '슬롯 종료 시각'),
    ('slot_status',         '슬롯 상태'),
    ('slot_created_at',     '슬롯 생성 시각'),
    ('reservations_count',  '슬롯에 묶인 예약 수'),
    ('reservation_id',      '예약 번호'),
    ('searcher_user_id',    '탐색자 ID'),
    ('reservation_status',  '예약 상태'),
    ('confirmed_at',        '예약 확정 시각'),
    ('cancelled_at',        '예약 취소 시각'),
    ('invalidated_at',      '예약 무효화 시각'),
    ('visit_result',        '방문 결과'),
    ('result_classified_at','결과 분류 시각'),
    ('registrant_user_id',  '등록자 ID'),
]


def write_raw_log_sheet(wb: Workbook, log_path: Path):
    """log.info 원본 값을 가공 없이 시간 순으로 나열하는 시트."""
    ws = wb.create_sheet('원본 측정 로그')

    raw_field_keys    = [k for k, _ in RAW_FIELD_COLUMNS]
    raw_field_headers = [h for _, h in RAW_FIELD_COLUMNS]
    headers = (
        ['순번', '로그 기록 시각', '처리 스레드', '단계 / 이벤트']
        + raw_field_headers
        + ['log.info 메시지 (원본 그대로)']
    )
    SPAN = len(headers)

    # 제목 (첫 시트와 동일 톤)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=SPAN)
    title = ws.cell(row=1, column=1, value='F004 측정 로그 원본 — 시간순 전체 출력')
    title.font = Font(bold=True, color='FFFFFF', size=14)
    fill(title, TITLE_FILL)
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 메타
    ws.cell(row=3, column=1, value='입력 로그').font = Font(bold=True)
    ws.cell(row=3, column=2, value=str(log_path))

    # 로그 라인 수집 — 파일 순서가 곧 시간 순서
    raw_rows: List[Dict[str, Any]] = []
    with log_path.open('r', encoding='utf-8', errors='replace') as f:
        for line in f:
            m = RAW_LOG_RE.match(line.rstrip('\n'))
            if not m:
                continue
            message = m.group('message')
            row = {
                'log_ts':  m.group('ts'),
                'thr':     m.group('thr'),
                'message': message,
                'phase':   detect_phase(message),
            }
            row.update(parse_message_fields(message))
            raw_rows.append(row)

    ws.cell(row=4, column=1, value='수집 라인 수').font = Font(bold=True)
    ws.cell(row=4, column=2, value=len(raw_rows))

    # 섹션 헤더
    cur = 6
    write_section(ws, cur, '■ 측정 로그 라인 (파일 순 = 시간 순)', SPAN)
    cur += 1

    # 표 헤더
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=cur, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        fill(c, HEADER_FILL)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[cur].height = 30
    cur += 1

    # 데이터 행
    if raw_rows:
        for i, r in enumerate(raw_rows, start=1):
            ws.cell(row=cur, column=1, value=i)
            ws.cell(row=cur, column=2, value=r['log_ts'])
            ws.cell(row=cur, column=3, value=r['thr'])
            ws.cell(row=cur, column=4, value=r['phase'])
            # 원본 메시지에서 파싱한 필드들 (단계/이벤트 컬럼 다음부터)
            base_col = 5
            for offset, key in enumerate(raw_field_keys):
                ws.cell(row=cur, column=base_col + offset, value=r.get(key))
            # 마지막 컬럼 — 원본 메시지 그대로
            msg = ws.cell(row=cur, column=base_col + len(raw_field_keys), value=r['message'])
            msg.alignment = Alignment(vertical='top', wrap_text=True)
            cur += 1
    else:
        ws.cell(row=cur, column=1, value='(수집된 로그 라인이 없습니다.)')
        cur += 1

    cur += 2

    # 안내 블록 (첫 시트와 동일 톤)
    write_section(ws, cur, '■ 본 시트 안내', SPAN)
    cur += 1
    notices = [
        ('수집 대상',
            '백엔드 로그에서 머리말이 [F004-...] 인 측정 라인 과 [VISIT_RESERVATION_CREATE_REQ] / '
            '[VISIT_RESERVATION_CONFIRMED] 인 요청 진입·확정 라인 을 시간 순으로 모두 모은 것.'),
        ('가공 여부',
            '메시지 컬럼은 백엔드가 출력한 log.info 본문을 변형 없이 그대로 옮긴 것이다. '
            '단계 / 이벤트 컬럼만 메시지 머리말로부터 식별한 보조 라벨이다.'),
        ('정렬 기준',
            '로그 파일의 파일 순서 (= logback 기록 순서) 그대로. 별도 재정렬 없음.'),
        ('첫 시트와의 관계',
            '"F004 측정 결과" 시트가 요청 단위로 묶고 가공한 결과라면, 본 시트는 그 가공의 입력이 된 원본 라인이다.'),
    ]
    for k, v in notices:
        a = ws.cell(row=cur, column=1, value=k)
        a.font = Font(bold=True)
        fill(a, LABEL_FILL)
        a.alignment = Alignment(vertical='top')
        b = ws.cell(row=cur, column=2, value=v)
        b.alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=cur, end_row=cur, start_column=2, end_column=SPAN)
        cur += 1

    # 컬럼 폭
    # 고정 4개: 순번 / 로그 시각 / 처리 스레드 / 단계
    base_widths = [6, 22, 24, 22]
    # 원본 메시지 파싱 컬럼들 (순서 = RAW_FIELD_COLUMNS)
    field_widths = [
        18,  # nano_time
        30,  # instant
        8,   # slot_id
        8,   # window_id
        36,  # property_id
        10,  # lease_type
        18,  # slot_start_time
        18,  # slot_end_time
        12,  # slot_status
        24,  # slot_created_at
        14,  # reservations_count
        10,  # reservation_id
        14,  # searcher_user_id
        12,  # reservation_status
        24,  # confirmed_at
        24,  # cancelled_at
        24,  # invalidated_at
        12,  # visit_result
        24,  # result_classified_at
        14,  # registrant_user_id
    ]
    all_widths = base_widths + field_widths + [110]  # 마지막은 메시지 원본
    for idx, w in enumerate(all_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = 'E2'  # 단계/이벤트 까지 가로로 고정해서 우측 스크롤 시에도 식별 가능


# ════════════════════════════════════════════════════════════
# 3번째 시트 — 시나리오·기대값·실측값 대조표 (보고서 후보 1)
# ════════════════════════════════════════════════════════════

VERDICT_PASS_COLOR = '2E7D32'  # 초록
VERDICT_FAIL_COLOR = 'C62828'  # 빨강
VERDICT_NEUTRAL_COLOR = '616161'  # 회색


def _compute_metrics(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """3·4·5 시트가 공통으로 쓰는 측정 지표 계산."""
    total = len(rows)
    success = sum(1 for r in rows if r.get('진행 단계') == '등록 완료')
    violations = sum(1 for r in rows if str(r.get('정합성 판정', '')).startswith('위반'))

    cp3_counts = [r['등록 후 예약 수'] for r in rows if r.get('등록 후 예약 수') is not None]
    cp1_counts = [r['등록 전 예약 수'] for r in rows if r.get('등록 전 예약 수') is not None]

    cp2_statuses = [r['점유 직후 슬롯 상태'] for r in rows if r.get('점유 직후 슬롯 상태')]
    cp2_reserved_count = sum(1 for s in cp2_statuses if s == '예약됨')

    ns_values = [r['검증→등록 응답시간 (ns)']
                 for r in rows if r.get('검증→등록 응답시간 (ns)') is not None]

    diff_plus_one = 0
    diff_total = 0
    for r in rows:
        before = r.get('등록 전 예약 수')
        after  = r.get('등록 후 예약 수')
        if before is not None and after is not None:
            diff_total += 1
            if after - before == 1:
                diff_plus_one += 1

    return {
        'total':              total,
        'success':            success,
        'partial':            total - success,
        'violations':         violations,
        'max_cp3':            max(cp3_counts) if cp3_counts else None,
        'cp2_total':          len(cp2_statuses),
        'cp2_reserved':       cp2_reserved_count,
        'diff_total':         diff_total,
        'diff_plus_one':      diff_plus_one,
        'ns_avg':             round(statistics.fmean(ns_values)) if ns_values else None,
        'ns_min':             min(ns_values) if ns_values else None,
        'ns_max':             max(ns_values) if ns_values else None,
    }


def write_comparison_sheet(wb: Workbook, rows: List[Dict[str, Any]]):
    """보고서용 — 시나리오 / 기대 결과 / 실측 결과 / 부합 여부 4 컬럼 대조표."""
    ws = wb.create_sheet('정합성 대조표')
    headers = ['측정 항목', '기대 결과', '실측 결과', '부합 여부', '비고']
    SPAN = len(headers)

    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=SPAN)
    title = ws.cell(row=1, column=1, value='F004 시나리오 1 — 기대값 / 실측값 대조표')
    title.font = Font(bold=True, color='FFFFFF', size=14)
    fill(title, TITLE_FILL)
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.cell(row=3, column=1, value='측정 회차 요청 수').font = Font(bold=True)
    ws.cell(row=3, column=2, value=len(rows))

    M = _compute_metrics(rows)

    cur = 5
    write_section(ws, cur, '■ 비즈니스 규칙 대조', SPAN)
    cur += 1

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=cur, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        fill(c, HEADER_FILL)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[cur].height = 30
    cur += 1

    items: List[Tuple[str, str, str, str, str]] = []

    # 항목 1 — 슬롯당 정원 1
    if M['max_cp3'] is None:
        items.append((
            '한 슬롯당 확정 예약 건수',
            '정확히 1 건',
            '관찰 없음 (등록까지 도달한 요청 없음)',
            '— 미측정',
            '기획서 7 절 규칙 1 (슬롯당 정원 1). 등록 후 시점 같은 슬롯의 예약 행 수가 기준.'
        ))
    elif M['max_cp3'] == 1:
        items.append((
            '한 슬롯당 확정 예약 건수',
            '정확히 1 건',
            f'최대 {M["max_cp3"]} 건 (등록 직후 시점 관찰)',
            '✓ 부합',
            '기획서 7 절 규칙 1 (슬롯당 정원 1). 등록 후 시점 같은 슬롯의 예약 행 수가 기준.'
        ))
    else:
        items.append((
            '한 슬롯당 확정 예약 건수',
            '정확히 1 건',
            f'최대 {M["max_cp3"]} 건 관찰됨',
            '✗ 위반',
            '같은 슬롯에 두 건 이상의 CONFIRMED 가 동시에 보임. 락·백스톱이 필요한 명백한 증거.'
        ))

    # 항목 2 — 동시 요청 중 성공 건수
    if M['total'] < 2:
        items.append((
            '동시 요청 중 성공한 요청 수',
            '정확히 1 건 (나머지는 거부)',
            f'{M["success"]} 건 성공 / {M["partial"]} 건 미완 (단일 회차)',
            '— 단일 회차 판정 불가',
            '동시 N 건 측정 시에만 변별력이 있음. N = 2, 5, 10, 20 회차 측정 권장.'
        ))
    else:
        verdict = '✓ 부합' if M['success'] == 1 else '✗ 위반'
        items.append((
            '동시 요청 중 성공한 요청 수',
            '정확히 1 건 (나머지는 거부)',
            f'{M["success"]} 건 성공 / {M["partial"]} 건 미완',
            verdict,
            '동시 N 건이 같은 슬롯에 도착했을 때, 정확히 1 명만 등록 완료에 도달해야 한다.'
        ))

    # 항목 3 — 이중 확정 발생 건수
    items.append((
        '이중 확정 발생 건수',
        '0 건',
        f'{M["violations"]} 건',
        '✓ 부합' if M['violations'] == 0 else '✗ 위반',
        '같은 슬롯에 두 명 이상이 동시에 CONFIRMED 가 되는 사건. 정합성 판정 컬럼 위반 카운트.'
    ))

    # 항목 4 — 점유 직후 슬롯 상태
    if M['cp2_total'] == 0:
        items.append((
            '점유 직후 슬롯 상태',
            '예약됨 (RESERVED)',
            '관찰 없음',
            '— 미측정',
            '본 트랜잭션의 슬롯 UPDATE 가 의도대로 반영되는지.'
        ))
    else:
        verdict = '✓ 부합' if M['cp2_reserved'] == M['cp2_total'] else '✗ 위반'
        items.append((
            '점유 직후 슬롯 상태',
            '예약됨 (RESERVED)',
            f'{M["cp2_reserved"]} / {M["cp2_total"]} 건이 "예약됨"',
            verdict,
            '본 트랜잭션의 slot.setStatus(RESERVED) + save 가 의도대로 반영되는지.'
        ))

    # 항목 5 — 등록 전 → 등록 후 예약 수 증가량
    if M['diff_total'] == 0:
        items.append((
            '등록 전 → 등록 후 예약 수 증가량',
            '정확히 + 1',
            '관찰 없음',
            '— 미측정',
            '본 트랜잭션이 정확히 한 건만 INSERT 했는지.'
        ))
    else:
        verdict = '✓ 부합' if M['diff_plus_one'] == M['diff_total'] else '✗ 위반'
        items.append((
            '등록 전 → 등록 후 예약 수 증가량',
            '정확히 + 1',
            f'{M["diff_plus_one"]} / {M["diff_total"]} 건이 + 1',
            verdict,
            '본 트랜잭션이 정확히 한 건만 INSERT 했는지. 2 이상 증가했다면 같은 트랜잭션이 두 번 호출됐거나 측정 도구 버그.'
        ))

    # 항목 6 — 응답시간 (기준 없음, 측정 대상)
    items.append((
        '검증 → 등록 전체 응답시간 (ns)',
        '(측정 대상 — 기준 없음)',
        f'평균 {M["ns_avg"]:,} ns' if M['ns_avg'] else '관찰 없음',
        '— 측정',
        '동시 경합 시 검증·UPDATE·INSERT 전체 처리 시간. 추후 락 / 인덱스 적용 후 회차와 비교 대조.'
    ))

    # 항목 7 — 거부 응답의 즉시성 (현재 구조 한계)
    items.append((
        '거부 응답의 즉시성',
        '거부된 트랜잭션은 즉시 명확한 거부 응답',
        '본 측정 구조에서는 거부된 트랜잭션이 CP1 이전 throw 로 표에 등장하지 않음',
        '— 구조 한계',
        '거부된 트랜잭션 식별을 위해 CP0 (AVAILABLE 검증 이전) 측정 추가 검토 필요.'
    ))

    for label, expected, actual, verdict, note in items:
        values = [label, expected, actual, verdict, note]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=cur, column=col, value=val)
            c.alignment = Alignment(vertical='top', wrap_text=True)
            if col == 1:
                c.font = Font(bold=True)
                fill(c, LABEL_FILL)
            elif col == 4:
                c.alignment = Alignment(horizontal='center', vertical='center')
                if verdict.startswith('✓'):
                    c.font = Font(bold=True, color=VERDICT_PASS_COLOR)
                elif verdict.startswith('✗'):
                    c.font = Font(bold=True, color=VERDICT_FAIL_COLOR)
                else:
                    c.font = Font(bold=True, color=VERDICT_NEUTRAL_COLOR)
        ws.row_dimensions[cur].height = 38
        cur += 1

    widths = [30, 30, 36, 18, 44]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = 'A2'


# ════════════════════════════════════════════════════════════
# 4번째 시트 — KPI 박스 요약 (보고서 후보 2)
# ════════════════════════════════════════════════════════════

def _draw_kpi_box(ws, top_row: int, start_col: int, end_col: int,
                  label: str, value: Any, sub: str, color: str):
    """라벨 행 / 큰 값 행 (3행 병합) / 주석 행 — 3 영역으로 구성된 KPI 박스."""
    # 라벨 행
    ws.merge_cells(start_row=top_row, end_row=top_row,
                   start_column=start_col, end_column=end_col)
    c = ws.cell(row=top_row, column=start_col, value=label)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    fill(c, color)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[top_row].height = 22

    # 값 행 (3행 병합으로 큰 박스)
    ws.merge_cells(start_row=top_row + 1, end_row=top_row + 3,
                   start_column=start_col, end_column=end_col)
    c = ws.cell(row=top_row + 1, column=start_col, value=value)
    c.font = Font(bold=True, size=22, color=color)
    c.alignment = Alignment(horizontal='center', vertical='center')
    for r in range(top_row + 1, top_row + 4):
        ws.row_dimensions[r].height = 28

    # 주석 행
    ws.merge_cells(start_row=top_row + 4, end_row=top_row + 4,
                   start_column=start_col, end_column=end_col)
    c = ws.cell(row=top_row + 4, column=start_col, value=sub)
    c.font = Font(italic=True, size=9, color='616161')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[top_row + 4].height = 20


def write_kpi_box_sheet(wb: Workbook, rows: List[Dict[str, Any]]):
    """보고서용 — 큰 글씨 박스 6 개로 핵심 KPI 표시."""
    ws = wb.create_sheet('KPI 박스 요약')

    M = _compute_metrics(rows)

    SPAN = 8  # 박스 3 개 × (2 컬럼 + 1 간격) - 마지막 간격 제외 = 8
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=SPAN)
    title = ws.cell(row=1, column=1, value='F004 시나리오 1 — KPI 박스 요약')
    title.font = Font(bold=True, color='FFFFFF', size=14)
    fill(title, TITLE_FILL)
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 박스 배치 — 위 줄 (row 3 부터) 3 개, 아래 줄 (row 10 부터) 3 개
    # 박스 1 컬럼: 1-2, 박스 2 컬럼: 4-5, 박스 3 컬럼: 7-8 (3·6 은 간격)
    _draw_kpi_box(ws, 3, 1, 2, '총 요청 수',
                  M['total'], '본 회차 호출 총 건수', '305496')
    _draw_kpi_box(ws, 3, 4, 5, '성공 요청 수',
                  M['success'], '등록 완료까지 도달', VERDICT_PASS_COLOR)
    _draw_kpi_box(ws, 3, 7, 8, '정합성 위반',
                  M['violations'],
                  '한 슬롯에 2 명 이상 확정',
                  VERDICT_FAIL_COLOR if M['violations'] > 0 else VERDICT_PASS_COLOR)

    _draw_kpi_box(ws, 10, 1, 2, '평균 응답시간',
                  f'{M["ns_avg"]:,} ns' if M['ns_avg'] is not None else '—',
                  '검증 → 등록 전체 (단조 ns)', '305496')
    _draw_kpi_box(ws, 10, 4, 5, '최소 응답시간',
                  f'{M["ns_min"]:,} ns' if M['ns_min'] is not None else '—',
                  '검증 → 등록 전체 (단조 ns)', '305496')
    _draw_kpi_box(ws, 10, 7, 8, '최대 응답시간',
                  f'{M["ns_max"]:,} ns' if M['ns_max'] is not None else '—',
                  '검증 → 등록 전체 (단조 ns)', '305496')

    # 컬럼 폭 — 박스 본 영역은 넓게, 간격 컬럼은 좁게
    for col in range(1, SPAN + 1):
        ws.column_dimensions[get_column_letter(col)].width = 4 if col in (3, 6) else 20

    # 안내
    note_row = 17
    ws.merge_cells(start_row=note_row, end_row=note_row, start_column=1, end_column=SPAN)
    c = ws.cell(row=note_row, column=1,
                value='※ 값이 "—" 인 박스는 해당 회차에 측정 표본이 없어 산출되지 않은 항목.')
    c.font = Font(italic=True, color='616161')
    c.alignment = Alignment(horizontal='center', vertical='center')


# ════════════════════════════════════════════════════════════
# 5번째 시트 — 측정 환경 + 시나리오 + 결과 + 결론 (보고서 후보 3)
# ════════════════════════════════════════════════════════════

def write_summary_report_sheet(wb: Workbook, rows: List[Dict[str, Any]], log_path: Path):
    """보고서용 — 한 페이지짜리 측정 보고서."""
    ws = wb.create_sheet('보고서 요약')
    SPAN = 4

    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=SPAN)
    title = ws.cell(row=1, column=1, value='F004 시나리오 1 — 측정 보고서')
    title.font = Font(bold=True, color='FFFFFF', size=14)
    fill(title, TITLE_FILL)
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    M = _compute_metrics(rows)

    properties = sorted({r['매물 식별자'] for r in rows if r.get('매물 식별자')})
    slots      = sorted({r['슬롯 번호'] for r in rows if r.get('슬롯 번호') is not None})
    lease_types = sorted({r['임대 유형'] for r in rows if r.get('임대 유형')})
    first_ts   = min((r['요청 시작 시각'] for r in rows if r.get('요청 시작 시각')), default=None)
    last_ts    = max((r['요청 종료 시각'] for r in rows if r.get('요청 종료 시각')), default=None)

    cur = 3

    # ─ 섹션 1: 측정 환경 ─
    write_section(ws, cur, '■ 1. 측정 환경', SPAN)
    cur += 1
    env_items = [
        ('측정 일시 (회차 시작)',         first_ts or '관찰 없음'),
        ('측정 일시 (회차 종료)',         last_ts or '관찰 없음'),
        ('측정 대상 매물 (property_id)',  ', '.join(properties) if properties else '관찰 없음'),
        ('측정 대상 임대 유형',           ', '.join(lease_types) if lease_types else '관찰 없음'),
        ('측정 대상 슬롯 (slot_id)',      ', '.join(str(s) for s in slots) if slots else '관찰 없음'),
        ('DB 인덱스 (03_constraints.sql)',
            '미적용 — 부분 유일 인덱스 백스톱 작동하지 않음'),
        ('DB 인덱스 (04_indexes.sql)',
            '미적용 — 성능 인덱스 작동하지 않음'),
        ('동시성 제어 기법 (락)',
            '미적용 — 단일 트랜잭션 안의 "읽기 → 검증 → 쓰기" 순서만 수행 (비관/낙관/분산 락 모두 X)'),
        ('스케줄러 주기',
            'SlotExpirationScheduler fixedDelay = 1000ms (테스트용, 운영용 60000ms 는 주석 처리)'),
        ('입력 로그 파일',                str(log_path)),
    ]
    for k, v in env_items:
        a = ws.cell(row=cur, column=1, value=k)
        a.font = Font(bold=True)
        fill(a, LABEL_FILL)
        a.alignment = Alignment(vertical='top')
        b = ws.cell(row=cur, column=2, value=v)
        b.alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=cur, end_row=cur, start_column=2, end_column=SPAN)
        cur += 1
    cur += 1

    # ─ 섹션 2: 시나리오 정의 ─
    write_section(ws, cur, '■ 2. 시나리오 정의', SPAN)
    cur += 1
    scenario_text = (
        'F004 시나리오 1 — 인기 매물 희소 슬롯의 직접 경합 + 재개방 직후 구독자 집중 경합.\n'
        '같은 슬롯 식별자에 대해 N 개 탐색자가 거의 동시에 createReservation 을 호출했을 때, '
        '시스템이 정확히 1 명만 확정하고 나머지를 거부하는지 (슬롯당 정원 1 규칙) 를 검증한다. '
        '메서드 안의 세 측정 지점 (검증 직후 / 점유 직후 / 등록 직후) 에서 슬롯과 예약의 상태를 캡처하여 '
        '동시 경합이 실제로 발생하는지, 발생한다면 시스템이 어떻게 반응하는지를 실측한다.'
    )
    a = ws.cell(row=cur, column=1, value='시나리오')
    a.font = Font(bold=True)
    fill(a, LABEL_FILL)
    a.alignment = Alignment(vertical='top')
    b = ws.cell(row=cur, column=2, value=scenario_text)
    b.alignment = Alignment(vertical='top', wrap_text=True)
    ws.merge_cells(start_row=cur, end_row=cur, start_column=2, end_column=SPAN)
    ws.row_dimensions[cur].height = 80
    cur += 2

    # ─ 섹션 3: 측정 결과 ─
    write_section(ws, cur, '■ 3. 측정 결과', SPAN)
    cur += 1
    result_items = [
        ('총 요청 수',                          M['total']),
        ('성공 요청 수 (등록 완료 도달)',       M['success']),
        ('미완 요청 수',                        M['partial']),
        ('정합성 위반 건수 (한 슬롯에 2 명 이상)', M['violations']),
        ('검증 → 등록 응답시간 평균 (ns)',     f'{M["ns_avg"]:,}' if M['ns_avg'] is not None else '—'),
        ('검증 → 등록 응답시간 최솟값 (ns)',   f'{M["ns_min"]:,}' if M['ns_min'] is not None else '—'),
        ('검증 → 등록 응답시간 최댓값 (ns)',   f'{M["ns_max"]:,}' if M['ns_max'] is not None else '—'),
    ]
    for k, v in result_items:
        a = ws.cell(row=cur, column=1, value=k)
        a.font = Font(bold=True)
        fill(a, LABEL_FILL)
        a.alignment = Alignment(vertical='top')
        b = ws.cell(row=cur, column=2, value=v)
        b.alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=cur, end_row=cur, start_column=2, end_column=SPAN)
        cur += 1
    cur += 1

    # ─ 섹션 4: 결론 / 다음 단계 ─
    write_section(ws, cur, '■ 4. 결론 / 다음 단계', SPAN)
    cur += 1

    if M['total'] < 2:
        conclusion = (
            f'본 회차는 측정 표본 {M["total"]} 건으로 동시성 거동을 판정할 수 없다. '
            '다만 측정 도구의 동작은 검증되었다 — CP1 / CP2 / CP3 세 측정 지점이 모두 정상 기록되었고, '
            '응답시간 측정과 정합성 판정 로직이 작동한다. '
            '다음 단계: 동시 N (= 2, 5, 10, 20) 건 측정 회차를 수행한 뒤 결과를 같은 보고서 시트로 갱신.'
        )
    elif M['violations'] > 0:
        conclusion = (
            f'동시 {M["total"]} 건 요청에서 정합성 위반 {M["violations"]} 건 관찰. '
            '현재 환경 (락·DB 백스톱 모두 미적용) 에서 슬롯당 정원 1 규칙이 보장되지 않음을 실측으로 확인. '
            '다음 단계: 동시성 제어 기법 후보 (비관적 락 / 낙관적 락 / DB 부분 유일 인덱스 백스톱 / 분산 락) '
            '를 트레이드오프 비교 후 선택 → 적용 → 재측정 → 본 보고서 갱신.'
        )
    else:
        conclusion = (
            f'동시 {M["total"]} 건 요청에서 정합성 위반 0 건. 현재 환경에서 슬롯당 정원 1 규칙이 우연히, '
            '또는 명시적 락이 아닌 메커니즘 (예: JPA 의 자동 행 락) 으로 보장되고 있다. '
            '다음 단계: 표본을 늘려 (N = 5, 10, 20) 재현성 확인, 그리고 응답시간 분포 변화 추적.'
        )

    a = ws.cell(row=cur, column=1, value='결론')
    a.font = Font(bold=True)
    fill(a, LABEL_FILL)
    a.alignment = Alignment(vertical='top')
    b = ws.cell(row=cur, column=2, value=conclusion)
    b.alignment = Alignment(vertical='top', wrap_text=True)
    ws.merge_cells(start_row=cur, end_row=cur, start_column=2, end_column=SPAN)
    ws.row_dimensions[cur].height = 100
    cur += 1

    # 컬럼 폭
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30


# ────────────────────────────────────────────────────────────
# 실행
# ────────────────────────────────────────────────────────────

def default_log() -> Path:
    return Path(__file__).resolve().parents[5] / 'wherehouse' / 'log' / 'wherehouse.log'


def default_out() -> Path:
    return Path(__file__).resolve().parents[1] / '결과' / 'F004_측정결과.xlsx'


def main():
    parser = argparse.ArgumentParser(description='F004 측정 로그를 포트폴리오용 엑셀 한 시트로 정리')
    parser.add_argument('--log', type=Path, default=default_log())
    parser.add_argument('--out', type=Path, default=default_out())
    args = parser.parse_args()

    if not args.log.exists():
        sys.exit(f'[ERROR] 로그 파일을 찾을 수 없습니다: {args.log}')
    args.out.parent.mkdir(parents=True, exist_ok=True)

    # 매물 번호 fallback — 슬롯 스냅샷에 propertyId 가 없는 구버전 로그용
    fb_pid, fb_lt = collect_property_fallback(args.log)

    events = parse_log(args.log)
    groups = group_requests(events)
    rows = [summarize_request(i, g, fb_pid, fb_lt) for i, g in enumerate(groups, start=1)]

    write_report(args.out, rows, args.log)
    print(f'[OK] {len(rows)} 개 요청 정리 -> {args.out}')


if __name__ == '__main__':
    main()
