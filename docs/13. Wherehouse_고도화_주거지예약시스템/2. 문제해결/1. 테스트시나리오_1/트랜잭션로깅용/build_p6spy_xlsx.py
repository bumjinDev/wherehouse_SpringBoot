"""
F004 시나리오 1 — P6Spy 로그에서 N 동시 요청의 트랜잭션별 SQL 시퀀스를 추출해 엑셀로 정리.

사용법:
    py -3.13 build_p6spy_xlsx.py

상단 "사용자 설정" 블록의 LOG_PATH / OUT_PATH / N_CONCURRENT 만 측정 회차에 맞춰 수정.
한 사이클 = 같은 슬롯에 N 건 동시 호출. 로그의 가장 최근 N 건의 INSERT VISIT_RESERVATION
을 한 사이클로 자동 인식하고, 그 스레드의 ±10초 범위 SQL 만 추출한다.

시트 구성:
    1. 사용자별 SQL 시퀀스             — P6Spy 가 찍은 binding SQL 그대로
    2. 사용자별 SQL 시퀀스 (Oracle 실행형) — SQL Developer 에 그대로 붙여넣어 발행 가능
    3. 전체 시간순 인터리브             — N 스레드의 모든 SQL 시간순 정렬 (race 관찰)
    4. SQL 종류 카운트                  — 사용자별 SQL 종류별 발행 횟수

P6Spy 가 같은 SQL 을 두 번 (P6Spy wrapper + underlying driver) 찍기 때문에,
url 이 jdbc:p6spy: 로 시작하는 wrapper 라인만 사용해 중복 제거.

요구 패키지: openpyxl (>= 3.1)
"""

import re
import sys
from collections import defaultdict, Counter
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# ════════════════════════════════════════════════════════════════════════
# 사용자 설정 — 측정 회차마다 여기만 수정
# ════════════════════════════════════════════════════════════════════════
LOG_PATH = Path(
    r"E:\devSpace\SpringBootProjects\wherehouse_SpringBoot-master"
    r"\wherehouse\log\wherehouse.log"
)
OUT_PATH = Path(
    r"E:\devSpace\SpringBootProjects\wherehouse_SpringBoot-master"
    r"\docs\13. Wherehouse_고도화_주거지예약시스템\2. 문제해결\1. 테스트시나리오_1"
    r"\트랜잭션로깅용\F004_P6Spy_SQL_시퀀스.xlsx"
)
N_CONCURRENT = 5  # 한 사이클의 동시 요청 수

# 메타 정보 (시트 상단에 표시될 텍스트). 회차마다 수정.
META_SLOT_DESC = "slot_id=401, window_id=42, 20:55~21:25 (CHARTER, 매물 693c35...)"
META_ENV_DESC = "UQ 인덱스 제거 / 동시성 제어 없음 (Step 2 baseline 재현)"
META_RESULT_DESC = "5건 모두 commit → 정합성 위반 (CONFIRMED 5건 잔존)"

# ════════════════════════════════════════════════════════════════════════
# 정규식
# ════════════════════════════════════════════════════════════════════════
P6SPY_RE = re.compile(
    r"^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3})\s+"
    r"\[(?P<thread>http-nio-\d+-exec-\d+)\]\s+INFO\s+p6spy\s+-\s+"
    r"#(?P<nano>\d+)\s+\|\s+took\s+(?P<took>\d+)ms\s+\|\s+(?P<kind>\w+)\s+\|\s+"
    r"connection\s+(?P<conn>\d+)\|\s*url\s+(?P<url>\S+)"
)
INSERT_USER_RE = re.compile(
    r"insert into VISIT_RESERVATION.*?'(?P<user>searcher\d+)'", re.IGNORECASE
)
# ISO 8601 with timezone offset: '2026-05-28T20:05:32.110+0900'
ISO_TS_RE = re.compile(
    r"'(\d{4}-\d{2}-\d{2})T(\d{2}:\d{2}:\d{2}\.\d+)\+\d{4}'"
)

# ════════════════════════════════════════════════════════════════════════
# 스타일
# ════════════════════════════════════════════════════════════════════════
TITLE_FONT = Font(bold=True, size=13)
META_FONT = Font(bold=True)
SECTION_FILL = PatternFill("solid", fgColor="305496")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="8EA9DB")
HEADER_FONT = Font(bold=True, color="FFFFFF")
UPDATE_FILL = PatternFill("solid", fgColor="FFE699")
INSERT_FILL = PatternFill("solid", fgColor="C6EFCE")
COMMIT_FILL = PatternFill("solid", fgColor="D9D9D9")


# ════════════════════════════════════════════════════════════════════════
# SQL 분류 / 변환
# ════════════════════════════════════════════════════════════════════════
def classify_sql(sql: str) -> str:
    s = sql.lower()
    if s.startswith("select") and "visit_slot" in s and "join" not in s and "order by" not in s:
        return "SELECT 슬롯 (findById)"
    if s.startswith("select") and "visit_window" in s and "count" not in s:
        return "SELECT 윈도우 (findById)"
    if s.startswith("select") and "visit_reservation" in s and "order by" in s:
        return "SELECT 예약 (logReservationsOnSlot)"
    if s.startswith("select") and "properties_charter" in s:
        return "SELECT 매물 (CHARTER)"
    if s.startswith("select") and "membertbl" in s:
        return "SELECT 회원 (등록자)"
    if s.startswith("select count") and "visit_window" in s and "lease_type" in s:
        return "COUNT 동일 매물 중복"
    if s.startswith("select count") and "start_time" in s and "end_time" in s:
        return "COUNT 시간 겹침"
    if "nextval" in s and "seq_visit_reservation" in s:
        return "SEQ 예약 ID"
    if "nextval" in s and "seq_visit_notification" in s:
        return "SEQ 알림 ID"
    if s.startswith("select") and "reopen_subscription" in s:
        return "SELECT 구독 (본인 활성)"
    if s.startswith("insert into visit_reservation"):
        return "INSERT 예약 (CONFIRMED)"
    if s.startswith("insert into visit_notification"):
        return "INSERT 알림"
    if s.startswith("update visit_slot"):
        return "UPDATE 슬롯 (→ RESERVED)"
    if s.upper().startswith("COMMIT"):
        return "COMMIT"
    return "OTHER"


def to_oracle_sql(sql: str) -> str:
    """P6Spy binding SQL → Oracle SQL Developer 실행 가능 형태.

    변환 사항:
      - ISO 8601 + timezone offset '2026-05-28T20:05:32.110+0900'
            → TIMESTAMP '2026-05-28 20:05:32.110'
      - 끝에 세미콜론 보장
    """
    if sql.strip().upper().startswith("COMMIT"):
        return "COMMIT;"
    sql = ISO_TS_RE.sub(r"TIMESTAMP '\1 \2'", sql)
    sql = sql.rstrip().rstrip(";") + ";"
    return sql


# ════════════════════════════════════════════════════════════════════════
# 로그 파싱
# ════════════════════════════════════════════════════════════════════════
def parse_log(log_path: Path):
    lines = log_path.read_text(encoding="utf-8").splitlines()
    records = []
    i = 0
    while i < len(lines):
        line = lines[i]
        m = P6SPY_RE.match(line)
        if not m:
            i += 1
            continue
        ts = m.group("ts")
        thread = m.group("thread")
        took = int(m.group("took"))
        kind = m.group("kind")
        conn = m.group("conn")
        url = m.group("url")
        if kind == "statement":
            placeholder = lines[i + 1] if i + 1 < len(lines) else ""
            binding = lines[i + 2].rstrip(";").strip() if i + 2 < len(lines) else ""
            records.append({
                "ts": ts, "thread": thread, "took": took, "kind": "statement",
                "conn": conn, "url": url, "placeholder": placeholder, "sql": binding,
            })
            i += 3
        elif kind == "commit":
            records.append({
                "ts": ts, "thread": thread, "took": took, "kind": "commit",
                "conn": conn, "url": url, "placeholder": "", "sql": "COMMIT",
            })
            i += 3 if i + 2 < len(lines) else 1
        else:
            i += 1
    return records


def select_cycle(records, n_concurrent):
    """가장 최근 N 건의 INSERT VISIT_RESERVATION 을 한 사이클로 잡고
    그 스레드의 라인만 ±10초 범위에서 추출."""
    recs = [r for r in records if r["url"].startswith("jdbc:p6spy:")]
    inserts = [r for r in recs
               if r["kind"] == "statement"
               and r["sql"].lower().startswith("insert into visit_reservation")]
    if len(inserts) < n_concurrent:
        raise RuntimeError(
            f"INSERT VISIT_RESERVATION 라인이 {n_concurrent}건 미만 ({len(inserts)} found)"
        )
    target_inserts = inserts[-n_concurrent:]
    target_threads = {r["thread"] for r in target_inserts}

    def parse_ts(ts):
        return datetime.strptime(ts, "%Y-%m-%d %H:%M:%S.%f")

    ts_min = parse_ts(min(r["ts"] for r in target_inserts))
    ts_max = parse_ts(max(r["ts"] for r in target_inserts))
    win_start = ts_min - timedelta(seconds=10)
    win_end = ts_max + timedelta(seconds=10)
    cycle_recs = [r for r in recs
                  if r["thread"] in target_threads
                  and win_start <= parse_ts(r["ts"]) <= win_end]
    return cycle_recs, ts_min, ts_max


def map_users(by_thread):
    thread_to_user = {}
    for thread, thread_recs in by_thread.items():
        for r in thread_recs:
            if r["kind"] == "statement":
                m = INSERT_USER_RE.search(r["sql"])
                if m:
                    thread_to_user[thread] = m.group("user")
                    break
    return thread_to_user


# ════════════════════════════════════════════════════════════════════════
# 시트 채우기 헬퍼
# ════════════════════════════════════════════════════════════════════════
def fill_color(ws, row, kind_label, n_cols):
    if "UPDATE" in kind_label:
        fill = UPDATE_FILL
    elif "INSERT 예약" in kind_label:
        fill = INSERT_FILL
    elif kind_label == "COMMIT":
        fill = COMMIT_FILL
    else:
        return
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).fill = fill


def fill_user_sequence_sheet(ws, by_thread, sorted_user_thread,
                             ts_min, ts_max, sql_transform=None):
    """사용자별 SQL 시퀀스 시트 채우기. sql_transform 으로 SQL 셀 값 변형 가능."""
    row = 1
    title = "F004 시나리오 1 — N 동시 요청의 사용자별 트랜잭션 SQL 시퀀스 (P6Spy)"
    if sql_transform:
        title += "  ※ Oracle SQL Developer 실행형 (타임스탬프 literal 변환)"
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    row += 2

    meta = [
        ("측정 시각",
         f"{ts_min.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]} ~ "
         f"{ts_max.strftime('%H:%M:%S.%f')[:-3]}"),
        ("대상 슬롯", META_SLOT_DESC),
        ("환경", META_ENV_DESC),
        ("결과", META_RESULT_DESC),
        ("로그 출처", str(LOG_PATH)),
        ("중복 처리", "P6Spy wrapper 라인 (url=jdbc:p6spy:...) 만 사용"),
    ]
    for k, v in meta:
        ws.cell(row=row, column=1, value=k).font = META_FONT
        ws.cell(row=row, column=2, value=v)
        row += 1
    row += 1

    for user, thread in sorted_user_thread:
        thread_recs = by_thread[thread]
        n_sql = sum(1 for r in thread_recs if r["kind"] == "statement")
        n_commit = sum(1 for r in thread_recs if r["kind"] == "commit")
        ws.cell(row=row, column=1,
                value=f"■ {user}  ({thread})  — SQL {n_sql}건 + COMMIT {n_commit}건"
                ).font = SECTION_FONT
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        row += 1

        headers = ["#", "시각 (HH:MM:SS.ms)", "SQL 종류", "took (ms)",
                   "SQL (Oracle 실행형)" if sql_transform else "SQL (바인드 박힌 완성형)"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        for idx, r in enumerate(thread_recs, 1):
            kind_label = "COMMIT" if r["kind"] == "commit" else classify_sql(r["sql"])
            ws.cell(row=row, column=1, value=idx
                    ).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=r["ts"][11:]
                    ).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=kind_label)
            ws.cell(row=row, column=4, value=r["took"]
                    ).alignment = Alignment(horizontal="right")
            sql_value = sql_transform(r["sql"]) if sql_transform else r["sql"]
            cell = ws.cell(row=row, column=5, value=sql_value)
            cell.alignment = Alignment(vertical="top", wrap_text=bool(sql_transform))
            fill_color(ws, row, kind_label, 5)
            row += 1
        row += 2

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 140


def fill_interleave_sheet(ws, cycle_recs, thread_to_user):
    row = 1
    ws.cell(row=row, column=1,
            value="N 의 모든 스레드 SQL 을 시간순으로 인터리브 정렬 (race 패턴 관찰)"
            ).font = TITLE_FONT
    row += 2
    headers = ["#", "시각", "사용자", "스레드", "SQL 종류", "took (ms)",
               "SQL (바인드 박힌 완성형)"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    all_lines = sorted(cycle_recs, key=lambda r: r["ts"])
    for idx, r in enumerate(all_lines, 1):
        kind_label = "COMMIT" if r["kind"] == "commit" else classify_sql(r["sql"])
        user = thread_to_user.get(r["thread"], "")
        ws.cell(row=row, column=1, value=idx
                ).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=r["ts"][11:]
                ).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=user
                ).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=r["thread"]
                ).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=kind_label)
        ws.cell(row=row, column=6, value=r["took"]
                ).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=7, value=r["sql"])
        fill_color(ws, row, kind_label, 7)
        row += 1
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 140


def fill_count_sheet(ws, by_thread, sorted_user_thread):
    row = 1
    ws.cell(row=row, column=1, value="사용자별 / SQL 종류별 발행 횟수"
            ).font = TITLE_FONT
    row += 2
    all_kinds_by_user = defaultdict(Counter)
    for user, thread in sorted_user_thread:
        for r in by_thread[thread]:
            kind = "COMMIT" if r["kind"] == "commit" else classify_sql(r["sql"])
            all_kinds_by_user[user][kind] += 1
    ordered_kinds = [
        "SELECT 슬롯 (findById)",
        "SELECT 윈도우 (findById)",
        "SELECT 예약 (logReservationsOnSlot)",
        "SELECT 매물 (CHARTER)",
        "SELECT 회원 (등록자)",
        "COUNT 동일 매물 중복",
        "COUNT 시간 겹침",
        "SEQ 예약 ID",
        "INSERT 예약 (CONFIRMED)",
        "UPDATE 슬롯 (→ RESERVED)",
        "SELECT 구독 (본인 활성)",
        "SEQ 알림 ID",
        "INSERT 알림",
        "COMMIT",
        "OTHER",
    ]
    headers = ["SQL 종류"] + [u for u, _ in sorted_user_thread] + ["합계"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    for kind in ordered_kinds:
        total = sum(all_kinds_by_user[u][kind] for u, _ in sorted_user_thread)
        if total == 0:
            continue
        ws.cell(row=row, column=1, value=kind)
        for ci, (user, _) in enumerate(sorted_user_thread, 2):
            ws.cell(row=row, column=ci, value=all_kinds_by_user[user][kind]
                    ).alignment = Alignment(horizontal="center")
        c = ws.cell(row=row, column=len(sorted_user_thread) + 2, value=total)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        fill_color(ws, row, kind, len(headers))
        row += 1
    ws.column_dimensions["A"].width = 35
    for ci in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14


# ════════════════════════════════════════════════════════════════════════
# 엔트리
# ════════════════════════════════════════════════════════════════════════
def build_workbook(records, n_concurrent):
    cycle_recs, ts_min, ts_max = select_cycle(records, n_concurrent)
    by_thread = defaultdict(list)
    for r in cycle_recs:
        by_thread[r["thread"]].append(r)
    thread_to_user = map_users(by_thread)
    sorted_user_thread = sorted(
        [(u, t) for t, u in thread_to_user.items()],
        key=lambda x: x[0],
    )

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "사용자별 SQL 시퀀스"
    fill_user_sequence_sheet(ws1, by_thread, sorted_user_thread,
                             ts_min, ts_max, sql_transform=None)

    ws2 = wb.create_sheet("사용자별 SQL 시퀀스 (Oracle 실행형)", index=1)
    fill_user_sequence_sheet(ws2, by_thread, sorted_user_thread,
                             ts_min, ts_max, sql_transform=to_oracle_sql)

    ws3 = wb.create_sheet("전체 시간순 인터리브")
    fill_interleave_sheet(ws3, cycle_recs, thread_to_user)

    ws4 = wb.create_sheet("SQL 종류 카운트")
    fill_count_sheet(ws4, by_thread, sorted_user_thread)

    return wb, sorted_user_thread, len(cycle_recs)


def main():
    records = parse_log(LOG_PATH)
    wb, sorted_user_thread, n_records = build_workbook(records, N_CONCURRENT)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"saved: {OUT_PATH}")
    print(f"users mapped: {sorted_user_thread}")
    print(f"records in cycle: {n_records}")


if __name__ == "__main__":
    main()
