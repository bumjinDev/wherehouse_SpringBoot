"""
F004 시나리오 1 — P6Spy SQL 로그 + 애플리케이션 CP 마커를 합쳐
N 동시 요청의 "동시성 타임라인" 엑셀로 정리.

목적(재작성 사유):
    조건부 UPDATE / 비관적 락 회차는 INSERT 가 1건뿐이라(승자만), 옛 스크립트의
    "최근 N 건 INSERT = 한 사이클" 가정이 깨진다. 또 D vs B 의 본질 차이는
    "읽기가 동시냐 직렬이냐 / 락을 SELECT 에서 잡냐 UPDATE 에서 잡냐" 이고,
    그건 각 문장의 took(경과)과 CP1 도달 여부로만 드러난다.
    → 요청 스레드(http-nio-…-exec-N)의 모든 P6Spy 문장과 F004-CP 마커를
       시간순으로 병합해, 선착/후착 트랜잭션의 읽기·쓰기 실행 순서를 한 판에 본다.

사용법:
    py -3.13 build_p6spy_xlsx.py --in "<wherehouse.log>" --out "<결과.xlsx>"
    py -3.13 build_p6spy_xlsx.py --in ... --out ... --n 5 --env "조건부 UPDATE, slot 446"

시트 구성:
    1. 동시성 타임라인        — 전 스레드의 SQL+CP 를 시간순. took 큰 문장(락 대기) 강조. ★핵심
    2. 스레드별 시퀀스        — 트랜잭션별(=스레드별) SQL+CP 순서
    3. 스레드별 SQL (Oracle 실행형) — SQL Developer 붙여넣기용 (타임스탬프 literal 변환)
    4. SQL 종류 카운트        — 스레드별 SQL 종류별 발행 횟수

로깅 이중화 처리:
    과거 설정은 같은 SQL 을 2번(jdbc:p6spy wrapper + jdbc:oracle 실연결) 찍었다.
    (thread, #epoch, kind, sql) 키로 중복 제거 → 단일/이중 로깅 로그 모두 동작.
    application.yml 을 단일 레벨(디코레이터)로 고친 뒤에는 애초에 한 줄만 찍힌다.

요구 패키지: openpyxl (>= 3.1).  실행: py -3.13 (openpyxl 설치 버전)
"""

import argparse
import re
import sys
from collections import defaultdict, Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# ════════════════════════════════════════════════════════════════════════
# 기본 경로 (─-in/--out 미지정 시 폴백)
# ════════════════════════════════════════════════════════════════════════
DEFAULT_IN = (
    r"E:\devSpace\SpringBootProjects\wherehouse_SpringBoot-master"
    r"\wherehouse\log\wherehouse.log"
)
DEFAULT_OUT = (
    r"E:\devSpace\SpringBootProjects\wherehouse_SpringBoot-master"
    r"\docs\13. Wherehouse_고도화_주거지예약시스템\2. 문제해결\1. 테스트시나리오_1"
    r"\트랜잭션로깅용\F004_트랜잭션_타임라인.xlsx"
)

WAIT_THRESHOLD_MS = 10  # took 가 이 값 이상이면 "락 대기" 로 강조

# ════════════════════════════════════════════════════════════════════════
# 정규식
# ════════════════════════════════════════════════════════════════════════
LOG_PREFIX_RE = re.compile(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3}\s")

# P6Spy 한 문장 헤더: url 은 jdbc:p6spy:… / jdbc:oracle:… 둘 다 매칭(\S+)
P6SPY_RE = re.compile(
    r"^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3})\s+"
    r"\[(?P<thread>http-nio-\d+-exec-\d+)\]\s+\w+\s+p6spy\s+-\s+"
    r"#(?P<epoch>\d+)\s+\|\s+took\s+(?P<took>\d+)ms\s+\|\s+(?P<kind>\w+)\s+\|\s+"
    r"connection\s+(?P<conn>\d+)\s*\|\s*url\s+(?P<url>\S+)"
)

# 애플리케이션 측정 마커: [F004-CP1-AFTER-AVAILABLE-CHECK] …
CP_RE = re.compile(
    r"^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3})\s+"
    r"\[(?P<thread>http-nio-\d+-exec-\d+)\]\s+\w+\s+\S+\s+-\s+"
    r"\[F004-(?P<cp>CP\d-[A-Z-]+|REJECT)\]\s+(?P<rest>.*)$"
)

SEARCHER_RE = re.compile(r"searcher\d+")
INSTANT_RE = re.compile(r"instant=\S*?T\d{2}:\d{2}:(?P<sec>\d{2}\.\d+)Z")
NANO_RE = re.compile(r"nano=(?P<nano>\d+)")
STATUS_RE = re.compile(r"status=(?P<st>\w+)")
RCOUNT_RE = re.compile(r"reservationsCount=(?P<n>\d+)")
RES_RE = re.compile(
    r"reservation=\(reservationId=(?P<rid>\d+).*?"
    r"searcherUserId=(?P<user>searcher\d+).*?status=(?P<st>\w+)"
)
# 따옴표로 묶인 '시각' 리터럴(시:분:초 포함)을 Oracle 실행형 TIMESTAMP literal 로 변환하는 패턴.
# p6spy 가 내보내는 모든 변형을 흡수한다(문장 종류 SELECT/UPDATE/INSERT/COUNT 무관):
#   '2026-06-09T23:35:00.000+0900'   ISO + 오프셋   (현재 측정 로그의 형태)
#   '2026-06-09T23:35:00+0900'       소수초 없음
#   '2026-06-09T23:35:00.000Z'       UTC(Z)
#   '2026-06-09 23:35:00.000+09:00'  공백 구분 / 콜론 오프셋
#   '2026-06-09T23:35:00'            오프셋 없음
# → TIMESTAMP '2026-06-09 23:35:00.000'  (T 를 공백으로, 오프셋·Z 제거)
# (?<!TIMESTAMP ) : 이미 변환된 리터럴을 다시 감싸지 않도록(멱등성) 방지.
TS_LITERAL_RE = re.compile(
    r"(?<!TIMESTAMP )'(\d{4}-\d{2}-\d{2})[T ](\d{2}:\d{2}:\d{2}(?:\.\d+)?)"
    r"(?:\s*(?:Z|[+-]\d{2}:?\d{2}))?'"
)

# Hibernate 자동 생성 테이블 별칭 토큰: vse1_0, vwe1_0, vre1_0, pce1_0, me1_0, rse1_0 …
#   (소문자+ 숫자 + '_' + 숫자). 컬럼명(대문자)·따옴표 값('693c…','searcher05','ownerTest1')과
#   겹치지 않는다(값에는 '_숫자' 패턴이 없고 컬럼은 대문자라 \b[a-z]+\d+_\d+\b 에 안 걸림).
HIB_ALIAS_RE = re.compile(r"\b[a-z]+\d+_\d+\b")

CP_LABEL = {
    # 신규 cp() 마커 (createReservation 계측)
    "CP0-ENTER": "CP0 진입",
    "CP1-READ": "CP1 읽기후",
    "CP2-WRITE": "CP2 쓰기후",
    "CP3-INSERT": "CP3 등록후",
    "REJECT": "REJECT 거부",
    # 구 logSlotSnapshot 마커 (하위호환)
    "CP1-AFTER-AVAILABLE-CHECK": "CP1 확인조회",
    "CP2-AFTER-SLOT-UPDATE": "CP2 점유",
    "CP3-AFTER-RESERVATION-INSERT": "CP3 등록",
}

# Spring 트랜잭션 라이프사이클 (begin/commit/rollback).
# createReservation 의 외곽 @Transactional 만 "Creating new transaction" 으로 잡힌다
# (repository 호출은 "Participating in existing transaction" 이라 begin 미발생) → 요청당 1 begin.
TX_BEGIN_RE = re.compile(
    r"^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3})\s+"
    r"\[(?P<thread>http-nio-\d+-exec-\d+)\]\s+\w+\s+\S+\s+-\s+"
    r"Creating new transaction with name \[(?P<method>[^\]]+)\]:\s*(?P<props>.+)$"
)
TX_COMMIT_RE = re.compile(
    r"^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3})\s+"
    r"\[(?P<thread>http-nio-\d+-exec-\d+)\]\s+\w+\s+\S+\s+-\s+Committing JPA transaction"
)
TX_ROLLBACK_RE = re.compile(
    r"^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3})\s+"
    r"\[(?P<thread>http-nio-\d+-exec-\d+)\]\s+\w+\s+\S+\s+-\s+Rolling back JPA transaction"
)

# ════════════════════════════════════════════════════════════════════════
# 스타일
# ════════════════════════════════════════════════════════════════════════
TITLE_FONT = Font(bold=True, size=13)
META_FONT = Font(bold=True)
SECTION_FILL = PatternFill("solid", fgColor="305496")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="8EA9DB")
HEADER_FONT = Font(bold=True, color="FFFFFF")
UPDATE_FILL = PatternFill("solid", fgColor="FFE699")  # UPDATE 슬롯
INSERT_FILL = PatternFill("solid", fgColor="C6EFCE")  # INSERT 예약
COMMIT_FILL = PatternFill("solid", fgColor="D9D9D9")  # COMMIT
ROLLBACK_FILL = PatternFill("solid", fgColor="FCE4D6")  # ROLLBACK
CP_FILL = PatternFill("solid", fgColor="DDEBF7")       # CP 마커
TX_FILL = PatternFill("solid", fgColor="E4DFEC")       # 트랜잭션 begin/commit/rollback (JPA)
REJECT_FILL = PatternFill("solid", fgColor="FFC7CE")   # REJECT 거부 마커
WAIT_FONT = Font(bold=True, color="C00000")           # took >= 임계 (락 대기)


# ════════════════════════════════════════════════════════════════════════
# SQL 분류 / 변환
# ════════════════════════════════════════════════════════════════════════
def classify_sql(sql: str) -> str:
    s = sql.lower()
    if s.startswith("select") and "visit_slot" in s and "for update" in s:
        return "SELECT 슬롯 FOR UPDATE (비관적 락)"
    if s.startswith("select") and "visit_slot" in s and "join" not in s and "order by" not in s:
        return "SELECT 슬롯 (findById)"
    if s.startswith("select") and "visit_window" in s and "count" not in s:
        return "SELECT 윈도우 (findById)"
    if s.startswith("select") and "visit_reservation" in s and "order by" in s:
        return "SELECT 예약 (logReservationsOnSlot)"
    if s.startswith("select") and "properties_charter" in s:
        return "SELECT 매물 (CHARTER)"
    if s.startswith("select") and "membertbl" in s:
        return "SELECT 회원 (등록자)"
    if s.startswith("select count") and "visit_window" in s and "lease_type" in s:
        return "COUNT 동일 매물 중복"
    if s.startswith("select count") and "start_time" in s and "end_time" in s:
        return "COUNT 시간 겹침"
    if "nextval" in s and "seq_visit_reservation" in s:
        return "SEQ 예약 ID"
    if "nextval" in s and "seq_visit_notification" in s:
        return "SEQ 알림 ID"
    if s.startswith("select") and "reopen_subscription" in s:
        return "SELECT 구독 (본인 활성)"
    if s.startswith("insert into visit_reservation"):
        return "INSERT 예약 (CONFIRMED)"
    if s.startswith("insert into visit_notification"):
        return "INSERT 알림"
    if s.startswith("update visit_slot"):
        return "UPDATE 슬롯 (→ RESERVED)"
    if s.upper().startswith("COMMIT"):
        return "COMMIT"
    return "OTHER"


def clean_hibernate_aliases(sql: str) -> str:
    """SELECT 의 Hibernate 자동 별칭을 사람이 쓴 Oracle SQL 형태로 정리.

    - 단일 테이블(별칭 1개)   : 별칭 제거.  vse1_0.SLOT_ID → SLOT_ID,  from VISIT_SLOT vse1_0 → from VISIT_SLOT
    - 다중 테이블(JOIN, 2개+) : 짧고 깔끔한 별칭(t1, t2 …)으로 치환(조인은 별칭이 있어야 실행 가능).
    INSERT/UPDATE 처럼 별칭이 없는 문장은 그대로 둔다(HIB_ALIAS_RE 매치 0건).
    """
    aliases = []
    for a in HIB_ALIAS_RE.findall(sql):
        if a not in aliases:
            aliases.append(a)                                    # 등장 순서 보존
    if not aliases:
        return sql
    if len(aliases) == 1:
        a = aliases[0]
        sql = re.sub(r"\b" + re.escape(a) + r"\.", "", sql)      # 'alias.' 한정자 제거
        sql = re.sub(r"\s+" + re.escape(a) + r"\b", "", sql)     # FROM 절 별칭 선언 제거
        return sql
    for i, a in enumerate(aliases, 1):                           # JOIN: t1, t2, … 로 치환
        sql = re.sub(r"\b" + re.escape(a) + r"\b", f"t{i}", sql)
    return sql


def to_oracle_sql(sql: str) -> str:
    """P6Spy binding SQL → Oracle SQL Developer 실행 가능 + 사람이 읽는 형태.

    문장 종류(SELECT/UPDATE/INSERT/COUNT)와 무관하게 동일하게 적용된다:
      - COMMIT/ROLLBACK → 대문자 + ';'
      - 모든 따옴표 '시각' 리터럴(오프셋·Z·소수초 유무 무관) → Oracle TIMESTAMP literal
        (T 를 공백으로, 오프셋/Z 제거). SLOT_ID·STATUS 같은 비(非)시각 리터럴은 건드리지 않는다.
      - Hibernate 자동 별칭(vse1_0 …) 정리 → 단일 테이블은 별칭 제거, JOIN 은 t1/t2 로 치환.
    """
    if sql.strip().upper().startswith(("COMMIT", "ROLLBACK")):
        return sql.strip().upper() + ";"
    sql = TS_LITERAL_RE.sub(r"TIMESTAMP '\1 \2'", sql)
    sql = clean_hibernate_aliases(sql)
    sql = sql.rstrip().rstrip(";") + ";"
    return sql


def short_thread(thread: str) -> str:
    """http-nio-8185-exec-1 → exec-1."""
    m = re.search(r"exec-\d+", thread)
    return m.group(0) if m else thread


def parse_ts(ts: str) -> datetime:
    return datetime.strptime(ts, "%Y-%m-%d %H:%M:%S.%f")


# ════════════════════════════════════════════════════════════════════════
# 로그 파싱
# ════════════════════════════════════════════════════════════════════════
def cp_detail(rest: str):
    """CP 라인 본문 → (상세문자열, 정밀시각μs|None, nano|None, 추출유저|None).

    신규 cp() 포맷   : "thread=… nano=<n> <detail>"  (detail = status/affected/resId/reason 등)
    구 logSlotSnapshot: "… instant=…Z slot=(…)/reservationsCount/reservation=(…)"
    """
    m = NANO_RE.search(rest)
    if m:  # 신규 cp()
        nano = int(m.group("nano"))
        detail = rest[m.end():].strip() or "—"
        um = SEARCHER_RE.search(detail)
        return detail, None, nano, (um.group(0) if um else None)
    # 구 포맷 (하위호환)
    inst = INSTANT_RE.search(rest)
    instant = inst.group("sec") if inst else None
    if "slot=(" in rest:
        st = STATUS_RE.search(rest)
        return (f"슬롯 status={st.group('st')}" if st else "슬롯 스냅샷"), instant, None, None
    rm = RES_RE.search(rest)
    if rm:
        return (f"예약 {rm.group('user')} status={rm.group('st')} (id={rm.group('rid')})",
                instant, None, rm.group("user"))
    rc = RCOUNT_RE.search(rest)
    if rc:
        return f"reservationsCount={rc.group('n')}", instant, None, None
    return rest[:80], instant, None, None


def parse_log(log_path: str):
    """요청 스레드(exec-N)의 P6Spy 문장/commit/rollback + F004-CP 마커를 레코드로.

    스케줄러([scheduling-N]) 등 다른 스레드는 정규식 단계에서 자동 제외된다.
    (thread, epoch, kind, sql) 키로 P6Spy 이중로깅을 중복 제거한다.
    """
    with open(log_path, encoding="utf-8") as f:
        lines = f.read().splitlines()

    records = []
    seen = set()
    i = 0
    n = len(lines)
    while i < n:
        line = lines[i]

        m = P6SPY_RE.match(line)
        if m:
            ts = m.group("ts")
            thread = m.group("thread")
            took = int(m.group("took"))
            kind = m.group("kind")
            epoch = m.group("epoch")
            url = m.group("url")

            if kind == "statement":
                # 다음 비-로그 라인들(= SQL 조각)을 모은다. 마지막 = 값 바인딩 SQL.
                frags = []
                j = i + 1
                while j < n and not LOG_PREFIX_RE.match(lines[j]):
                    if lines[j].strip():
                        frags.append(lines[j].strip())
                    j += 1
                binding = frags[-1].rstrip(";").strip() if frags else ""
                placeholder = frags[0] if frags else ""
                key = (thread, epoch, "statement", binding)
                if key not in seen:
                    seen.add(key)
                    records.append({
                        "ts": ts, "thread": thread, "took": took, "kind": "statement",
                        "epoch": epoch, "url": url, "placeholder": placeholder,
                        "sql": binding, "cp": None, "detail": None, "instant": None,
                        "idx": i,
                    })
                i = j
                continue

            if kind in ("commit", "rollback"):
                key = (thread, epoch, kind, "")
                if key not in seen:
                    seen.add(key)
                    records.append({
                        "ts": ts, "thread": thread, "took": took, "kind": kind,
                        "epoch": epoch, "url": url, "placeholder": "",
                        "sql": kind.upper(), "cp": None, "detail": None, "instant": None,
                        "idx": i,
                    })
                # commit/rollback 뒤 빈 줄 흡수
                j = i + 1
                while j < n and not LOG_PREFIX_RE.match(lines[j]) and not lines[j].strip():
                    j += 1
                i = j
                continue

            i += 1
            continue

        cp = CP_RE.match(line)
        if cp:
            label = CP_LABEL.get(cp.group("cp"), cp.group("cp"))
            detail, instant, nano, _user = cp_detail(cp.group("rest"))
            records.append({
                "ts": cp.group("ts"), "thread": cp.group("thread"), "took": None,
                "kind": "cp", "epoch": None, "url": None, "placeholder": "",
                "sql": "", "cp": label, "detail": detail, "instant": instant,
                "nano": nano, "idx": i,
            })
            i += 1
            continue

        tb = TX_BEGIN_RE.match(line)
        if tb:
            props = tb.group("props").strip().replace(",", ", ")
            if "ISOLATION_DEFAULT" in props:
                props += " (Oracle 기본 = READ COMMITTED)"
            method = tb.group("method").rsplit(".", 1)[-1]
            records.append({
                "ts": tb.group("ts"), "thread": tb.group("thread"), "took": None,
                "kind": "tx-begin", "epoch": None, "url": None, "placeholder": "",
                "sql": "", "cp": "TX 시작", "detail": f"{props}  [{method}]",
                "instant": None, "idx": i,
            })
            i += 1
            continue
        tc = TX_COMMIT_RE.match(line)
        if tc:
            records.append({
                "ts": tc.group("ts"), "thread": tc.group("thread"), "took": None,
                "kind": "tx-commit", "epoch": None, "url": None, "placeholder": "",
                "sql": "", "cp": "TX 커밋(JPA)", "detail": "JPA 트랜잭션 커밋",
                "instant": None, "idx": i,
            })
            i += 1
            continue
        tr = TX_ROLLBACK_RE.match(line)
        if tr:
            records.append({
                "ts": tr.group("ts"), "thread": tr.group("thread"), "took": None,
                "kind": "tx-rollback", "epoch": None, "url": None, "placeholder": "",
                "sql": "", "cp": "TX 롤백(JPA)", "detail": "JPA 트랜잭션 롤백",
                "instant": None, "idx": i,
            })
            i += 1
            continue

        i += 1

    return records


def map_users(records):
    """스레드 → searcher 식별자 (best-effort).

    조건부 UPDATE 는 모든 스레드가 COUNT 쿼리에 searcher 를 실으므로 전부 매핑되고,
    비관적 락의 후착은 SELECT FOR UPDATE 에서 막혀 COUNT 전에 throw 하므로 매핑 안 될 수 있다
    (그 경우 스레드명으로 표기)."""
    thread_to_user = {}
    for r in records:
        t = r["thread"]
        if t in thread_to_user:
            continue
        text = r["sql"] or r["detail"] or ""
        m = SEARCHER_RE.search(text)
        if m:
            thread_to_user[t] = m.group(0)
    return thread_to_user


def thread_label(thread, thread_to_user):
    u = thread_to_user.get(thread)
    return f"{u} ({short_thread(thread)})" if u else short_thread(thread)


def type_label(r):
    if r["kind"] == "cp" or r["kind"].startswith("tx-"):
        return r["cp"]
    if r["kind"] == "commit":
        return "COMMIT (JDBC)"
    if r["kind"] == "rollback":
        return "ROLLBACK (JDBC)"
    return classify_sql(r["sql"])


def row_fill(r):
    if r["kind"] == "cp":
        return REJECT_FILL if r["cp"].startswith("REJECT") else CP_FILL
    if r["kind"].startswith("tx-"):
        return TX_FILL
    if r["kind"] == "commit":
        return COMMIT_FILL
    if r["kind"] == "rollback":
        return ROLLBACK_FILL
    label = classify_sql(r["sql"])
    if "UPDATE" in label:
        return UPDATE_FILL
    if "INSERT 예약" in label:
        return INSERT_FILL
    return None


# ════════════════════════════════════════════════════════════════════════
# 시트 — 1. 동시성 타임라인 (핵심)
# ════════════════════════════════════════════════════════════════════════
def fill_timeline_sheet(ws, records, thread_to_user, meta):
    row = 1
    ws.cell(row=row, column=1,
            value="F004 시나리오 1 — 동시성 타임라인 (요청 스레드 전체 SQL + CP 마커, 시간순)"
            ).font = TITLE_FONT
    row += 2
    for k, v in meta:
        ws.cell(row=row, column=1, value=k).font = META_FONT
        ws.cell(row=row, column=2, value=v)
        row += 1
    ws.cell(row=row, column=1,
            value=f"읽는 법: took ≥ {WAIT_THRESHOLD_MS}ms(빨강) = 행 락 대기. "
                  "CP1 도달 스레드 수 = '동시에 AVAILABLE 로 읽은' 수. "
                  "TX 시작/커밋(JPA)/롤백(JPA) 행 = @Transactional 경계(격리·전파 표시)."
            ).font = Font(italic=True, color="808080")
    row += 2

    headers = ["#", "시각(ms)", "정밀시각", "스레드", "사용자/스레드",
               "유형", "took(ms)", "상세 (SQL / CP)"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    cp_nanos = [r["nano"] for r in records if r.get("nano")]
    base_nano = min(cp_nanos) if cp_nanos else None
    ordered = sorted(records, key=lambda r: (r["ts"], r["idx"]))
    for seq, r in enumerate(ordered, 1):
        is_wait = r["kind"] == "statement" and r["took"] is not None and r["took"] >= WAIT_THRESHOLD_MS
        detail = r["sql"] if r["kind"] in ("statement", "commit", "rollback") else r["detail"]
        if r.get("nano") and base_nano:
            precise = f"+{(r['nano'] - base_nano) / 1e6:.3f}ms"   # 버스트 시작 대비 nano 상대시각
        else:
            precise = r["instant"] or ""
        cells = [
            (seq, "center"),
            (r["ts"][11:], "center"),
            (precise, "center"),
            (short_thread(r["thread"]), "center"),
            (thread_label(r["thread"], thread_to_user), "left"),
            (type_label(r), "left"),
            (r["took"] if r["took"] is not None else "", "right"),
            (detail, "left"),
        ]
        fill = row_fill(r)
        for ci, (val, align) in enumerate(cells, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.alignment = Alignment(horizontal=align, vertical="top",
                                    wrap_text=(ci == 8 and r["kind"] == "statement"))
            if fill:
                c.fill = fill
        if is_wait:
            tc = ws.cell(row=row, column=7)
            tc.font = WAIT_FONT
            ws.cell(row=row, column=8).font = WAIT_FONT
        row += 1

    widths = [6, 14, 12, 9, 22, 30, 10, 130]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A" + str(row - len(ordered))


# ════════════════════════════════════════════════════════════════════════
# 시트 — 2. 스레드별 시퀀스
# ════════════════════════════════════════════════════════════════════════
def fill_per_thread_sheet(ws, by_thread, threads_sorted, thread_to_user,
                          sql_transform=None):
    row = 1
    title = "스레드(=트랜잭션)별 SQL + CP 시퀀스"
    if sql_transform:
        title += "  ※ Oracle SQL Developer 실행형 (타임스탬프 literal 변환)"
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    row += 2

    for thread in threads_sorted:
        recs = sorted(by_thread[thread], key=lambda r: (r["ts"], r["idx"]))
        n_sql = sum(1 for r in recs if r["kind"] == "statement")
        n_commit = sum(1 for r in recs if r["kind"] == "commit")
        n_rb = sum(1 for r in recs if r["kind"] == "rollback")
        end = "COMMIT" if n_commit else ("ROLLBACK" if n_rb else "—")
        ws.cell(row=row, column=1,
                value=f"■ {thread_label(thread, thread_to_user)}  — SQL {n_sql}건 / 종료 {end}"
                ).font = SECTION_FONT
        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        row += 1

        headers = ["#", "시각", "유형", "took(ms)",
                   "SQL (Oracle 실행형)" if sql_transform else "상세 (SQL / CP)"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        for idx, r in enumerate(recs, 1):
            if sql_transform and (r["kind"] == "cp" or r["kind"].startswith("tx-")):
                # Oracle 실행형 시트는 실행 가능한 SQL/commit/rollback 만 (CP·트랜잭션 마커 제외)
                continue
            is_wait = r["kind"] == "statement" and r["took"] is not None and r["took"] >= WAIT_THRESHOLD_MS
            if r["kind"] == "statement":
                detail = sql_transform(r["sql"]) if sql_transform else r["sql"]
            elif r["kind"] in ("commit", "rollback"):
                detail = r["sql"] + (";" if sql_transform else "")
            else:
                detail = r["detail"]
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=r["ts"][11:]).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=type_label(r))
            ws.cell(row=row, column=4,
                    value=r["took"] if r["took"] is not None else ""
                    ).alignment = Alignment(horizontal="right")
            c5 = ws.cell(row=row, column=5, value=detail)
            c5.alignment = Alignment(vertical="top", wrap_text=bool(sql_transform))
            fill = row_fill(r)
            if fill:
                for c in range(1, 6):
                    ws.cell(row=row, column=c).fill = fill
            if is_wait:
                ws.cell(row=row, column=4).font = WAIT_FONT
            row += 1
        row += 2

    widths = [6, 16, 30, 10, 140]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ════════════════════════════════════════════════════════════════════════
# 시트 — 4. SQL 종류 카운트
# ════════════════════════════════════════════════════════════════════════
def fill_count_sheet(ws, by_thread, threads_sorted, thread_to_user):
    row = 1
    ws.cell(row=row, column=1, value="스레드별 / SQL 종류별 발행 횟수").font = TITLE_FONT
    row += 2

    counts = {t: Counter() for t in threads_sorted}
    for t in threads_sorted:
        for r in by_thread[t]:
            if r["kind"] == "cp" or r["kind"].startswith("tx-"):
                continue
            counts[t][type_label(r)] += 1

    ordered_kinds = [
        "SELECT 슬롯 FOR UPDATE (비관적 락)",
        "SELECT 슬롯 (findById)",
        "SELECT 윈도우 (findById)",
        "SELECT 예약 (logReservationsOnSlot)",
        "SELECT 매물 (CHARTER)",
        "SELECT 회원 (등록자)",
        "COUNT 동일 매물 중복",
        "COUNT 시간 겹침",
        "SEQ 예약 ID",
        "INSERT 예약 (CONFIRMED)",
        "UPDATE 슬롯 (→ RESERVED)",
        "SELECT 구독 (본인 활성)",
        "SEQ 알림 ID",
        "INSERT 알림",
        "COMMIT (JDBC)",
        "ROLLBACK (JDBC)",
        "OTHER",
    ]
    headers = ["SQL 종류"] + [thread_label(t, thread_to_user) for t in threads_sorted] + ["합계"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    seen_kinds = set()
    for kind in ordered_kinds:
        total = sum(counts[t][kind] for t in threads_sorted)
        if total == 0:
            continue
        seen_kinds.add(kind)
        ws.cell(row=row, column=1, value=kind)
        for ci, t in enumerate(threads_sorted, 2):
            ws.cell(row=row, column=ci, value=counts[t][kind]
                    ).alignment = Alignment(horizontal="center")
        c = ws.cell(row=row, column=len(threads_sorted) + 2, value=total)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        row += 1

    ws.column_dimensions["A"].width = 34
    for ci in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16


# ════════════════════════════════════════════════════════════════════════
# 빌드
# ════════════════════════════════════════════════════════════════════════
def build_workbook(records, n_expected, in_path, env):
    if not records:
        raise RuntimeError("요청 스레드(http-nio-…-exec-N)의 P6Spy/CP 레코드가 없다. "
                           "로그 경로·측정 여부 확인.")
    by_thread = defaultdict(list)
    for r in records:
        by_thread[r["thread"]].append(r)
    threads_sorted = sorted(by_thread.keys(),
                            key=lambda t: int(re.search(r"exec-(\d+)", t).group(1))
                            if re.search(r"exec-(\d+)", t) else 0)
    thread_to_user = map_users(records)

    ts_all = [r["ts"] for r in records]
    cp1_threads = {r["thread"] for r in records
                   if r["kind"] == "cp" and r["cp"].startswith("CP1")}
    span = f"{min(ts_all)[11:]} ~ {max(ts_all)[11:]}"
    meta = [
        ("측정 시각", span),
        ("입력 로그", in_path),
        ("회차 환경", env or "(미지정 — --env 로 입력)"),
        ("스레드 수 (관측/기대)", f"{len(threads_sorted)} / {n_expected}"),
        ("CP1(읽기후) 도달 스레드 수", f"{len(cp1_threads)}  "
            "(= 동시에 슬롯을 AVAILABLE 로 읽은 트랜잭션 수)"),
    ]

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "동시성 타임라인"
    fill_timeline_sheet(ws1, records, thread_to_user, meta)

    ws2 = wb.create_sheet("스레드별 시퀀스")
    fill_per_thread_sheet(ws2, by_thread, threads_sorted, thread_to_user)

    ws3 = wb.create_sheet("스레드별 SQL (Oracle 실행형)")
    fill_per_thread_sheet(ws3, by_thread, threads_sorted, thread_to_user,
                          sql_transform=to_oracle_sql)

    ws4 = wb.create_sheet("SQL 종류 카운트")
    fill_count_sheet(ws4, by_thread, threads_sorted, thread_to_user)

    return wb, threads_sorted, thread_to_user, len(cp1_threads)


def main():
    ap = argparse.ArgumentParser(
        description="P6Spy SQL 로그 + F004 CP 마커 → 동시성 타임라인 엑셀")
    ap.add_argument("--in", dest="in_path", default=DEFAULT_IN,
                    help="입력 로그 파일 경로 (wherehouse.log)")
    ap.add_argument("--out", dest="out_path", default=DEFAULT_OUT,
                    help="출력 xlsx 경로")
    ap.add_argument("--n", dest="n", type=int, default=5,
                    help="동시 요청 수 (사니티 체크용, 기본 5)")
    ap.add_argument("--env", dest="env", default="",
                    help="회차 환경 설명 (시트 상단 메타)")
    args = ap.parse_args()

    records = parse_log(args.in_path)
    wb, threads_sorted, thread_to_user, cp1_n = build_workbook(
        records, args.n, args.in_path, args.env)

    from pathlib import Path
    Path(args.out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out_path)

    print(f"saved: {args.out_path}")
    print(f"threads ({len(threads_sorted)}): "
          + ", ".join(thread_label(t, thread_to_user) for t in threads_sorted))
    print(f"records: {len(records)}  /  CP1 도달 스레드: {cp1_n} (= 동시 AVAILABLE 읽기 수)")
    if len(threads_sorted) != args.n:
        print(f"  ⚠ 스레드 수({len(threads_sorted)}) != 기대({args.n}) — "
              "로그에 다른 요청이 섞였는지 확인.")


if __name__ == "__main__":
    main()
