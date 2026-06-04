#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
F004 슬롯 예약 동시성 측정 — 다회차 비교 보고서 생성기.

여러 측정 회차의 wherehouse.log 를 받아 단일 엑셀로 정리한다.
회차 = "동일한 코드 + 동일한 부하 시나리오 + 다른 동시성 제어 환경" 의 한 차례.
예: 1회차 = unique 인덱스 적용, 2회차 = unique 인덱스 제거.

사용:
    py -3.13 report.py \
        --log "1_unique적용=결과/Step1_unique적용_T1/wherehouse.log" \
        --log "2_unique제거=결과/Step2_unique삭제_T1/wherehouse.log" \
        --out 결과/F004_측정결과.xlsx

--log 옵션은 여러 번 지정 가능하며 NAME=PATH 형식이다. NAME 은 엑셀의 시트 이름·
표 컬럼 이름으로 사용되므로 짧고 의미 있게 (Excel 시트 이름 제약 31자) 적는다.
"""
from __future__ import annotations

import argparse
import re
import statistics
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ════════════════════════════════════════════════════════════════════
# 1. 로그 파싱 — 한 회차의 wherehouse.log 를 dict 리스트로
# ════════════════════════════════════════════════════════════════════

LINE_RE = re.compile(
    r'^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3})\s+'
    r'\[[^\]]+\]\s+\w+\s+\S+\s+-\s+'
    r'\[F004-(?P<cp>[^\]]+)\]\s+'
    r'thread=\[name=(?P<thr>[^,]+),\s*id=(?P<tid>\d+)\]\s+'
    r'nanoTime=(?P<nt>\d+)'
    r'(?:\s+instant=\S+)?'
    r'\s+(?P<rest>.*)$'
)

SLOT_BODY_RE = re.compile(r'^slot=\((?P<body>.*)\)\s*$')
RES_HDR_RE   = re.compile(r'^slotId=(\d+)\s+reservationsCount=(\d+)\s*$')
RES_ROW_RE   = re.compile(
    r'reservation=\(reservationId=(\d+),\s*slotId=\d+,\s*searcherUserId=(\S+?),\s*status=(\w+)'
)
QUERY_REQ_RE = re.compile(
    r'\[VISIT_SLOT_QUERY_REQ\]\s+propertyId=(?P<pid>\S+?),\s+leaseType=(?P<lt>\S+)'
)
BACKSTOP_RE  = re.compile(r'\[VISIT_INTEGRITY_BACKSTOP\]')
ORA_UQ_RE    = re.compile(
    r'ORA-00001:.*?(?P<cn>UQ_VISIT_RESERVATION_CONFIRMED_SLOT|UQ_REOPEN_SUBSCRIPTION_ACTIVE|UQ_VISIT_SLOT_WINDOW_START)'
)
ORA_GENERIC_RE = re.compile(r'ORA-\d{5}')

# 측정 단계 라벨 — 로그의 영문 라벨을 한국어로 정리
CHECKPOINT_LABEL = {
    'CP1-AFTER-AVAILABLE-CHECK':    '확인 조회',
    'CP2-AFTER-SLOT-UPDATE':        '점유',
    'CP3-AFTER-RESERVATION-INSERT': '등록',
}
CHECKPOINT_SHORT = {
    'CP1-AFTER-AVAILABLE-CHECK':    'CHECK',
    'CP2-AFTER-SLOT-UPDATE':        'UPDATE',
    'CP3-AFTER-RESERVATION-INSERT': 'INSERT',
}

LEASE_KOR = {
    'CHARTER': '전세',
    'MONTHLY': '월세',
}

SLOT_STATUS_KOR = {
    'AVAILABLE': '예약 가능',
    'RESERVED':  '예약됨',
    'CLOSED':    '종료됨',
    'WITHDRAWN': '철회됨',
}


def kor_lease(v):
    if v is None:
        return None
    return LEASE_KOR.get(v, v)


def kor_slot_status(v):
    if v is None:
        return None
    return SLOT_STATUS_KOR.get(v, v)


def parse_kv(body: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for pair in body.split(', '):
        if '=' in pair:
            k, v = pair.split('=', 1)
            out[k.strip()] = v.strip()
    return out


def norm(v: Optional[str]) -> Optional[str]:
    if v is None or v == 'null':
        return None
    return v


def collect_property_fallback(path: Path) -> Tuple[Optional[str], Optional[str]]:
    pairs = set()
    with path.open('r', encoding='utf-8', errors='replace') as f:
        for line in f:
            m = QUERY_REQ_RE.search(line)
            if m:
                pairs.add((m.group('pid'), m.group('lt')))
    return next(iter(pairs)) if len(pairs) == 1 else (None, None)


def parse_log_events(path: Path) -> List[Dict[str, Any]]:
    events: List[Dict[str, Any]] = []
    with path.open('r', encoding='utf-8', errors='replace') as f:
        for line in f:
            m = LINE_RE.match(line.rstrip('\n'))
            if not m:
                continue
            cp_raw = m.group('cp')
            ev: Dict[str, Any] = {
                'log_ts':      m.group('ts'),
                'cp_raw':      cp_raw,
                'cp_short':    CHECKPOINT_SHORT.get(cp_raw, cp_raw),
                'cp_label':    CHECKPOINT_LABEL.get(cp_raw, cp_raw),
                'thr':         m.group('thr'),
                'tid':         int(m.group('tid')),
                'nano':        int(m.group('nt')),
                'slot_id':     None, 'slot_status': None,
                'property_id': None, 'lease_type':  None,
                'start_time':  None, 'end_time':    None,
                'res_count':   None,
                'res_id':      None, 'searcher':    None, 'res_status': None,
            }
            rest = m.group('rest')
            sm = SLOT_BODY_RE.match(rest)
            rh = RES_HDR_RE.match(rest)
            rr = RES_ROW_RE.search(rest)
            if sm:
                kv = parse_kv(sm.group('body'))
                ev['slot_id']     = int(kv['slotId']) if 'slotId' in kv else None
                ev['slot_status'] = norm(kv.get('status'))
                ev['property_id'] = norm(kv.get('propertyId'))
                ev['lease_type']  = norm(kv.get('leaseType'))
                ev['start_time']  = norm(kv.get('startTime'))
                ev['end_time']    = norm(kv.get('endTime'))
            elif rh:
                ev['slot_id']   = int(rh.group(1))
                ev['res_count'] = int(rh.group(2))
            elif rr:
                ev['res_id']     = int(rr.group(1))
                ev['searcher']   = rr.group(2)
                ev['res_status'] = rr.group(3)
            events.append(ev)
    return events


def group_requests(events: List[Dict[str, Any]]) -> List[List[Dict[str, Any]]]:
    events_sorted = sorted(events, key=lambda e: (e['thr'], e['nano']))
    groups: List[List[Dict[str, Any]]] = []
    current: List[Dict[str, Any]] = []
    prev_thr: Optional[str] = None
    for e in events_sorted:
        is_new = (e['cp_short'] == 'CHECK' and e['slot_status'] is not None)
        if (e['thr'] != prev_thr) or is_new:
            if current:
                groups.append(current)
                current = []
        current.append(e)
        prev_thr = e['thr']
    if current:
        groups.append(current)
    return groups


def short_time(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    return s.split('T', 1)[1] if 'T' in s else s


def summarize_request(seq: int,
                      group: List[Dict[str, Any]],
                      fb_pid: Optional[str],
                      fb_lt: Optional[str]) -> Dict[str, Any]:
    check_slot   = next((e for e in group if e['cp_short'] == 'CHECK'  and e['slot_status']), None)
    check_count  = next((e for e in group if e['cp_short'] == 'CHECK'  and e['res_count'] is not None), None)
    update_slot  = next((e for e in group if e['cp_short'] == 'UPDATE' and e['slot_status']), None)
    insert_count = next((e for e in group if e['cp_short'] == 'INSERT' and e['res_count'] is not None), None)

    if insert_count:
        reached = '등록 완료'
    elif update_slot:
        reached = '점유 직후'
    else:
        reached = '확인 조회 직후'

    elapsed_ns: Optional[int] = None
    if check_slot and insert_count:
        elapsed_ns = insert_count['nano'] - check_slot['nano']

    after_count = (insert_count or {}).get('res_count')
    if after_count is None:
        rule = '등록 단계 미도달'
    elif after_count == 1:
        rule = '정상'
    else:
        rule = f'위반 ({after_count}건 관찰)'

    property_id = (check_slot or {}).get('property_id') or fb_pid
    lease_type  = (check_slot or {}).get('lease_type')  or fb_lt
    start_t = short_time((check_slot or {}).get('start_time'))
    end_t   = short_time((check_slot or {}).get('end_time'))
    slot_time = f'{start_t} ~ {end_t}' if (start_t and end_t) else None

    return {
        '요청 번호':                seq,
        '처리 스레드명':             group[0]['thr'],
        '매물 식별자':               property_id,
        '임대 유형':                 kor_lease(lease_type),
        '슬롯 번호':                 (check_slot or check_count or {}).get('slot_id'),
        '방문 예정 시간':            slot_time,
        '요청 시작 시각':            (check_slot or {}).get('log_ts'),
        '요청 종료 시각':            (insert_count or {}).get('log_ts'),
        '확인→등록 응답시간 (ns)':   elapsed_ns,
        '진행 단계':                 reached,
        '확인 조회 시점 슬롯 상태':  kor_slot_status((check_slot or {}).get('slot_status')),
        '점유 직후 슬롯 상태':       kor_slot_status((update_slot or {}).get('slot_status')),
        '등록 전 예약 수':           (check_count or {}).get('res_count'),
        '등록 후 예약 수':           after_count,
        '정합성 판정':               rule,
    }


def count_anomalies(path: Path) -> Dict[str, int]:
    """한 회차의 DB 거부·ORA 발생 카운트."""
    db_reject = 0
    ora_uq = 0
    ora_generic = 0
    with path.open('r', encoding='utf-8', errors='replace') as f:
        for line in f:
            if BACKSTOP_RE.search(line):
                db_reject += 1
            if ORA_UQ_RE.search(line):
                ora_uq += 1
            elif ORA_GENERIC_RE.search(line):
                ora_generic += 1
    return {
        'db_reject': db_reject,
        'ora_uq': ora_uq,
        'ora_other': ora_generic,
    }


# ════════════════════════════════════════════════════════════════════
# 2. 원본 로그 시트용 — 메시지 필드 파싱
# ════════════════════════════════════════════════════════════════════

RAW_LOG_RE = re.compile(
    r'^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3})\s+'
    r'\[(?P<thr>[^\]]+)\]\s+'
    r'(?P<level>\w+)\s+'
    r'(?P<logger>\S+)\s+-\s+'
    r'(?P<message>(?:\[F004-|\[VISIT_RESERVATION_|\[VISIT_INTEGRITY_BACKSTOP).*)$'
)


def detect_phase(message: str) -> str:
    if 'CP1-AFTER-AVAILABLE-CHECK' in message:
        return '1. 확인 조회 직후'
    if 'CP2-AFTER-SLOT-UPDATE' in message:
        return '2. 점유 직후'
    if 'CP3-AFTER-RESERVATION-INSERT' in message:
        return '3. 등록 직후'
    if 'VISIT_RESERVATION_CREATE_REQ' in message:
        return '요청 진입 (컨트롤러)'
    if 'VISIT_RESERVATION_CONFIRMED' in message:
        return '예약 확정 (서비스 완료)'
    if 'VISIT_INTEGRITY_BACKSTOP' in message:
        return 'DB 단 거부 (무결성 위반)'
    return '기타'


_NANO_INSTANT_RE = re.compile(r'nanoTime=(?P<nano>\d+)(?:\s+instant=(?P<inst>\S+))?')
_SLOT_PAYLOAD_RE = re.compile(r'slot=\((?P<body>[^)]*)\)')
_RES_PAYLOAD_RE  = re.compile(r'reservation=\((?P<body>[^)]*)\)')
_RES_COUNT_RE    = re.compile(r'slotId=(?P<sid>\d+)\s+reservationsCount=(?P<rc>\d+)')
_CREATE_REQ_RE   = re.compile(r'\[VISIT_RESERVATION_CREATE_REQ\]\s+userId=(?P<uid>\S+?),\s+slotId=(?P<sid>\d+)')
_CONFIRMED_RE    = re.compile(
    r'\[VISIT_RESERVATION_CONFIRMED\]\s+'
    r'reservationId=(?P<rid>\d+),\s+slotId=(?P<sid>\d+),\s+'
    r'searcher=(?P<searcher>\S+?),\s+registrant=(?P<reg>\S+)'
)


def parse_message_fields(message: str) -> Dict[str, Any]:
    f: Dict[str, Any] = {
        'nano_time':           None,
        'instant':             None,
        'slot_id':             None,
        'window_id':           None,
        'property_id':         None,
        'lease_type':          None,
        'slot_start_time':     None,
        'slot_end_time':       None,
        'slot_status':         None,
        'slot_created_at':     None,
        'reservations_count':  None,
        'reservation_id':      None,
        'searcher_user_id':    None,
        'reservation_status':  None,
        'confirmed_at':        None,
        'cancelled_at':        None,
        'invalidated_at':      None,
        'visit_result':        None,
        'result_classified_at':None,
        'registrant_user_id':  None,
    }

    ni = _NANO_INSTANT_RE.search(message)
    if ni:
        f['nano_time'] = int(ni.group('nano'))
        f['instant']   = ni.group('inst')

    sm = _SLOT_PAYLOAD_RE.search(message)
    if sm:
        kv = parse_kv(sm.group('body'))
        if kv.get('slotId'):   f['slot_id']   = int(kv['slotId'])
        if kv.get('windowId'): f['window_id'] = int(kv['windowId'])
        f['property_id']     = norm(kv.get('propertyId'))
        f['lease_type']      = norm(kv.get('leaseType'))
        f['slot_start_time'] = norm(kv.get('startTime'))
        f['slot_end_time']   = norm(kv.get('endTime'))
        f['slot_status']     = norm(kv.get('status'))
        f['slot_created_at'] = norm(kv.get('createdAt'))

    rm = _RES_PAYLOAD_RE.search(message)
    if rm:
        kv = parse_kv(rm.group('body'))
        if kv.get('reservationId'): f['reservation_id'] = int(kv['reservationId'])
        if kv.get('slotId') and f['slot_id'] is None:
            f['slot_id'] = int(kv['slotId'])
        f['searcher_user_id']     = norm(kv.get('searcherUserId'))
        f['reservation_status']   = norm(kv.get('status'))
        f['confirmed_at']         = norm(kv.get('confirmedAt'))
        f['cancelled_at']         = norm(kv.get('cancelledAt'))
        f['invalidated_at']       = norm(kv.get('invalidatedAt'))
        f['visit_result']         = norm(kv.get('visitResult'))
        f['result_classified_at'] = norm(kv.get('resultClassifiedAt'))

    cm = _RES_COUNT_RE.search(message)
    if cm:
        if f['slot_id'] is None:
            f['slot_id'] = int(cm.group('sid'))
        f['reservations_count'] = int(cm.group('rc'))

    rq = _CREATE_REQ_RE.search(message)
    if rq:
        f['searcher_user_id'] = rq.group('uid')
        if f['slot_id'] is None:
            f['slot_id'] = int(rq.group('sid'))

    cf = _CONFIRMED_RE.search(message)
    if cf:
        if f['reservation_id'] is None:
            f['reservation_id'] = int(cf.group('rid'))
        if f['slot_id'] is None:
            f['slot_id'] = int(cf.group('sid'))
        if f['searcher_user_id'] is None:
            f['searcher_user_id'] = cf.group('searcher')
        f['registrant_user_id'] = cf.group('reg')

    return f


RAW_FIELD_COLUMNS: List[Tuple[str, str]] = [
    ('nano_time',            '측정 시각 (nanoTime, ns)'),
    ('instant',              '절대 시각 (instant, UTC)'),
    ('slot_id',              '슬롯 번호'),
    ('window_id',            '윈도우 번호'),
    ('property_id',          '매물 식별자'),
    ('lease_type',           '임대 유형'),
    ('slot_start_time',      '슬롯 시작 시각'),
    ('slot_end_time',        '슬롯 종료 시각'),
    ('slot_status',          '슬롯 상태'),
    ('slot_created_at',      '슬롯 생성 시각'),
    ('reservations_count',   '슬롯에 묶인 예약 수'),
    ('reservation_id',       '예약 번호'),
    ('searcher_user_id',     '탐색자 ID'),
    ('reservation_status',   '예약 상태'),
    ('confirmed_at',         '예약 확정 시각'),
    ('cancelled_at',         '예약 취소 시각'),
    ('invalidated_at',       '예약 무효화 시각'),
    ('visit_result',         '방문 결과'),
    ('result_classified_at', '결과 분류 시각'),
    ('registrant_user_id',   '등록자 ID'),
]


# ════════════════════════════════════════════════════════════════════
# 3. 회차 결과 컨테이너
# ════════════════════════════════════════════════════════════════════

class RunResult:
    """한 회차의 측정 결과를 담는 컨테이너."""

    def __init__(self, name: str, log_path: Path, meta: Optional[Dict[str, str]] = None):
        self.name = name
        self.log_path = log_path
        self.meta: Dict[str, str] = meta or {}

        # 요청별 요약
        fb_pid, fb_lt = collect_property_fallback(log_path)
        events = parse_log_events(log_path)
        groups = group_requests(events)
        self.rows: List[Dict[str, Any]] = [
            summarize_request(i, g, fb_pid, fb_lt)
            for i, g in enumerate(groups, start=1)
        ]

        # 회차 단위 카운트
        anomalies = count_anomalies(log_path)
        self.db_reject_count: int = anomalies['db_reject']
        self.ora_uq_count: int = anomalies['ora_uq']
        self.ora_other_count: int = anomalies['ora_other']

        # 집계
        self.total: int = len(self.rows)
        self.reached_check: int = sum(
            1 for r in self.rows
            if r['진행 단계'] in ('확인 조회 직후', '점유 직후', '등록 완료')
        )
        self.reached_update: int = sum(
            1 for r in self.rows if r['진행 단계'] in ('점유 직후', '등록 완료')
        )
        self.reached_insert: int = sum(
            1 for r in self.rows if r['진행 단계'] == '등록 완료'
        )
        self.confirmed_rows: int = self.reached_insert
        self.violation_requests: int = sum(
            1 for r in self.rows if r['정합성 판정'].startswith('위반')
        )
        after_counts = [
            r['등록 후 예약 수'] for r in self.rows if r['등록 후 예약 수'] is not None
        ]
        self.max_confirmed_observed: int = max(after_counts) if after_counts else 0

        ns_values = [
            r['확인→등록 응답시간 (ns)']
            for r in self.rows if r['확인→등록 응답시간 (ns)'] is not None
        ]
        self.mean_ns: Optional[int] = round(statistics.fmean(ns_values)) if ns_values else None
        self.min_ns: Optional[int] = min(ns_values) if ns_values else None
        self.max_ns: Optional[int] = max(ns_values) if ns_values else None
        self.stdev_ns: Optional[int] = (
            round(statistics.stdev(ns_values)) if len(ns_values) >= 2 else None
        )

        # 원본 로그 라인
        self.raw_lines: List[Dict[str, Any]] = []
        with log_path.open('r', encoding='utf-8', errors='replace') as f:
            for line in f:
                m = RAW_LOG_RE.match(line.rstrip('\n'))
                if not m:
                    continue
                message = m.group('message')
                row = {
                    'log_ts':  m.group('ts'),
                    'thr':     m.group('thr'),
                    'message': message,
                    'phase':   detect_phase(message),
                }
                row.update(parse_message_fields(message))
                self.raw_lines.append(row)


# ════════════════════════════════════════════════════════════════════
# 4. 엑셀 출력 — 공통 스타일
# ════════════════════════════════════════════════════════════════════

TITLE_FILL   = '305496'
SECTION_FILL = '8EA9DB'
HEADER_FILL  = '203864'
LABEL_FILL   = 'F2F2F2'
DIFF_FILL    = 'FFE699'  # 회차 간 차이 강조
GOOD_FILL    = 'C6EFCE'  # 정합성 충족
BAD_FILL     = 'F4B084'  # 정합성 위반


def fill(cell, color: str):
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')


def write_section(ws: Worksheet, row: int, title: str, span: int):
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=span)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    fill(c, SECTION_FILL)
    c.alignment = Alignment(vertical='center')
    ws.row_dimensions[row].height = 22


def write_title(ws: Worksheet, title_text: str, span: int):
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=span)
    title = ws.cell(row=1, column=1, value=title_text)
    title.font = Font(bold=True, color='FFFFFF', size=14)
    fill(title, TITLE_FILL)
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30


def write_table_header(ws: Worksheet, row: int, headers: List[str]):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        fill(c, HEADER_FILL)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30


def write_label_value(ws: Worksheet, row: int, label: str, *values, span: int = 0):
    a = ws.cell(row=row, column=1, value=label)
    a.font = Font(bold=True)
    fill(a, LABEL_FILL)
    a.alignment = Alignment(vertical='center')
    for offset, v in enumerate(values, start=2):
        ws.cell(row=row, column=offset, value=v).alignment = Alignment(vertical='center', wrap_text=True)
    if span and len(values) == 1:
        ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=span)


def col_letter(idx: int) -> str:
    return get_column_letter(idx)


def fmt_ns_with_ms(v: Optional[int]) -> Optional[str]:
    """나노초 정수를 사람이 읽기 좋게: 12,345,678 ns  (≈ 12.35 ms)"""
    if v is None:
        return None
    return f'{v:,} ns  (≈ {v / 1_000_000:.2f} ms)'


# ════════════════════════════════════════════════════════════════════
# 5. 시트 1 — 회차 비교 요약
# ════════════════════════════════════════════════════════════════════

def write_sheet_comparison(ws: Worksheet, runs: List[RunResult]):
    ws.title = '회차 비교 요약'
    n_runs = len(runs)
    span = 1 + n_runs

    write_title(ws, 'F004 슬롯 예약 동시성 측정 — 회차 비교 요약', span)

    ws.cell(row=3, column=1, value='회차 수').font = Font(bold=True)
    ws.cell(row=3, column=2, value=n_runs)
    ws.cell(row=4, column=1, value='측정 대상 코드').font = Font(bold=True)
    ws.cell(row=4, column=2, value='com.wherehouse.VisitReservation.service.VisitReservationWriteService.createReservation')

    # 회차 정보 / 측정 환경
    cur = 6
    write_section(ws, cur, '■ 회차 정보 / 측정 환경', span); cur += 1
    write_table_header(ws, cur, ['항목'] + [r.name for r in runs]); cur += 1
    write_label_value(ws, cur, '입력 로그 경로', *(str(r.log_path) for r in runs)); cur += 1
    write_label_value(ws, cur, '총 요청 수',     *(r.total for r in runs)); cur += 1

    # 회차별 메타 정보 (--meta 로 전달된 키들의 합집합, 첫 등장 순서 유지)
    all_meta_keys: List[str] = []
    for r in runs:
        for k in r.meta:
            if k not in all_meta_keys:
                all_meta_keys.append(k)
    for key in all_meta_keys:
        write_label_value(ws, cur, key, *(r.meta.get(key, '—') for r in runs))
        cur += 1

    cur += 1

    # 단계별 통과 분포
    write_section(ws, cur, '■ 동시 요청의 단계별 통과 분포', span); cur += 1
    write_table_header(ws, cur, ['측정 지점'] + [r.name for r in runs]); cur += 1
    write_label_value(ws, cur,
        '확인 조회 통과 (슬롯을 "예약 가능" 으로 인지)',
        *(r.reached_check for r in runs)); cur += 1
    write_label_value(ws, cur,
        '점유 통과 (슬롯을 "예약됨" 으로 UPDATE 성공)',
        *(r.reached_update for r in runs)); cur += 1
    _row_with_diff(ws, cur,
        '등록 통과 (예약 INSERT 성공)',
        [r.reached_insert for r in runs]); cur += 1
    cur += 1

    # 정합성
    write_section(ws, cur, '■ 정합성 (비즈니스 규칙 1 — 슬롯당 확정 예약 = 1)', span); cur += 1
    write_table_header(ws, cur, ['항목'] + [r.name for r in runs]); cur += 1
    _row_with_judge(ws, cur, 'CONFIRMED 행 수 (DB 에 남은 확정 예약)',
                    [r.confirmed_rows for r in runs],
                    judge=lambda v: 'good' if v == 1 else ('bad' if v >= 2 else None)); cur += 1
    _row_with_judge(ws, cur, '슬롯당 동시 관측된 최대 CONFIRMED 수',
                    [r.max_confirmed_observed for r in runs],
                    judge=lambda v: 'good' if v <= 1 else 'bad'); cur += 1
    _row_with_judge(ws, cur, '정합성 위반 요청 수 (등록 후 예약 수 ≥ 2 로 관측)',
                    [r.violation_requests for r in runs],
                    judge=lambda v: 'good' if v == 0 else 'bad'); cur += 1
    cur += 1

    # DB 단 거부 발생
    write_section(ws, cur, '■ DB 단 거부 발생 (부분 유일 인덱스 UQ_VISIT_RESERVATION_CONFIRMED_SLOT)', span); cur += 1
    write_table_header(ws, cur, ['이벤트'] + [r.name for r in runs]); cur += 1
    write_label_value(ws, cur,
        'DB 무결성 위반 처리 진입 수 ([VISIT_INTEGRITY_BACKSTOP] 로그)',
        *(r.db_reject_count for r in runs)); cur += 1
    write_label_value(ws, cur,
        'ORA-00001 발생 수 (UQ_VISIT_RESERVATION_CONFIRMED_SLOT 등)',
        *(r.ora_uq_count for r in runs)); cur += 1
    write_label_value(ws, cur,
        '그 외 ORA-xxxxx 발생 수',
        *(r.ora_other_count for r in runs)); cur += 1
    cur += 1

    # 응답시간 분포
    write_section(ws, cur, '■ 확인→등록 응답시간 (확인 조회 직후 → 등록 직후 구간, 등록까지 도달한 요청 한정)', span); cur += 1
    write_table_header(ws, cur, ['통계 항목'] + [r.name for r in runs]); cur += 1
    write_label_value(ws, cur, '평균',     *(fmt_ns_with_ms(r.mean_ns) for r in runs)); cur += 1
    write_label_value(ws, cur, '최솟값',   *(fmt_ns_with_ms(r.min_ns) for r in runs)); cur += 1
    write_label_value(ws, cur, '최댓값',   *(fmt_ns_with_ms(r.max_ns) for r in runs)); cur += 1
    write_label_value(ws, cur, '표준편차', *(fmt_ns_with_ms(r.stdev_ns) for r in runs)); cur += 1
    cur += 2

    # 해석 가이드
    write_section(ws, cur, '■ 본 시트 해석 가이드', span); cur += 1
    guides = [
        ('확인 조회 통과 = 총 요청 수',
         '같은 슬롯에 동시 요청한 N 건 모두 슬롯을 "예약 가능" 으로 인지한 것이다. 즉 확인 조회와 상태 변경 사이의 시간 간격에서 동시 N 건이 모두 같은 옛 값을 본다.'),
        ('점유 통과 = 총 요청 수',
         'N 건 모두 슬롯 상태를 "예약됨" 으로 UPDATE 성공한 것이다. Oracle 의 자동 행 락은 N 개 UPDATE 를 순차로 실행시킬 뿐 두 번째 UPDATE 를 거부하지는 않는다.'),
        ('등록 통과 = 1 vs N',
         '예약 INSERT 단계의 통과 인원. 1 이라면 DB 단의 부분 유일 인덱스가 후속 트랜잭션을 거부한 것이고, N 이라면 추가 메커니즘 없이 모두 INSERT 됨 — 한 슬롯에 N 명 확정 (정합성 위반) 의 직접 관측.'),
        ('CONFIRMED 행 수 / 슬롯당 최대 관측',
         '"CONFIRMED 행 수" 는 측정 종료 후 DB 잔여. "슬롯당 최대 관측" 은 측정 중 어느 한 트랜잭션이 등록 직후 시점에 관측한 동일 슬롯의 누적 예약 수 최댓값. 둘 다 1 이어야 비즈니스 규칙 1 충족.'),
        ('DB 무결성 위반 처리 진입 = N-1',
         '예약 INSERT 단계에서 부분 유일 인덱스가 N-1 건을 ORA-00001 로 거부했다는 의미. 이 값이 N-1 이면 DB 단 보호망 활성, 0 이면 비활성 환경.'),
        ('응답시간이 회차마다 다른 의미',
         'DB 단 보호망이 활성인 회차에서는 거부된 트랜잭션의 ORA-00001 처리 비용이 일부 포함되어 분산이 커진다. 비활성 회차에서는 모두 같은 코드 경로를 통과해 분산이 작다.'),
    ]
    for k, v in guides:
        write_label_value(ws, cur, k, v, span=span)
        ws.merge_cells(start_row=cur, end_row=cur, start_column=2, end_column=span)
        ws.row_dimensions[cur].height = max(30, 15 * (1 + len(v) // 80))
        cur += 1

    ws.column_dimensions['A'].width = 60
    for i in range(n_runs):
        ws.column_dimensions[col_letter(2 + i)].width = 28
    ws.freeze_panes = 'B2'


def _row_with_diff(ws: Worksheet, row: int, label: str, values: List[Any]):
    write_label_value(ws, row, label, *values)
    if len(set(values)) > 1:
        for i in range(len(values)):
            fill(ws.cell(row=row, column=2 + i), DIFF_FILL)


def _row_with_judge(ws: Worksheet, row: int, label: str,
                    values: List[Any], judge):
    write_label_value(ws, row, label, *values)
    for i, v in enumerate(values):
        verdict = judge(v) if v is not None else None
        if verdict == 'good':
            fill(ws.cell(row=row, column=2 + i), GOOD_FILL)
        elif verdict == 'bad':
            fill(ws.cell(row=row, column=2 + i), BAD_FILL)


# ════════════════════════════════════════════════════════════════════
# 6. 시트 2 — 회차별 측정 결과 (회차별 섹션 분리)
# ════════════════════════════════════════════════════════════════════

PER_RUN_HEADERS = [
    '요청 번호', '처리 스레드명', '매물 식별자', '임대 유형',
    '슬롯 번호', '방문 예정 시간', '요청 시작 시각', '요청 종료 시각',
    '확인→등록 응답시간 (ns)', '진행 단계', '확인 조회 시점 슬롯 상태',
    '점유 직후 슬롯 상태', '등록 전 예약 수', '등록 후 예약 수', '정합성 판정',
]


def write_sheet_per_run(ws: Worksheet, runs: List[RunResult]):
    ws.title = '회차별 측정 결과'
    span = len(PER_RUN_HEADERS)

    write_title(ws, '회차별 측정 결과 — 한 요청 = 한 행', span)

    ws.cell(row=3, column=1, value='총 요청 수 (모든 회차 합산)').font = Font(bold=True)
    ws.cell(row=3, column=2, value=sum(r.total for r in runs))
    ws.cell(row=4, column=1, value='회차 수').font = Font(bold=True)
    ws.cell(row=4, column=2, value=len(runs))

    cur = 6
    for run in runs:
        # 회차 섹션 헤더
        write_section(ws, cur,
                      f'■ {run.name}  —  요청 {run.total}건 / 등록 완료 {run.reached_insert}건 / '
                      f'정합성 위반 {run.violation_requests}건 / DB 단 거부 {run.db_reject_count}건',
                      span)
        cur += 1

        # 표 헤더
        write_table_header(ws, cur, PER_RUN_HEADERS)
        cur += 1

        # 데이터 행
        if not run.rows:
            ws.cell(row=cur, column=1, value='(측정된 요청이 없습니다.)')
            cur += 1
        else:
            for r in run.rows:
                for col_idx, h in enumerate(PER_RUN_HEADERS, start=1):
                    ws.cell(row=cur, column=col_idx, value=r.get(h))
                # 정합성 판정 셀 강조
                judge = r['정합성 판정']
                judge_cell = ws.cell(row=cur, column=span)
                if judge == '정상':
                    fill(judge_cell, GOOD_FILL)
                elif judge.startswith('위반'):
                    fill(judge_cell, BAD_FILL)
                cur += 1

        cur += 2  # 회차 사이 빈 줄 2 개

    # 컬럼 안내
    write_section(ws, cur, '■ 컬럼 안내', span); cur += 1
    notices = [
        ('진행 단계', '본 요청이 마지막으로 통과한 측정 지점. "확인 조회 직후" / "점유 직후" / "등록 완료" 중 하나.'),
        ('확인→등록 응답시간 (ns)', '확인 조회 직후부터 등록 직후까지 System.nanoTime() 차이. 등록 단계까지 도달한 요청만 측정값을 갖는다.'),
        ('확인 조회 시점 슬롯 상태', '확인 조회 직후 시점에 본 요청이 인지한 슬롯 상태. 정상이면 "예약 가능".'),
        ('점유 직후 슬롯 상태', '본 요청이 슬롯 상태를 "예약됨" 으로 UPDATE 한 직후의 상태. 정상이면 "예약됨".'),
        ('등록 전 / 등록 후 예약 수', '확인 조회 시점과 등록 직후 시점에 본 트랜잭션이 본 동일 슬롯의 예약 수. "등록 후" 가 1 이어야 정상.'),
        ('정합성 판정', '"등록 후 예약 수" 가 1 = 정상 (초록), 2 이상 = 위반 (주황). "등록 단계 미도달" = 예약 INSERT 까지 도달하지 못한 요청 (이전 단계에서 throw).'),
    ]
    for k, v in notices:
        a = ws.cell(row=cur, column=1, value=k); a.font = Font(bold=True); fill(a, LABEL_FILL)
        a.alignment = Alignment(vertical='top')
        b = ws.cell(row=cur, column=2, value=v); b.alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=cur, end_row=cur, start_column=2, end_column=span)
        cur += 1

    widths = [8, 22, 36, 10, 8, 14, 22, 22, 24, 14, 22, 18, 14, 14, 18]
    for idx, w in enumerate(widths[:span], start=1):
        ws.column_dimensions[col_letter(idx)].width = w
    ws.freeze_panes = 'A2'


# ════════════════════════════════════════════════════════════════════
# 7. 시트 3 — 원본 측정 로그
# ════════════════════════════════════════════════════════════════════

def write_sheet_raw_logs(ws: Worksheet, runs: List[RunResult]):
    ws.title = '원본 측정 로그'

    raw_keys    = [k for k, _ in RAW_FIELD_COLUMNS]
    raw_headers = [h for _, h in RAW_FIELD_COLUMNS]
    headers = ['회차', '순번', '로그 기록 시각', '처리 스레드', '단계 / 이벤트'] + raw_headers + ['log.info 메시지 (원본 그대로)']
    span = len(headers)

    write_title(ws, '원본 측정 로그 — 회차별 시간순', span)

    cur = 3
    for run in runs:
        ws.cell(row=cur, column=1, value=f'  · {run.name}').font = Font(bold=True)
        ws.cell(row=cur, column=2, value=f'{len(run.raw_lines)} 라인').font = Font(italic=True)
        ws.cell(row=cur, column=3, value=str(run.log_path))
        cur += 1
    cur += 1

    write_section(ws, cur, '■ 측정 로그 라인 (회차별로 묶고 회차 내부는 시간 순)', span); cur += 1
    write_table_header(ws, cur, headers); cur += 1

    for run in runs:
        if not run.raw_lines:
            ws.cell(row=cur, column=1, value=run.name)
            ws.cell(row=cur, column=2, value='(수집된 로그 라인 없음)')
            cur += 1
            continue
        for i, r in enumerate(run.raw_lines, start=1):
            ws.cell(row=cur, column=1, value=run.name).font = Font(bold=True)
            ws.cell(row=cur, column=2, value=i)
            ws.cell(row=cur, column=3, value=r['log_ts'])
            ws.cell(row=cur, column=4, value=r['thr'])
            ws.cell(row=cur, column=5, value=r['phase'])
            base = 6
            for off, key in enumerate(raw_keys):
                ws.cell(row=cur, column=base + off, value=r.get(key))
            msg = ws.cell(row=cur, column=base + len(raw_keys), value=r['message'])
            msg.alignment = Alignment(vertical='top', wrap_text=True)
            cur += 1

    fixed = [16, 8, 22, 24, 24]
    for i, w in enumerate(fixed, start=1):
        ws.column_dimensions[col_letter(i)].width = w
    base = len(fixed) + 1
    raw_widths = [22, 24, 12, 12, 36, 10, 22, 22, 14, 22, 18, 14, 18, 14, 22, 22, 22, 14, 22, 18]
    for off, w in enumerate(raw_widths[:len(raw_keys)]):
        ws.column_dimensions[col_letter(base + off)].width = w
    ws.column_dimensions[col_letter(base + len(raw_keys))].width = 80
    ws.freeze_panes = 'A2'


# ════════════════════════════════════════════════════════════════════
# 8. 시트 4 — 시나리오 예상 동작 ↔ 실측 대조표
# ════════════════════════════════════════════════════════════════════

def write_sheet_scenario_check(ws: Worksheet, runs: List[RunResult]):
    ws.title = '예상 동작 대조'
    n = len(runs)
    span = 2 + n  # 항목 + 예상 동작 + N 회차

    write_title(ws, '시나리오 1 예상 동작 ↔ 실측 대조', span)

    ws.cell(row=3, column=1, value='참조 문서').font = Font(bold=True)
    ws.cell(row=3, column=2, value='1. Wherehouse_F004_시나리오1.md  (섹션 2.3, 3.1)')

    cur = 5
    write_section(ws, cur,
                  '■ 시나리오 1 이 측정 전에 적은 예상 동작과 실측의 회차별 일치 여부', span); cur += 1
    write_table_header(ws, cur, ['항목', '예상 동작'] + [r.name for r in runs]); cur += 1

    n_per_run         = [r.total for r in runs]
    check_per_run     = [r.reached_check for r in runs]
    update_per_run    = [r.reached_update for r in runs]
    insert_per_run    = [r.reached_insert for r in runs]
    reject_per_run    = [r.db_reject_count for r in runs]
    confirmed_per_run = [r.confirmed_rows for r in runs]

    items = [
        (
            '확인 조회 race',
            '동시 N 건이 모두 슬롯을 "예약 가능" 으로 인지 (확인 조회와 상태 변경 사이 간격에서의 race)',
            [f'{ck}/{n_per_run[i]}  → ' + ('✓ 예상대로' if ck == n_per_run[i] else '✗ 불일치')
             for i, ck in enumerate(check_per_run)],
        ),
        (
            '슬롯 UPDATE 순차 실행',
            'Oracle 자동 행 락은 N 개 UPDATE 를 순차로 실행시킬 뿐 거부하지 않는다 → N 모두 점유 통과',
            [f'{up}/{n_per_run[i]}  → ' + ('✓ 예상대로' if up == n_per_run[i] else '✗ 불일치')
             for i, up in enumerate(update_per_run)],
        ),
        (
            'DB 단 보호망 부재 시 다수 INSERT',
            '별도 DB 메커니즘 없으면 예약 INSERT 도 모두 통과되어 한 슬롯에 N 개 CONFIRMED 가 남는다',
            [_judge_no_protection(insert_per_run[i], reject_per_run[i], n_per_run[i])
             for i in range(n)],
        ),
        (
            'DB 단 보호망 활성 시 후속 거부',
            '부분 유일 인덱스 적용 시 후속 트랜잭션의 INSERT 가 ORA-00001 로 거부되고 ROLLBACK',
            [_judge_with_protection(reject_per_run[i], n_per_run[i])
             for i in range(n)],
        ),
        (
            '비즈니스 규칙 1',
            '슬롯당 확정 예약 = 1 (어느 시점에나 최대 1건)',
            [_judge_rule1(confirmed_per_run[i]) for i in range(n)],
        ),
    ]

    for label, expectation, results in items:
        ws.cell(row=cur, column=1, value=label).font = Font(bold=True)
        ws.cell(row=cur, column=1).alignment = Alignment(vertical='top')
        ws.cell(row=cur, column=2, value=expectation).alignment = Alignment(vertical='top', wrap_text=True)
        for i, txt in enumerate(results, start=3):
            c = ws.cell(row=cur, column=i, value=txt)
            c.alignment = Alignment(vertical='top', wrap_text=True)
            if '✓' in txt:
                fill(c, GOOD_FILL)
            elif '✗' in txt:
                fill(c, BAD_FILL)
        ws.row_dimensions[cur].height = 40
        cur += 1

    cur += 1
    write_section(ws, cur, '■ 표 해석', span); cur += 1
    notes = [
        ('이 표의 의미',
         '시나리오 1 문서는 측정 이전에 단계별 메커니즘을 예상으로 적어 두었다. 본 표는 그 예상이 실제 측정으로 어떻게 확인되거나 어긋나는지를 회차별로 한 줄씩 매핑한다.'),
        ('확인 조회 race / 슬롯 UPDATE 순차 실행',
         '이 두 항목은 동시성 제어 기법과 무관하게 항상 N/N 이 나와야 한다. 어긋난다면 측정 도구 또는 코드 변경 의심.'),
        ('DB 단 보호망 부재 시 다수 INSERT',
         '등록 통과가 1 (보호망 활성) 이면 DB 단 보호망이 작동한 것, N (보호망 비활성) 이면 시나리오 문서가 경고한 정합성 위반이 직접 관측된 것이다.'),
        ('DB 단 보호망 활성 시 후속 거부',
         'DB 무결성 위반 처리 진입 수가 N-1 이면 부분 유일 인덱스가 후속 N-1 건을 정확히 거부한 것. 0 이면 보호망이 없는 환경.'),
    ]
    for k, v in notes:
        a = ws.cell(row=cur, column=1, value=k); a.font = Font(bold=True); fill(a, LABEL_FILL)
        a.alignment = Alignment(vertical='top')
        b = ws.cell(row=cur, column=2, value=v); b.alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=cur, end_row=cur, start_column=2, end_column=span)
        ws.row_dimensions[cur].height = max(30, 15 * (1 + len(v) // 80))
        cur += 1

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 56
    for i in range(n):
        ws.column_dimensions[col_letter(3 + i)].width = 30
    ws.freeze_panes = 'C2'


def _judge_no_protection(insert_pass: int, reject: int, total: int) -> str:
    if insert_pass == 1 and reject >= total - 1:
        return f'{insert_pass}/{total}  → ✓ DB 단 보호망 활성 (예상 미적용 환경)'
    if insert_pass == total and reject == 0:
        return f'{insert_pass}/{total}  → ✓ 예상대로 (보호망 없음, 다수 INSERT 직접 관측)'
    return f'{insert_pass}/{total} (DB 거부 {reject})  → ? 추가 분석 필요'


def _judge_with_protection(reject: int, total: int) -> str:
    if reject == 0:
        return f'DB 거부 {reject}건  → (해당 없음 — 인덱스 미적용)'
    if reject == total - 1:
        return f'DB 거부 {reject}건  → ✓ 예상대로 ({total - 1}건 거부)'
    return f'DB 거부 {reject}건  → ? 예상과 다름 (예상 {total - 1}건)'


def _judge_rule1(confirmed: int) -> str:
    if confirmed == 1:
        return f'CONFIRMED {confirmed}건  → ✓ 충족'
    if confirmed == 0:
        return f'CONFIRMED {confirmed}건  → ? 측정 표본 부족'
    return f'CONFIRMED {confirmed}건  → ✗ 위반 (한 슬롯에 {confirmed}건)'


# ════════════════════════════════════════════════════════════════════
# 9. 엑셀 출력 — 시트 합치기
# ════════════════════════════════════════════════════════════════════

def write_workbook(runs: List[RunResult], out_path: Path):
    wb = Workbook()
    write_sheet_comparison(wb.active, runs)
    write_sheet_per_run(wb.create_sheet(), runs)
    write_sheet_raw_logs(wb.create_sheet(), runs)
    write_sheet_scenario_check(wb.create_sheet(), runs)
    wb.save(out_path)


# ════════════════════════════════════════════════════════════════════
# 10. CLI
# ════════════════════════════════════════════════════════════════════

def parse_run_arg(s: str) -> Tuple[str, Path]:
    if '=' not in s:
        raise argparse.ArgumentTypeError(
            f'--log 인자는 NAME=PATH 형식이어야 합니다 (현재: {s!r})'
        )
    name, path_str = s.split('=', 1)
    name = name.strip()
    path_obj = Path(path_str.strip())
    if not name:
        raise argparse.ArgumentTypeError(f'--log 인자의 NAME 이 비어있습니다: {s!r}')
    if len(name) > 31:
        raise argparse.ArgumentTypeError(
            f'--log NAME 은 31자 이하여야 합니다 (Excel 시트 이름 제약). 현재 {len(name)}자: {name!r}'
        )
    if not path_obj.exists():
        raise argparse.ArgumentTypeError(f'--log 경로의 파일을 찾을 수 없습니다: {path_obj}')
    return name, path_obj


def parse_meta_arg(s: str) -> Tuple[str, Dict[str, str]]:
    """--meta 인자 파싱: NAME=KEY1:VALUE1;KEY2:VALUE2;... → (name, {KEY1:VALUE1, ...})"""
    if '=' not in s:
        raise argparse.ArgumentTypeError(
            f'--meta 인자는 NAME=KEY:VALUE;KEY:VALUE;... 형식이어야 합니다: {s!r}'
        )
    name, body = s.split('=', 1)
    name = name.strip()
    if not name:
        raise argparse.ArgumentTypeError(f'--meta 의 NAME 이 비어있습니다: {s!r}')
    meta: Dict[str, str] = {}
    for item in body.split(';'):
        item = item.strip()
        if not item:
            continue
        if ':' not in item:
            raise argparse.ArgumentTypeError(
                f'--meta 의 각 항목은 KEY:VALUE 형식이어야 합니다 '
                f'(잘못된 항목: {item!r}, 전체: {s!r})'
            )
        k, v = item.split(':', 1)
        meta[k.strip()] = v.strip()
    return name, meta


def main():
    parser = argparse.ArgumentParser(
        description='F004 동시성 측정 다회차 비교 보고서 생성기',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            '예:\n'
            '  py -3.13 report.py \\\n'
            '    --log "1_unique적용=결과/Step1_unique적용_T1/wherehouse.log" \\\n'
            '    --log "2_unique제거=결과/Step2_unique삭제_T1/wherehouse.log" \\\n'
            '    --meta "1_unique적용=격리:READ COMMITTED;UQ제약:적용;동시성제어:DB 부분 유일 인덱스" \\\n'
            '    --meta "2_unique제거=격리:READ COMMITTED;UQ제약:제거;동시성제어:없음 (baseline)" \\\n'
            '    --out 결과/F004_측정결과.xlsx\n'
        ),
    )
    parser.add_argument(
        '--log', type=parse_run_arg, action='append', required=True, metavar='NAME=PATH',
        help='회차 이름과 wherehouse.log 경로 (여러 번 지정 가능, 회차 이름이 표 컬럼명으로 사용됨)',
    )
    parser.add_argument(
        '--meta', type=parse_meta_arg, action='append', default=None,
        metavar='NAME=KEY:VALUE;KEY:VALUE;...',
        help='(선택) 회차의 측정 환경 메타 정보. NAME 은 --log 의 NAME 과 매칭. '
             '예: --meta "1_unique적용=격리:READ COMMITTED;UQ제약:적용;동시성제어:DB 부분 유일 인덱스"',
    )
    parser.add_argument(
        '--out', type=Path, required=True, metavar='PATH',
        help='출력 xlsx 경로',
    )
    args = parser.parse_args()

    args.out.parent.mkdir(parents=True, exist_ok=True)

    names = [n for n, _ in args.log]
    if len(set(names)) != len(names):
        sys.exit(f'[ERROR] --log NAME 이 중복됩니다: {names}')

    # --meta 매핑 (선택). 매칭 안 되는 NAME 은 경고
    meta_map: Dict[str, Dict[str, str]] = {}
    if args.meta:
        log_names = set(names)
        for meta_name, meta_dict in args.meta:
            if meta_name not in log_names:
                print(f'[WARN] --meta NAME 이 --log NAME 과 매칭되지 않음: {meta_name!r}',
                      file=sys.stderr)
            meta_map[meta_name] = meta_dict

    runs = [RunResult(name, path, meta_map.get(name, {})) for name, path in args.log]
    write_workbook(runs, args.out)

    print(f'[OK] {len(runs)} 회차 정리 -> {args.out}')
    for r in runs:
        print(
            f'  · {r.name}: 요청 {r.total} / 등록 완료 {r.reached_insert} / '
            f'위반 {r.violation_requests} / DB 단 거부 {r.db_reject_count} / '
            f'CONFIRMED 잔여 {r.confirmed_rows}'
        )


if __name__ == '__main__':
    main()
